"""
百度热搜榜单 全量采集 → Excel
==========================================================
核心接口（热搜板块）：
  GET https://top.baidu.com/api/board
      ?platform=wise&tab={榜单类型}        # 全国各类榜单（公开，无需登录）
      ?platform=pc&tab=city&city={城市名}  # 城市榜（需中国大陆IP）

可用 tab：realtime 热搜榜 / livelihood 民生榜 / finance 财经榜 /
         sports 体育榜 / new_entertainment 文娱榜 / internation_news 国际榜 /
         challenge 挑战榜 / movie 电影榜 / teleplay 电视剧榜 /
         novel 小说榜 / drama 短剧榜

百度新闻搜索（获取对应内容）：
  GET https://news.baidu.com/ns
      ?word={关键词}&tn=news&from=news&cl=2&pn=0&rn=1&ct=1&ie=utf-8

数据字段：
  word=话题标题, index=排名, newHotName=热度标签("新"/"热"),
  labelTagName=内容标签("热议"/"辟谣"), url=百度搜索链接
"""

import urllib.parse
import urllib.request
import json
import os
import re
import time
import uuid
from datetime import datetime

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 榜单配置 ─────────────────────────────────────────────────────────────────

BOARD_TABS = {
    "realtime":          "热搜榜",
    "livelihood":        "民生榜",
    "finance":           "财经榜",
    "sports":            "体育榜",
    "new_entertainment": "文娱榜",
    "internation_news":  "国际榜",
    "challenge":         "挑战榜",
    "movie":             "电影榜",
    "teleplay":          "电视剧榜",
    "novel":             "小说榜",
    "drama":             "短剧榜",
}

# 城市榜使用城市中文名（需中国大陆IP）
CITIES_ALL = [
    "北京", "上海", "广州", "深圳", "成都",
    "杭州", "重庆", "武汉", "苏州", "西安",
    "南京", "长沙", "郑州", "天津", "合肥",
    "青岛", "东莞", "宁波", "佛山", "沈阳",
    "哈尔滨", "长春", "昆明", "福州", "厦门",
    "南昌", "济南", "石家庄", "太原", "贵阳",
]

# 默认采集的榜单
TARGET_BOARDS = ["realtime", "livelihood", "finance", "sports", "new_entertainment"]

# 默认城市（不传 --city 时仅抓榜单，传 --city-all 才抓全部）
TARGET_CITIES: list[str] = []

# 每个榜单最多采集条数（0 = 全部）
TOP_N = 0

BOARD_API = "https://top.baidu.com/api/board"
NEWS_API  = "https://news.baidu.com/ns"

UA_DESKTOP = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/132.0.0.0 Safari/537.36"
)
UA_MOBILE = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Version/17.0 Mobile/15E148 Safari/604.1"
)

COLUMNS = [
    "平台", "网站域名", "内容分类", "榜单",
    "标题", "排名", "热度", "标签",
    "对应内容url", "对应内容",
    "发布时间", "采集时间", "批次id", "原始来源",
    "外网图片链接", "内网图片链接", "图片分辨率",
]

# ─── HTTP Session ─────────────────────────────────────────────────────────────

_SESSION: requests.Session | None = None


def _make_session() -> requests.Session:
    s     = requests.Session()
    retry = Retry(total=2, backoff_factor=0.3,
                  status_forcelist=[500, 502, 503, 504])
    adp   = HTTPAdapter(max_retries=retry,
                        pool_connections=4, pool_maxsize=8)
    s.mount("https://", adp)
    s.mount("http://",  adp)
    return s


def _session() -> requests.Session:
    global _SESSION
    if _SESSION is None:
        _SESSION = _make_session()
    return _SESSION


# ─── 工具 ─────────────────────────────────────────────────────────────────────

def _strip_html(t: str) -> str:
    return re.sub(r"<[^>]+>", "", t or "").strip()


# ─── 接口一：榜单热搜列表 ──────────────────────────────────────────────────────

def _board_request(tab: str, city: str = "", platform: str = "wise") -> dict:
    params: dict = {"platform": platform, "tab": tab}
    if city:
        params["city"] = city
    url = BOARD_API + "?" + urllib.parse.urlencode(params, encoding="utf-8")
    headers = {
        "accept":          "application/json, text/plain, */*",
        "accept-language": "zh-CN,zh;q=0.9",
        "cache-control":   "no-cache",
        "referer":         "https://top.baidu.com/board",
        "user-agent":      UA_DESKTOP,
    }
    resp = _session().get(url, headers=headers, timeout=15)
    data = resp.json()
    if not data.get("success"):
        err = data.get("error", {}).get("message", "未知错误")
        raise RuntimeError(err)
    return data


def _parse_items(data: dict) -> list:
    for card in data.get("data", {}).get("cards", []):
        if card.get("component") == "tabTextList":
            for block in card.get("content", []):
                items = block.get("content", [])
                if items:
                    return items
    return []


def get_board_hot_list(tab: str, board_name: str, top_n: int = 0) -> list:
    """获取指定榜单热搜列表。top_n=0 表示不限。"""
    data  = _board_request(tab)
    items = _parse_items(data)
    result = []
    rank   = 0
    for item in items:
        if item.get("isTop"):
            result.append({
                "rank":      0,
                "word":      item.get("word", ""),
                "hot_tag":   item.get("newHotName", ""),
                "label":     item.get("labelTagName", ""),
                "url":       item.get("url", ""),
                "board":     board_name,
                "city":      "",
            })
        else:
            rank += 1
            result.append({
                "rank":      item.get("index", rank),
                "word":      item.get("word", ""),
                "hot_tag":   item.get("newHotName", ""),
                "label":     item.get("labelTagName", ""),
                "url":       item.get("url", ""),
                "board":     board_name,
                "city":      "",
            })
        if top_n > 0 and len(result) >= top_n:
            break
    return result


def get_city_hot_list(city: str, top_n: int = 0) -> list:
    """获取城市热榜（需中国大陆IP）。"""
    data  = _board_request(tab="city", city=city, platform="pc")
    items = _parse_items(data)
    result = []
    rank   = 0
    for item in items:
        if item.get("isTop"):
            result.append({
                "rank":      0,
                "word":      item.get("word", ""),
                "hot_tag":   item.get("newHotName", ""),
                "label":     item.get("labelTagName", ""),
                "url":       item.get("url", ""),
                "board":     "城市榜",
                "city":      city,
            })
        else:
            rank += 1
            result.append({
                "rank":      item.get("index", rank),
                "word":      item.get("word", ""),
                "hot_tag":   item.get("newHotName", ""),
                "label":     item.get("labelTagName", ""),
                "url":       item.get("url", ""),
                "board":     "城市榜",
                "city":      city,
            })
        if top_n > 0 and len(result) >= top_n:
            break
    return result


# ─── 接口二：百度新闻搜索（获取对应内容）────────────────────────────────────────

def fetch_top_news(keyword: str) -> dict:
    """
    搜索百度新闻，返回第一条结果，字段：
      title, url, snippet, source, pub_time, img_url
    任意步骤失败返回空字典。
    """
    params = urllib.parse.urlencode({
        "word": keyword,
        "tn":   "news",
        "from": "news",
        "cl":   "2",
        "pn":   "0",
        "rn":   "5",
        "ct":   "1",
        "ie":   "utf-8",
    }, encoding="utf-8")
    url = NEWS_API + "?" + params
    headers = {
        "User-Agent": UA_MOBILE,
        "Accept":     "text/html,application/xhtml+xml,*/*",
        "Referer":    "https://news.baidu.com/",
        "Accept-Language": "zh-CN,zh;q=0.9",
    }
    try:
        resp = _session().get(url, headers=headers, timeout=8)
        html = resp.text
    except Exception:
        return {}

    # 提取第一条新闻标题
    m_title = re.search(
        r'class="result[^"]*"[^>]*>.*?<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>',
        html, re.S)
    if not m_title:
        # 备用：匹配常见的 news-item 结构
        m_title = re.search(
            r'<a[^>]+class="[^"]*title[^"]*"[^>]+href="([^"]+)"[^>]*>(.*?)</a>',
            html, re.S)
    if not m_title:
        return {}

    art_url   = m_title.group(1).strip()
    art_title = _strip_html(m_title.group(2)).strip()

    # 提取正文摘要
    m_snip = re.search(
        r'<p[^>]*class="[^"]*abs[^"]*"[^>]*>(.*?)</p>',
        html, re.S)
    snippet = _strip_html(m_snip.group(1)).strip() if m_snip else ""

    # 提取来源
    m_src = re.search(
        r'class="[^"]*c-color-gray[^"]*"[^>]*>(.*?)</span>',
        html, re.S)
    source = _strip_html(m_src.group(1)).strip() if m_src else ""

    # 提取发布时间
    m_time = re.search(
        r'(\d{4}-\d{2}-\d{2}[\s\d:]*|\d+[分小时天]+前)',
        html)
    pub_time = m_time.group(1).strip() if m_time else ""

    # 提取图片
    m_img = re.search(
        r'src="(https?://[^"]+\.(?:jpg|jpeg|png|webp)[^"]*)"',
        html, re.I)
    img_url = m_img.group(1) if m_img else ""

    return {
        "title":    art_title,
        "url":      art_url,
        "snippet":  snippet,
        "source":   source,
        "pub_time": pub_time,
        "img_url":  img_url,
    }


# ─── 图片下载 ─────────────────────────────────────────────────────────────────

_IMG_HEADERS = {
    "User-Agent": UA_DESKTOP,
    "Referer":    "https://news.baidu.com/",
    "Accept":     "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
}


def download_image(url: str, save_dir: str, filename: str) -> str:
    """下载图片到本地，返回绝对路径；失败返回空字符串。"""
    if not url:
        return ""
    os.makedirs(save_dir, exist_ok=True)
    ext = os.path.splitext(url.split("?")[0])[-1]
    if not ext.startswith(".") or len(ext) > 5:
        ext = ".jpg"
    local_path = os.path.join(save_dir, filename + ext)
    if os.path.exists(local_path):
        return local_path
    try:
        resp = _session().get(url, headers=_IMG_HEADERS, timeout=8)
        if resp.status_code == 200:
            with open(local_path, "wb") as f:
                f.write(resp.content)
            return local_path
    except Exception:
        pass
    return ""


# ─── 构建记录 ─────────────────────────────────────────────────────────────────

def build_record(hot_item: dict, news: dict,
                 collect_time: str, batch_id: str,
                 img_dir: str = "") -> dict:
    word  = hot_item["word"]
    rank  = hot_item["rank"]
    board = hot_item.get("board", "")
    city  = hot_item.get("city", "")

    # 内容分类：城市榜 / 热搜榜单
    category = "城市热搜" if city else "热搜榜单"

    # 对应内容：新闻摘要
    content    = news.get("snippet", "") if news else ""
    content_url = news.get("url", "") or hot_item.get("url", "")

    # 发布时间
    pub_time = news.get("pub_time", "") if news else ""

    # 原始来源：新闻媒体来源
    source = news.get("source", "")
    if not source:
        source = f"百度{board or '热搜'}"

    # 图片
    ext_img = news.get("img_url", "") if news else ""

    # 标签：热度标签 + 内容标签合并
    tag = hot_item.get("label", "") or hot_item.get("hot_tag", "")

    # 下载图片
    local_img  = ""
    resolution = ""
    if ext_img and img_dir:
        safe_board = re.sub(r"[^\w一-鿿]", "_", board)
        safe_city  = city or safe_board
        fname      = f"{safe_city}_{rank:02d}_{uuid.uuid4().hex[:8]}"
        local_img  = download_image(ext_img, img_dir, fname)

    # 图片分辨率
    if not resolution and local_img:
        try:
            from PIL import Image as _PILImage
            with _PILImage.open(local_img) as _img:
                resolution = f"{_img.width}x{_img.height}"
        except Exception:
            pass

    return {
        "平台":        "百度",
        "网站域名":    "baidu.com",
        "内容分类":    category,
        "榜单":        board,
        "标题":        word,
        "排名":        "置顶" if rank == 0 else rank,
        "热度":        hot_item.get("hot_tag", ""),
        "标签":        tag,
        "对应内容url":  content_url,
        "对应内容":    content,
        "发布时间":    pub_time,
        "采集时间":    collect_time,
        "批次id":      batch_id,
        "原始来源":    source,
        "外网图片链接": ext_img,
        "内网图片链接": local_img,
        "图片分辨率":   resolution,
    }


# ─── Excel 写出 ────────────────────────────────────────────────────────────────

_BOARD_COLORS = [
    "DDEEFF", "EEF5FF", "DDF0E8", "EEF8F3",
    "FFF3DD", "FFF9EE", "F0DDEE", "F8EEF5",
    "DDEEEE", "EEF5F5", "F5DDEE", "F0EEF5",
]
_HEADER_FILL  = PatternFill("solid", fgColor="2B5FAA")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
_DATA_FONT    = Font(name="微软雅黑", size=9)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN         = Side(style="thin", color="CCCCCC")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COL_WIDTHS = {
    "平台": 8, "网站域名": 14, "内容分类": 12, "榜单": 12,
    "标题": 34, "排名": 6, "热度": 8, "标签": 10,
    "对应内容url": 52, "对应内容": 48,
    "发布时间": 20, "采集时间": 20, "批次id": 26, "原始来源": 18,
    "外网图片链接": 55, "内网图片链接": 55, "图片分辨率": 12,
}
_CENTER_COLS = {
    "平台", "网站域名", "内容分类", "榜单", "排名", "热度", "标签",
    "发布时间", "采集时间", "批次id", "原始来源", "图片分辨率",
}


def save_excel(records: list, filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "百度热搜榜"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 16)
    ws.row_dimensions[1].height = 22

    board_order = []
    for r in records:
        b = r.get("榜单", "")
        if b not in board_order:
            board_order.append(b)
    board_color = {b: _BOARD_COLORS[i % len(_BOARD_COLORS)] for i, b in enumerate(board_order)}

    for row_idx, rec in enumerate(records, 2):
        fill = PatternFill("solid", fgColor=board_color.get(rec.get("榜单", ""), "FFFFFF"))
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=rec.get(col_name, ""))
            cell.fill      = fill
            cell.font      = _DATA_FONT
            cell.border    = _BORDER
            cell.alignment = _ALIGN_CENTER if col_name in _CENTER_COLS else _ALIGN_LEFT
        ws.row_dimensions[row_idx].height = 45

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(filename)
    print(f"  Excel 已保存：{filename}  ({len(records)} 行)")


# ─── 全量采集主流程 ────────────────────────────────────────────────────────────

def collect_all(boards: dict = None, cities: list = None,
                delay: float = 1.5, img_dir: str = "",
                top_n: int = 0, no_content: bool = False) -> tuple:
    if boards is None:
        boards = {tab: BOARD_TABS[tab] for tab in TARGET_BOARDS}
    if cities is None:
        cities = TARGET_CITIES

    batch_id     = datetime.now().strftime("%Y%m%d%H%M%S") + "_" + str(uuid.uuid4())[:6]
    collect_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_records  = []
    start_dt     = datetime.now()

    if not img_dir:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        img_dir    = os.path.join(script_dir, "images", batch_id)
    os.makedirs(img_dir, exist_ok=True)

    n_label    = f"每榜 TOP {top_n}" if top_n > 0 else "每榜全量"
    mode_label = "基础模式（不抓内容）" if no_content else "完整模式（含新闻内容）"
    total_boards = len(boards) + len(cities)

    print(f"\n批次 ID   : {batch_id}")
    print(f"榜单数量  : {len(boards)}   城市数量: {len(cities)}   {n_label}")
    print(f"图片目录  : {img_dir}")
    print(f"采集模式  : {mode_label}")
    print(f"开始时间  : {collect_time}")
    print("=" * 70)

    # ── 采集各类榜单 ──
    for board_idx, (tab, board_name) in enumerate(boards.items(), 1):
        print(f"\n[{board_idx:02d}/{total_boards}] {board_name}", end="", flush=True)
        try:
            hot_list = get_board_hot_list(tab, board_name, top_n=top_n)
        except Exception as e:
            print(f"  热榜失败: {e}")
            continue

        print(f"  共 {len(hot_list)} 条", end="", flush=True)
        board_records = []

        for idx, hot_item in enumerate(hot_list):
            news = {}
            if not no_content:
                try:
                    news = fetch_top_news(hot_item["word"])
                except Exception:
                    pass

            rec = build_record(hot_item, news, collect_time, batch_id, img_dir)
            board_records.append(rec)

            has_img     = bool(rec["内网图片链接"])
            has_content = bool(rec["对应内容"])
            symbol      = "o" if has_img else "-"
            if has_content:
                symbol = symbol.upper()
            print(f" {symbol}", end="", flush=True)

            if not no_content and delay > 0:
                time.sleep(delay + 0.5 * (idx % 3))   # 0.5s 抖动

        all_records.extend(board_records)
        print(f"  OK {len(board_records)} 条入库")
        time.sleep(0.5)

    # ── 采集城市榜 ──
    for city_idx, city in enumerate(cities, len(boards) + 1):
        print(f"\n[{city_idx:02d}/{total_boards}] {city}城市榜", end="", flush=True)
        try:
            hot_list = get_city_hot_list(city, top_n=top_n)
        except Exception as e:
            print(f"  热榜失败（需中国大陆IP）: {e}")
            continue

        print(f"  共 {len(hot_list)} 条", end="", flush=True)
        city_records = []

        for idx, hot_item in enumerate(hot_list):
            news = {}
            if not no_content:
                try:
                    news = fetch_top_news(hot_item["word"])
                except Exception:
                    pass

            rec = build_record(hot_item, news, collect_time, batch_id, img_dir)
            city_records.append(rec)

            symbol = "O" if (rec["内网图片链接"] and rec["对应内容"]) else \
                     ("o" if rec["内网图片链接"] else ("C" if rec["对应内容"] else "-"))
            print(f" {symbol}", end="", flush=True)

            if not no_content and delay > 0:
                time.sleep(delay + 0.5 * (idx % 3))

        all_records.extend(city_records)
        print(f"  OK {len(city_records)} 条入库")
        time.sleep(0.5)

    elapsed = int((datetime.now() - start_dt).total_seconds())
    print(f"\n{'=' * 70}")
    print(f"采集完成  总计 {len(all_records)} 条  耗时 {elapsed} 秒")
    return all_records, batch_id


# ─── 主函数 ───────────────────────────────────────────────────────────────────

def main():
    import argparse, sys
    p = argparse.ArgumentParser(description="百度热搜榜单全量采集 → Excel")
    p.add_argument("--all-boards",  action="store_true", help="抓取全部11个榜单")
    p.add_argument("--board",       nargs="+",           help="指定榜单 tab，如 realtime finance")
    p.add_argument("--city",        nargs="+",           help="指定城市，如 北京 上海（需中国大陆IP）")
    p.add_argument("--city-all",    action="store_true", help="抓取全部城市榜（需中国大陆IP）")
    p.add_argument("--top",         type=int, default=0, help="每榜取前 N 条（默认全部）")
    p.add_argument("--no-content",  action="store_true", help="不抓新闻内容，只抓热榜列表（速度快3倍）")
    p.add_argument("--delay",       type=float, default=1.5,
                   help="每条间隔秒数（默认1.5，--no-content 时无效）")
    p.add_argument("--out-dir",     default=".",         help="Excel/JSON 保存目录（默认当前目录）")
    args = p.parse_args()

    # 解析榜单
    if args.board:
        boards = {tab: BOARD_TABS[tab] for tab in args.board if tab in BOARD_TABS}
        unknown = [t for t in args.board if t not in BOARD_TABS]
        if unknown:
            print(f"  [WARN] 未知 tab: {unknown}，可用：{list(BOARD_TABS.keys())}")
    elif args.all_boards:
        boards = BOARD_TABS.copy()
    else:
        boards = {tab: BOARD_TABS[tab] for tab in TARGET_BOARDS}

    # 解析城市
    if args.city_all:
        cities = list(CITIES_ALL)
    elif args.city:
        cities = [c for c in args.city if c in CITIES_ALL]
        unknown_c = [c for c in args.city if c not in CITIES_ALL]
        if unknown_c:
            print(f"  [WARN] 未知城市: {unknown_c}")
    else:
        cities = list(TARGET_CITIES)

    if not boards and not cities:
        print("没有可采集的榜单或城市，退出。")
        sys.exit(1)

    records, batch_id = collect_all(
        boards     = boards,
        cities     = cities,
        delay      = args.delay,
        top_n      = args.top,
        no_content = args.no_content,
    )

    if not records:
        print("无数据")
        return

    os.makedirs(args.out_dir, exist_ok=True)
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx   = os.path.join(args.out_dir, f"百度热搜_{ts}.xlsx")
    json_f = os.path.join(args.out_dir, f"百度热搜_{ts}.json")

    save_excel(records, xlsx)

    with open(json_f, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"  JSON 备份：{json_f}")


if __name__ == "__main__":
    main()
