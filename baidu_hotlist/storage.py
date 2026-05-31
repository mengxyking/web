import json
import csv
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


# ─── JSON ─────────────────────────────────────────────────────────────────────

def save_json(data: dict, out_dir: str = "data") -> str:
    _ensure_dir(out_dir)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"baidu_hotlist_{ts}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


# ─── CSV ──────────────────────────────────────────────────────────────────────

_CSV_FIELDS = ["board", "city", "rank", "word", "hot_tag", "label", "url"]


def save_csv(board_data: dict, out_dir: str = "data") -> str:
    _ensure_dir(out_dir)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"baidu_hotlist_{ts}.csv")
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_CSV_FIELDS)
        writer.writeheader()
        for items in board_data.values():
            writer.writerows(items)
    return path


# ─── Excel ────────────────────────────────────────────────────────────────────

COLUMNS = [
    "平台", "网站域名", "内容分类", "榜单", "城市",
    "标题", "排名", "热度标签", "内容标签",
    "对应内容url", "对应内容",
    "发布时间", "采集时间", "批次id", "原始来源",
    "外网图片链接", "内网图片链接", "图片分辨率",
]

_BOARD_COLORS = [
    "DDEEFF", "EEF5FF", "DDF0E8", "EEF8F3",
    "FFF3DD", "FFF9EE", "F0DDEE", "F8EEF5",
    "DDEEEE", "EEF5F5", "F5DDEE", "F0EEF5",
]
_HEADER_FILL  = PatternFill("solid", fgColor="2B5FAA")   # 百度蓝
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
_DATA_FONT    = Font(name="微软雅黑", size=9)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN         = Side(style="thin", color="CCCCCC")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COL_WIDTHS = {
    "平台": 8, "网站域名": 14, "内容分类": 12, "榜单": 12, "城市": 8,
    "标题": 34, "排名": 6, "热度标签": 8, "内容标签": 10,
    "对应内容url": 52, "对应内容": 48,
    "发布时间": 20, "采集时间": 20, "批次id": 26, "原始来源": 18,
    "外网图片链接": 55, "内网图片链接": 55, "图片分辨率": 12,
}
_CENTER_COLS = {
    "平台", "网站域名", "内容分类", "榜单", "城市", "排名",
    "热度标签", "内容标签", "发布时间", "采集时间", "批次id", "原始来源", "图片分辨率",
}


def save_excel(records: list, filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "百度热搜榜"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell            = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill       = _HEADER_FILL
        cell.font       = _HEADER_FONT
        cell.alignment  = _ALIGN_CENTER
        cell.border     = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 16)
    ws.row_dimensions[1].height = 22

    board_order = []
    for r in records:
        b = r.get("榜单", "")
        if b not in board_order:
            board_order.append(b)
    board_color = {b: _BOARD_COLORS[i % len(_BOARD_COLORS)] for i, b in enumerate(board_order)}

    for row_idx, rec in enumerate(records, 2):
        fill = PatternFill("solid", fgColor=board_color.get(rec.get("榜单", ""), "FFFFFF"))
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=rec.get(col_name, ""))
            cell.fill       = fill
            cell.font       = _DATA_FONT
            cell.border     = _BORDER
            cell.alignment  = _ALIGN_CENTER if col_name in _CENTER_COLS else _ALIGN_LEFT
        ws.row_dimensions[row_idx].height = 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(filename)
    print(f"  Excel 已保存：{filename}  ({len(records)} 行)")


# ─── 控制台摘要 ───────────────────────────────────────────────────────────────

def print_summary(board_data: dict, top_n: int = 5):
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = "=" * 62
    print(f"\n{line}")
    print(f"  百度热搜榜  {ts}")
    print(line)
    for name, items in board_data.items():
        if not items:
            print(f"\n  【{name}】 无数据")
            continue
        print(f"\n  【{name}】TOP {min(top_n, len(items))}")
        for item in items[:top_n]:
            rank = "置顶" if item["rank"] == 0 else str(item["rank"])
            tag  = f" [{item['hot_tag']}]" if item["hot_tag"] else ""
            lbl  = f" ({item['label']})"   if item["label"]   else ""
            print(f"    {rank:>4}. {item['word']}{tag}{lbl}")
    print(f"\n{line}\n")
