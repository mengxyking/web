from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import threading



class ThreadSafeExcelHandler:
    # 类级别的全局锁，确保所有实例共享同一把锁（同一时间只能有一个线程操作文件）
    _lock = threading.Lock()

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = None  # 工作簿对象（仅在操作时打开，避免长期占用）
        self.ws = None  # 当前工作表对象

    def _open(self, sheet_name=None):
        """内部方法：打开文件（仅在加锁后调用）"""
        self.wb = load_workbook(self.file_path)
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active

    def _close(self):
        """内部方法：关闭文件（仅在加锁后调用）"""
        if self.wb:
            self.wb.save(self.file_path)
            self.wb.close()
            self.wb = None  # 释放引用
            self.ws = None

    # -------------------------- 查（查询数据）--------------------------
    def get_cell_value(self, row, col, sheet_name=None):
        """获取指定单元格的值（线程安全）"""
        with self._lock:  # 加锁，确保唯一线程访问
            try:
                self._open(sheet_name)
                return self.ws.cell(row=row, column=col).value
            finally:
                self._close()

    def get_row(self, row_num, sheet_name=None):
        """获取指定行数据（线程安全）"""
        with self._lock:
            try:
                self._open(sheet_name)
                return [cell.value for cell in self.ws[row_num]]
            finally:
                self._close()

    def search(self, keyword, sheet_name=None):
        """搜索关键词（线程安全）"""
        with self._lock:
            try:
                self._open(sheet_name)
                results = []
                for row in self.ws.iter_rows(values_only=False):
                    for cell in row:
                        if cell.value and keyword in str(cell.value):
                            results.append({
                                "位置": f"{cell.column_letter}{cell.row}",
                                "值": cell.value
                            })
                return results
            finally:
                self._close()

    # -------------------------- 增（添加数据）--------------------------
    def add_row(self, data, row_num=None, sheet_name=None):
        """添加行（线程安全）"""
        with self._lock:
            try:
                self._open(sheet_name)
                if row_num:
                    self.ws.insert_rows(row_num)
                    target_row = row_num
                else:
                    target_row = self.ws.max_row + 1
                for col, value in enumerate(data, start=1):
                    self.ws.cell(row=target_row, column=col, value=value)
                print(f"已在第{target_row}行添加数据: {data}")
            finally:
                self._close()

    # -------------------------- 改（修改数据）--------------------------
    def update_cell(self, row, col, new_value, sheet_name=None):
        """修改单元格（线程安全）"""
        with self._lock:
            try:
                self._open(sheet_name)
                old_value = self.ws.cell(row=row, column=col).value
                self.ws.cell(row=row, column=col, value=new_value)
                print(f"已将({row},{col})从 {old_value} 修改为 {new_value}")
            finally:
                self._close()

    # -------------------------- 删（删除数据）--------------------------
    def delete_row(self, row_num, sheet_name=None):
        """删除行（线程安全）"""
        with self._lock:
            try:
                self._open(sheet_name)
                self.ws.delete_rows(row_num)
                print(f"已删除第{row_num}行")
            finally:
                self._close()

    def get_first_row_with_empty_3rd_4th(self, sheet_name=None):
        """
        获取第一行满足“第三列和第四列均为空（或None）”的行数据及行号
        返回：元组 (行数据列表, 行号)，若未找到则返回 (None, None)
        """
        with self._lock:  # 线程安全锁
            try:
                self._open(sheet_name)
                # 遍历所有行（从第1行开始，记录行号）
                for row_num, row in enumerate(self.ws.iter_rows(values_only=True), start=1):
                    # 处理第三列和第四列（索引2和3）
                    col3 = row[2] if len(row) > 2 else None
                    col4 = row[3] if len(row) > 3 else None

                    # 判断是否满足条件
                    if (col3 is None or str(col3).strip() == "") and (col4 is None or str(col4).strip() == ""):
                        return (list(row), row_num)  # 返回行数据和行号
                # 遍历完所有行仍未找到，返回(None, None)
                return None
            finally:
                self._close()
    def get_first_row_with_empty_4rd_5th(self, sheet_name=None):
        """
        获取第一行满足“第三列和第四列均为空（或None）”的行数据及行号
        返回：元组 (行数据列表, 行号)，若未找到则返回 (None, None)
        """
        with self._lock:  # 线程安全锁
            try:
                self._open(sheet_name)
                # 遍历所有行（从第1行开始，记录行号）
                for row_num, row in enumerate(self.ws.iter_rows(values_only=True), start=1):
                    # 处理第三列和第四列（索引2和3）
                    col3 = row[3] if len(row) > 3 else None
                    col4 = row[4] if len(row) > 4 else None

                    # 判断是否满足条件
                    if (col3 is None or str(col3).strip() == "") and (col4 is None or str(col4).strip() == ""):
                        return (list(row), row_num)  # 返回行数据和行号
                # 遍历完所有行仍未找到，返回(None, None)
                return None
            finally:
                self._close()
    def update_cell_by_row_col(self, row_num, col_num, value, sheet_name=None):
        """
        根据行号和列号修改单元格数据（行号、列号均从1开始，符合Excel规则）
        参数：
            row_num: 行号（int，从1开始）
            col_num: 列号（int，从1开始）
            value: 要设置的单元格值
            sheet_name: 工作表名称（None则使用当前工作表）
        返回：
            bool: 修改成功返回True，失败返回False（如行号/列号无效）
        """
        with self._lock:
            try:
                self._open(sheet_name)
                # 检查行号和列号是否有效（至少为1）
                if row_num < 1 or col_num < 1:
                    return False
                # 获取指定单元格并修改值（Excel单元格对象从1开始索引）
                cell = self.ws.cell(row=row_num, column=col_num)
                cell.value = value
                # 保存修改（如果你的_open/_close逻辑中没有自动保存，需要在这里添加保存代码）
                # 例如：self.wb.save(self.file_path)
                return True
            except Exception as e:
                # 捕获索引越界等异常（如行号超过工作表最大行数）
                print(f"修改单元格失败：{e}")
                return False
            finally:
                self._close()


