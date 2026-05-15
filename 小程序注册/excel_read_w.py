# 安装依赖：pip install openpyxl requests
from openpyxl import load_workbook
import requests

# ====================== 1. 读取全部数据（键值对格式） ======================
def read_excel(file_path):
    """
    读取Excel所有数据，返回【行号+标题键值对】格式
    返回示例：
    [
        {"行号": 1, "手机号": "手机号", "取码链接": "取码链接", "状态": "状态"},
        {"行号": 2, "手机号": "COM18----14798404719", ...},
    ]
    """
    wb = load_workbook(file_path)
    ws = wb.active
    all_data = []
    headers = [cell.value for cell in ws[1]]  # 获取表头（手机号、取码链接、状态）

    row_num = 1
    for row in ws.iter_rows(values_only=True):
        row_dict = {"行号": row_num}
        for idx, header in enumerate(headers):
            row_dict[header] = row[idx]
        all_data.append(row_dict)
        row_num += 1

    wb.close()
    return all_data


def get_row_with_empty_status(file_path):
    """
    查找表格中第一个「状态」为空的行（键值对格式）
    空值包含：None、空字符串、纯空白字符（如空格/制表符）
    :param file_path: Excel文件路径
    :return: 第一个状态为空的行字典，无则返回None
    """
    # 先读取所有数据
    all_data = read_excel(file_path)

    # 遍历所有行（跳过表头行，可选）
    for row_dict in all_data:
        # 排除表头行（如果表头是"状态"，不判定为空）
        if row_dict["行号"] == 1 and row_dict["状态"] == "状态":
            continue

        # 获取状态值，去除首尾空白字符
        status_value = row_dict.get("状态", "")
        if isinstance(status_value, str):
            status_value = status_value.strip()

        # 判断是否为空（None/空字符串/纯空白）
        if status_value is None or status_value == "":
            print(f"✅ 找到状态为空的行：行{row_dict['行号']} → {row_dict}")
            return row_dict

    # 没有找到状态为空的行
    print("❌ 未找到状态为空的行")
    return None

# ====================== 2. ✨ 新增：根据行号获取指定行的所有数据 ======================
def get_row_data(file_path, target_row_num):
    """
    根据行号获取对应行的完整数据（键值对格式）
    :param file_path: Excel文件路径
    :param target_row_num: 要获取的行号（比如2、3）
    :return: 该行的键值对字典，找不到返回None
    """
    all_data = read_excel(file_path)
    # 遍历查找指定行号
    for row_dict in all_data:
        if row_dict["行号"] == target_row_num:
            print(f"✅ 成功获取第{target_row_num}行数据：{row_dict}")
            return row_dict
    # 没找到的情况
    print(f"❌ 错误：表格中没有第{target_row_num}行")
    return None

# ====================== 5. ✨ 新增核心方法：按最大成功次数筛选有效行 ======================
def get_next_valid_row(file_path, max_allowed_count,erweima_suc_max_2,erweima_suc_max_3):
    """
    按顺序遍历Excel，找到【成功次数 <= 最大允许数】的第一条有效数据
    规则：
    1. 跳过表头（行号1）
    2. 如果 成功次数 > 最大数 → 跳过，取下一行
    3. 如果 成功次数为空/None/0 → 视为有效
    4. 找到第一条符合条件的行 → 返回该行完整数据
    5. 全部遍历完都没有 → 返回 None
    """
    all_data = read_excel(file_path)
    print(all_data)

    for row in all_data:
        # 跳过表头
        if row["行号"] == 1:
            continue

        # 获取当前行的成功次数
        success_count = row.get("成功次数")

        # 处理：空值 / None → 直接视为可用
        if success_count is None or success_count == "":
            print(f"✅ 找到有效行（成功次数为空）：行号 {row['行号']} | 数据：{row}")
            return row

        # 尝试转成数字（防止Excel存成字符串）
        try:
            count_num = int(success_count)
        except:
            count_num = 0

        erweima_t = row.get("二维码")
        if(erweima_t == 1):
            max_allowed_count = max_allowed_count
        elif(erweima_t == 2):
            max_allowed_count = erweima_suc_max_2
        elif (erweima_t == 3):
            max_allowed_count = erweima_suc_max_3
        else:
            max_allowed_count = 10

        # 判断：如果 <= 最大数 → 可用
        if int(count_num) < int(max_allowed_count):
            print(f"✅ 找到有效行：行号 {row['行号']} | 成功次数 {count_num} ≤ {max_allowed_count}")
            return row

        # 否则：跳过
        else:
            print(f"⏭️  跳过：行号 {row['行号']} | 成功次数 {count_num} > {max_allowed_count}")

    # 全部遍历完都没有
    print("❌ 所有行都已超过最大成功次数，无有效数据")
    return None
# ====================== 6. ✨ 新增：成功次数 +1 并保存 ======================
def increment_success_count(file_path, row_data):
    """
    对获取到的行数据，将【成功次数】+1，并写回Excel
    :param file_path: Excel文件路径
    :param row_data: 从 get_next_valid_row 获取到的行字典
    """
    if not row_data:
        print("❌ 无有效行数据，无法累加成功次数")
        return

    row_num = row_data["行号"]
    current_count = row_data.get("成功次数")
    print("current_count",current_count)

    # 处理空值/None → 默认为 0，然后 +1
    if current_count is None or current_count == "":
        print("none")
        new_count = 1
    else:
        try:
            new_count = int(current_count) + 1
        except:
            new_count = 1  # 转数字失败，默认设为1
    print("new_count=",new_count)

    # 调用你原来的 modify_data 方法修改
    modify_data(file_path, row_num, "成功次数", new_count)
    print(f"✅ 行{row_num} 成功次数 +1 → 最新值：{new_count}")
# ====================== 3. 根据【行号 + 标题】修改数据 ======================
def modify_data(file_path, row_num, header_name, new_value):
    """修改指定行+指定标题的数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    if header_name not in headers:
        print(f"❌ 错误：表格中没有「{header_name}」这个标题")
        wb.close()
        return

    col_num = headers.index(header_name) + 1
    ws.cell(row=row_num, column=col_num, value=new_value)
    wb.save(file_path)
    wb.close()
    print(f"✅ 修改成功：行{row_num} → 「{header_name}」 = {new_value}")

# ====================== 4. 访问取码链接获取验证码 ======================
def get_code_from_link(link):
    """访问取码链接获取验证码内容"""
    try:
        if "szfangmm.com:3000/" in link:
            token = link.split("3000/")[-1]
            api_url = f"http://sms.szfangmm.com:3000/api/smslist?token={token}"
            resp = requests.get(api_url, timeout=10)
            if resp.status_code == 200:
                return resp.text
        return None
    except Exception as e:
        print(f"❌ 访问链接失败：{e}")
        return None

# ====================== 使用示例（重点看新增方法） ======================
# if __name__ == '__main__':
#     #excel_file = r"D:\注册配置\ALTMVB3B17005679\手机号.xlsx"
#     excel_file = r"D:\注册配置\ALTMVB3B17005679\二维码.xlsx"
# #
# #     # 🌟 示例1：获取第2行的所有数据（最常用）
# #     row2_data = get_row_data(excel_file, target_row_num=2)
# #     if row2_data:
# #         # 直接通过标题取值，超方便！
# #         phone = row2_data["手机号"]
# #         link = row2_data["取码链接"]
# #         status = row2_data["状态"]
# #         print(f"\n📌 第2行详情：")
# #         print(f"手机号：{phone}")
# #         print(f"取码链接：{link}")
# #         print(f"状态：{status}")
# #
# #     # 🌟 示例2：获取第3行数据，并用取码链接拿验证码
# #     row3_data = get_row_data(excel_file, target_row_num=3)
# #     if row3_data and row3_data["取码链接"]:
# #         code_content = get_code_from_link(row3_data["取码链接"])
# #         print(f"\n📌 第3行取码链接返回内容：{code_content}")
# #         # 拿到验证码后，直接修改该行的状态
# #         modify_data(excel_file, row_num=3, header_name="状态", new_value="已获取验证码")
# #
# #     # 🌟 示例3：获取不存在的行（比如10行），会提示错误
# #     get_row_data(excel_file, target_row_num=10)
# #
# #
# #     # 1. 查看所有数据（带行号）
#     bbb = read_excel(excel_file)
#     print("---------ffffffffffffff--------------")
#     print(bbb)
#
    # excel_file = r"D:\注册配置\ALTMVB3B17005679\二维码.xlsx"
    # ccc = get_next_valid_row(excel_file,10,5,12)
    # print(ccc)
#     increment_success_count(excel_file,ccc)
# #
# #     # 2. 修改：第3行 + 标题“年龄” → 改成 30
#     #modify_data(excel_file, row_num=3, header_name="状态", new_value=301)
# #
# #     # 3. 再修改：第2行 + 标题“薪资” → 改成 10000
# #     modify_data(excel_file, row_num=2, header_name="状态", new_value=100001)
# #
# #     # 4. 查看最终结果
# #     read_excel(excel_file)