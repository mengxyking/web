import base64
import hashlib
import json
import platform
import shutil
import sys
import threading
import random
import traceback
import uuid
from datetime import datetime
#from RapidOCR_json.api.python.demo1 import OCRProcessor
import lxml.etree as etree
import uiautomator2 as u2
import excel_read_w
import zhuce

from PyQt6.QtGui import QBrush, QColor
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem,
    QScrollArea, QVBoxLayout, QWidget, QCheckBox, QHBoxLayout, QPushButton, QLabel, QRadioButton, QLineEdit,
    QFileDialog, QTextEdit, QTabWidget, QFrame, QSpinBox, QComboBox
)
from PyQt6.QtCore import Qt, QTimer

from paddleOCR_json_duixiang_0308 import OCRProcessor
ocr1 = OCRProcessor(id=1)
ocr2 = OCRProcessor(id=2)
# 2. 创建全局锁（保证轮询时的线程安全）
ocr_lock = threading.Lock()
#print(ocr)
current_scroll_position = 0
def get_available_ocr():
    """获取空闲的OCR对象，没有则等待（轮询ocr1→ocr2）"""
    while True:
        with ocr_lock:
            # 先查ocr1，空闲则返回
            if not ocr1.busy:
                return ocr1
            # 再查ocr2，空闲则返回
            if not ocr2.busy:
                return ocr2
            # if not ocr3.busy:
            #     return ocr3
            # if not ocr4.busy:
            #     return ocr4
            # if not ocr5.busy:
            #     return ocr5
        # 都忙则等待50ms再查，避免CPU空转
        time.sleep(0.05)


file_lock = threading.Lock()
video_lock = threading.Lock()

def get_device(serial):
    #d = ""
    #print("之前的d", d)
    #print(f"正在连接设备: {serial}")
    d = u2.connect(serial)
    d.watcher.remove()
    return d
def getPhotoPath():
    pan = os.getcwd().split(':')[0] + ":"
    pic_path = pan + '//yangmao/pic'  # 标志图片文件 新路径
    #print("00000000000000000000000000---------")
    #print(pic_path)
    if (os.path.exists(pic_path) == False):
        os.makedirs(pic_path)
    return pic_path
def photo(s):
    Ui_file_Name =  str(int(time.time()))+"_"+str(s)+"_ui.png"
    path = getPhotoPath()+"/"+Ui_file_Name
    return path

def get_files_in_directory(directory_path):
    files = []
    try:
        # 创建一个 Path 对象
        path = Path(directory_path)

        # 检查路径是否存在且是一个目录
        if not path.exists() or not path.is_dir():
            print(f"The directory {directory_path} does not exist or is not a directory.")
            return files

            # 遍历目录中的所有文件并添加到列表中
        for file_path in path.iterdir():
            if file_path.is_file():
                files.append(file_path)
    except PermissionError:
        print(f"Permission denied to access {directory_path}.")

    return files  # 将 Path 对象转换为字符串列表


from pathlib import Path


def create_directory_if_not_exists(directory_path):
    path = Path(directory_path)
    if not path.exists():
        path.mkdir(parents=True)
        print(f"Directory '{directory_path}' created.")
    else:
        print(f"Directory '{directory_path}' already exists.")

    # 示例用法


def random_click_view(d,view):
    bottom = view.info["bounds"]["top"]
    left = view.info["bounds"]["left"]

    random_x = int(left)+random.randint(0,5)
    random_y = int(bottom) + random.randint(0, 5)

    d.click(random_x,random_y)
def shell_neibu(cmd):
    os.system(cmd)

def connect_port(ip_t,port_t):
    cmd = f"adb connect {ip_t}:{port_t}"
    print("cmd=",cmd)
    shell_neibu(cmd)

import random
import os

def get_windows_desktop():
    """
    获取Windows系统的桌面目录（优先读取DESKTOP环境变量，兼容所有Windows版本）
    :return: 桌面绝对路径字符串
    """
    # 方案1：直接读取DESKTOP环境变量（Windows 10/11 优先推荐）
    if 'DESKTOP' in os.environ:
        desktop_path = os.environ['DESKTOP']
    # 方案2：拼接用户目录（兼容所有Windows版本）
    else:
        desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')

    # 确保目录存在（避免特殊情况，如桌面被手动删除）
    os.makedirs(desktop_path, exist_ok=True)
    return desktop_path
from datetime import datetime

def get_current_date(format_str="%Y_%m_%d"):
    """
    获取当前日期，支持自定义格式
    :param format_str: 日期格式，默认"%Y_%m_%d"
    :return: 格式化后的日期字符串
    """
    return datetime.now().strftime(format_str)
start_time = datetime.now()

win_deskTop_path = get_windows_desktop()

def tongji(service):
    service = service.replace(".", "_")
    service = service.replace(":", "_")

    rizhimulu = win_deskTop_path + "/签到记录"
    create_directory_if_not_exists(rizhimulu)
    rizhimulu_service = rizhimulu + "/" + str(service)
    create_directory_if_not_exists(rizhimulu_service)
    riqi = get_current_date()
    riqi_path = rizhimulu_service + "/" + riqi
    create_directory_if_not_exists(riqi_path)

    return riqi_path
# tongji("129.168.1.1")
# print("win_deskTop_path=",win_deskTop_path)
def operate_device(serial):
    count_zong = 0
    while(True):
        #try:
            import datetime
            import time
            if(int(datetime.datetime.now().timestamp()) > 1866582706):
                print("")
                #return

            result = main_control(serial)
            if (result == "88"):
                print("运行结束")
                filepath = './shuju/' + serial + ".pkl"
                print("filepath-->", filepath)
                if (os.path.isfile(filepath)):
                    updata_pkl(filepath, "执行状态", "运行结束")
                    updata_pkl(filepath, "进行的任务", "空闲")
                return
            count_zong += 1
        # except BaseException as ee:
        #     print("崩溃了",ee)
        #     operate_device(serial)


def check_time_difference(interval_seconds):
    if(interval_seconds == 0):
        return False
    # 获取当前时间
    end_time = datetime.now()
    # 计算时间差（以秒为单位）
    time_difference = (end_time - start_time).total_seconds()
    print("time_difference=",time_difference)
    # 如果时间差大于100秒，则返回True，否则返回False
    return time_difference > interval_seconds
#搜索路径、评论文件路径、任务列表、运行时长、更换频率小、更换频率大、视频滑动间隔小、视频滑动间隔大、视频滑动次数、收藏、评论、点赞、关注
def main_control(serial):
    # task = get_value_by_key_pkl("shuju_config.pkl", "task")
    # print("task=",task)
    # #d = get_device(serial)
    # if(str(task).count("douyinyanghao")):
    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "小程序注册")
    print("")
    main_douyin(serial)
    return "88"
def create_folder_on_current_disk():
    # 获取当前代码文件所在的路径
    current_script_path = os.path.abspath(__file__)
    # 提取当前代码所在的磁盘（如 'C:\\' 或 'D:\\'）
    current_disk = os.path.splitdrive(current_script_path)[0] + os.sep
    # 拼接新建文件夹的路径（以磁盘根目录为例）
    folder_path = os.path.join(current_disk, "dy_temp")
    # 新建文件夹
    try:
        os.makedirs(folder_path)
        print(f"已在 {current_disk} 上成功创建文件夹：{folder_path}")
    except FileExistsError:
        print(f"文件夹 {folder_path} 已存在")
    except Exception as e:
        print(f"创建文件夹失败：{e}")
    return folder_path
def take_screenshot(d):
    try:
        SAVE_DIR = create_folder_on_current_disk()
        # 生成带时间戳的文件名，避免重复
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        save_path = os.path.join(SAVE_DIR, f"screenshot_{timestamp}.png")
        # 截图并保存
        d.screenshot(save_path)
        print(f"截图已保存至：{save_path}")
        return save_path
    except BaseException as e:
        print("截图时，发生崩溃", str(e))
        return None
def take_screenshot_path_tt(d,path,huiyuan):
    try:
        SAVE_DIR = path + "会员截图"

        if not os.path.exists(SAVE_DIR):
            os.makedirs(SAVE_DIR)
            print("文件夹已新建")
        else:
            print("文件夹已存在")

        # 生成带时间戳的文件名，避免重复
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        save_path = os.path.join(SAVE_DIR, f"{huiyuan}_{timestamp}.png")
        # 截图并保存
        d.screenshot(save_path)
        print(f"截图已保存至：{save_path}")
        return save_path
    except BaseException as e:
        print("截图时，发生崩溃", str(e))
        return None

def take_screenshot_path(d,path,name):
    try:
        name  = str(name).replace(".","_")
        name = str(name).replace(":", "_")

        SAVE_DIR = path
        # 生成带时间戳的文件名，避免重复
        save_path = os.path.join(SAVE_DIR, f"{name}.png")
        # 截图并保存
        d.screenshot(save_path)
        print(f"截图已保存至：{save_path}")
        return save_path
    except BaseException as e:
        print("截图时，发生崩溃", str(e))
        return None

def check_and_reconnect(d, device,package_name, element_len=30):
    """
    检查当前界面元素数量，与指定长度比较，决定是否重建设备连接
    :param d: 现有uiautomator2设备对象
    :param element_len: 参考元素数量阈值
    :param device: 设备标识（序列号或IP:端口）
    :return: 处理后的uiautomator2设备对象（可能是原对象或重建的对象）
    """
    try:
        # 获取当前界面XML结构
        xml = d.dump_hierarchy()
        # 解析XML并获取所有node元素（即UI控件）
        tree = etree.fromstring(xml.encode('utf-8'))
        elements = tree.xpath('//node')
        current_len = len(elements)
        print(f"当前界面元素数量: {current_len}, 阈值: {element_len}")

        # 比较元素数量
        if current_len > element_len:
            # 元素数量大于阈值，返回原设备对象
            print("元素数量符合要求，使用原设备连接")
            return d
        else:
            # 元素数量小于等于阈值，重建连接
            print(f"元素数量不足，尝试重建设备连接...")

            d.stop_uiautomator(wait=True)
            if ":" in device:  # 仅对IP:端口格式的无线设备执行disconnectd
                subprocess.run(f"adb disconnect {device}", shell=True, capture_output=True, text=True)
                print(f"已执行adb disconnect {device}，彻底断开无线ADB连接")

            # 停止当前uiautomator服务
            time.sleep(3)
            new_d = u2.connect(device)
            new_d.start_uiautomator()
            new_d.app_start(package_name=package_name)
            print(f"设备 {device} 已重新连接")

            xml = d.dump_hierarchy()
            # 解析XML并获取所有node元素（即UI控件）
            tree = etree.fromstring(xml.encode('utf-8'))
            elements = tree.xpath('//node')
            current_len = len(elements)
            print(f"重连后: {current_len}, 阈值: {element_len}")

            return d

    except Exception as e:
        print(f"处理过程发生错误: {str(e)}，尝试重建连接...")
        # 发生异常时也尝试重建连接
        try:
            d.stop_uiautomator(wait=True)
        except:
            pass  # 忽略停止服务时的错误
        time.sleep(3)
        new_d = u2.connect(device)
        new_d.start_uiautomator()
        new_d.app_start(package_name="com.tencent.mm")

        return new_d

def backToHome_qiandao_tupianshibie(d,flag,flag_text):
    print(f"flag_text={flag_text}")
    dd =  0
    time.sleep(2)
    while(dd < 4):
        d.stop_uiautomator()
        print(d.dump_hierarchy())
        if(flag == 1):
            #print(d.dump_hierarchy())
            elements = d(text=flag_text)  # 获取所有文本为'some_text'的元素
            print(len(elements))
            if(len(elements)>0):
                return "1"
            time.sleep(1.5)
            d.press("back")
            time.sleep(1.5)
            dd += 1
        if (flag == 2):
            elements = d(description=flag_text)  # 获取所有文本为'some_text'的元素
            # print(len(elements))
            if (len(elements) > 0):
                return "1"
            time.sleep(1.5)
            d.press("back")
            time.sleep(1.5)
            dd += 1
def backToHome_qiandao(d,flag,flag_text):
    print(f"flag_text={flag_text}")
    dd =  0
    time.sleep(2)
    while(dd < 4):
        if(flag == 1):
            #print(d.dump_hierarchy())
            elements = d(text=flag_text)  # 获取所有文本为'some_text'的元素
            print(len(elements))
            if(len(elements)>0):
                return "1"
            time.sleep(1.5)
            d.press("back")
            time.sleep(1.5)
            dd += 1
        if (flag == 2):
            elements = d(description=flag_text)  # 获取所有文本为'some_text'的元素
            # print(len(elements))
            if (len(elements) > 0):
                return "1"
            time.sleep(1.5)
            d.press("back")
            time.sleep(1.5)
            dd += 1


import xml.etree.ElementTree as ET
import re
def count_nodes_with_y_gt_200(d):
    """
    统计XML中y坐标大于200的节点数量
    :param xml_content: XML内容字符串
    :return: 符合条件的节点数量
    """
    # 编译正则表达式提取bounds中的坐标
    xml_content = d.dump_hierarchy()
    bounds_pattern = re.compile(r'\[(\d+),(\d+)\]\[(\d+),(\d+)\]')
    # 解析XML
    root = ET.fromstring(xml_content)
    # 初始化计数器
    count = 0

    # 遍历所有node节点
    for node in root.iter('node'):

        bounds = node.get('bounds')
        if not bounds:
            continue

        # 提取坐标（x1, y1, x2, y2）
        match = bounds_pattern.match(bounds)
        title_text = node.get("text")
        if match:
            y1 = int(match.group(2))  # 取左上角y坐标作为判断依据
            if y1 > 500 and title_text != "":
                #print(node.get("text"))
                count += 1

    return count
def count_nodes_with_y_gt_200_text_is_null(d):
    """
    统计XML中y坐标大于200的节点数量
    :param xml_content: XML内容字符串
    :return: 符合条件的节点数量
    """
    # 编译正则表达式提取bounds中的坐标
    xml_content = d.dump_hierarchy()
    bounds_pattern = re.compile(r'\[(\d+),(\d+)\]\[(\d+),(\d+)\]')
    # 解析XML
    root = ET.fromstring(xml_content)
    # 初始化计数器
    count = 0

    # 遍历所有node节点
    for node in root.iter('node'):

        bounds = node.get('bounds')
        if not bounds:
            continue

        # 提取坐标（x1, y1, x2, y2）
        match = bounds_pattern.match(bounds)
        title_text = node.get("text")
        if match:
            y1 = int(match.group(2))  # 取左上角y坐标作为判断依据
            if y1 > 500:
                #print(node.get("text"))
                count += 1

    return count

def get_app_page(app_biaoji, target_app):
    """
    根据APP名称（key）获取对应的页面名称（value）
    :param app_biaoji: 原始APP-页面列表（如你的app_biaoji）
    :param target_app: 要查询的APP名称（key）
    :return: 对应的页面名称（value）/ None（未找到）
    """
    for item in app_biaoji:
        # 遍历每个字典项，取唯一的key-value对
        app_name = next(iter(item.keys()))
        page_name = next(iter(item.values()))
        if app_name == target_app:
            return page_name
    return None
def main_douyin(serial):
    task = get_value_by_key_pkl("shuju_config.pkl", "task_app") #
    file_path = get_value_by_key_pkl("shuju_config.pkl", "file_path")  # file_path
    erweima_suc_max = get_value_by_key_pkl("shuju_config.pkl", "tongchengguanjianzi")  # file_path
    d = get_device(serial)
    w,h = d.window_size()



    d.watcher.when("以后再说").click()
    d.watcher.when("我知道了").click()
    d.watcher.when("允许").click()
    d.watcher.when("取消").click()
    d.watcher.when("确定").click()
    d.watcher.start()
    phone_path_temp = file_path +"/" + serial
    shoujihao_file = phone_path_temp + "/" +"手机号.xlsx"
    erweima_file = phone_path_temp + "/" + "二维码.xlsx"
    print(phone_path_temp,shoujihao_file,erweima_file)
    if(not os.path.isdir(phone_path_temp) or not os.path.isfile(shoujihao_file) or not os.path.isfile(erweima_file)):
        print("当前手机没有配置数据")
        return 0
    error_couont = 0

    for i in range(1000):
        if(error_couont > 10):
            return 88
        try:
            result_yewu = zhuyao_yewu(d,serial ,shoujihao_file, erweima_file,phone_path_temp, erweima_suc_max,w)  # 0 代表一般错误，可以继续循环运行的，1代表成功，88代表 直接退出的  77代表需要多试几次的

        except BaseException as e:
            traceback.print_exc()  # 打印完整错误
            result_yewu = 77
        if(result_yewu == 1):
            print("当前成功了啊，继续")
        elif(result_yewu == 88):
            print("当前资源不足（二维码或者手机号），终止")
            return 88
        elif(result_yewu == 77):
            error_couont += 1
            print("当前需要重试")
def open_wechat_main(DEVICE_SERIAL):
    cmd = f"adb -s {DEVICE_SERIAL} shell am start -n com.tencent.mm/.ui.LauncherUI"
    subprocess.run(cmd, shell=True)


def click_wechat(d,index):
    """
    点击 text="微信" 的控件
    :param index: 1 = 第一个，2 = 第二个
    """
    time.sleep(1)
    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoints_by_data(all_data, "微信")
    print("point = ",point)
    print("len(point) = ", len(point))
    points_len = len(point)
    if index == 1:
        time.sleep(1)
        if(points_len>=1):
            print("开始点击1")
            print(point[0][0],point[0][1]-100)
            d.click(point[0][0],point[0][1]-100)

    if index == 2:
        time.sleep(1)
        if(points_len>=1):
            print("开始点击2")
            print(point[1][0],point[1][1]-100)
            d.click(point[1][0],point[1][1]-100)


    # try:
    #     if index == 1:
    #         c
    #         # 点击 第一个 text="微信"
    #         points_len = len(point)
    #         if(points_len>=1):
    #             print("开始点击1")
    #             print(point[points_len-2][0],point[points_len-2][1])
    #             d.click(point[points_len-2][0],point[points_len-2][1])
    #
    #     elif index == 2:
    #         print("点击第2个微信")
    #         # 点击 第二个 text="微信"
    #         points_len = len(point)
    #         if (points_len >= 2):
    #             print("开始点击1")
    #             print(point[points_len - 1][0], point[points_len - 1][1])
    #             d.click(point[points_len - 1][0], point[points_len - 1][1])
    #
    #     else:
    #         print("❌ 只能传入 1 或 2")
    #
    # except Exception as e:
    #     print(f"❌ 点击失败：{str(e)}")
def kill_all_wechat(DEVICE_SERIAL):
    # 杀死所有微信进程（主程序 + 分身 全部清空）
    cmd = f"adb -s {DEVICE_SERIAL} shell am force-stop com.tencent.mm"
    subprocess.run(cmd, shell=True)
def zhuyao_yewu(d,serial,shoujihao_file,erweima_file,phone_path_temp,erweima_suc_max,w):

    result_excel = excel_read_w.get_row_with_empty_status(shoujihao_file)
    print(result_excel)

    erweima_suc_max_1 = get_value_by_key_pkl("shuju_config.pkl", "tongchengguanjianzi")  # file_path
    erweima_suc_max_2 = get_value_by_key_pkl("shuju_config.pkl", "huifuxiaoxiyonghunicheng")  # file_path
    erweima_suc_max_3 = get_value_by_key_pkl("shuju_config.pkl", "huifuxiaoxiyonghunicheng_1")  # file_path


    if(result_excel == None):
        print("手机号已经用完了")
        return 88
    phone_safe2 = result_excel.get('手机号', '')
    print("安全取值的手机号：", phone_safe2)
    if(phone_safe2 == None or phone_safe2 == ""):
        print("手机号为空")
        return 88

    qumalianjie = result_excel.get('取码链接', '')
    hanghao = result_excel.get('行号', '')
    print("安全取值的手机号：", phone_safe2)
    if(not str(phone_safe2).count("----")):
        print("手机号meiyou ----")
        excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "手机号不是标准的")
        return 0
    phone_safe_t = str(phone_safe2).split("----")[1].strip()

    if(len(phone_safe_t) != 11):
        print("手机号不是十一位 ----")
        excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "手机号不是标准的")
        return 0

    result_excel_erweima = excel_read_w.get_next_valid_row(erweima_file,erweima_suc_max,erweima_suc_max_2,erweima_suc_max_3)
    print("-----erweima----->",result_excel_erweima)
    #excel_read_w.increment_success_count(erweima_file, result_excel_erweima)
    if (result_excel_erweima == None):
        print("二维码已经用完了")
        return 88
    # print("chengongcishujiayi")
    # excel_read_w.increment_success_count(erweima_file, result_excel_erweima)

    weixin_count = get_value_by_key_pkl("shuju_config.pkl", "weixin_count")
    kill_all_wechat(serial)  #weixin_count
    time.sleep(0.5)
    d.press("home")
    time.sleep(0.5)
    d.press("home")
    time.sleep(1.5)
    # open_wechat_main(serial)
    # time.sleep(2)
    weixinhao_tt = 1
    if(str(weixin_count).count("微信2")>0):
        print("点击微信2")
        weixinhao_tt = 2
        click_wechat(d,2)
    else:
        print("点击微信1")
        weixinhao_tt = 1
        click_wechat(d, 1)
    time.sleep(8)

    result_back = back_to_wx_home(d)
    if(result_back == 0):
        print("进入微信失败")
        return 77

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data,"发现")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(2)
    else:
        print("没有发现按钮")
        return 77

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "扫一扫")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(2)
    else:
        print("没有发现扫一扫")
        return 0

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "相册")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(2)
    else:
        print("没有发现相册")
        d.click(939,2107)#临时先点我的手机固定坐标
        time.sleep(3)
        #return 0

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "拍摄照片")
    if (point != None):
        d.click(split_num_to_four_parts(w,result_excel_erweima["行号"]-1), point[1] + random.randint(-3, 3))
        time.sleep(5)
    else:
        print("没有发现拍摄照片")
        return 0

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "允许")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(3)
    else:
        print("没有发现允许")
        #return 0



    for i in range(5):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "换个账号登录")
        if (point != None):
            d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
            time.sleep(5)
            break
        time.sleep(3)
    else:
        print("没有发现换个账号登录")
        return 0

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "请输入手机号")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(2)

        #输入手机号
        phone_num = phone_safe_t
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)

        for nn in phone_num:
            point = ocr.getPoint_by_data_back(all_data,nn)

            if (point != None):
                d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
                time.sleep(0.2)
            else:
                print("meiyou",nn)
                return 0
    else:
        print("没有发现请输入手机号")
        return 0
    #excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "已完成")
    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "获取验证码")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(3)
    else:
        print("没有发现获取验证码")
        return 0

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "同意并继续")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(10)
    else:
        print("没有发现同意并继续")
        #return 0

    result_zhuce = zhuce.get_shoulvrujia_code(qumalianjie,phone_num)
    if(result_zhuce == None):
        excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "获取不到验证码")
        return 0

    if(len(result_zhuce) != 6):
        excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "验证码个数不是六位")
        print("验证码不是六位")
        return 0

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "已发送验证码")
    point_denglu = ocr.getPoint_by_data_true(all_data, "登录")
    if (point_denglu != None):
        excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "老用户")
        return 0
    if (point != None):
        d.click(112, point[1] + 160)
        time.sleep(3)
    else:
        print("没有发现已发送验证码")
        return 0

    phone_num = result_zhuce
    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)

    for nn in phone_num:
        point = ocr.getPoint_by_data_back(all_data, nn)

        if (point != None):
            d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
            time.sleep(0.2)
        else:
            print("meiyou", nn)
            return 0

    path_photo = take_screenshot(d)
    ocr = get_available_ocr()
    all_data = ocr.yewu(path_photo)
    point = ocr.getPoint_by_data_back(all_data, "注册")
    if (point != None):
        d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
        time.sleep(3)
    else:
        print("没有发现注册按钮")
        return 0

    for i in range(5):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "立即认证")
        if (point != None):
            d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
            time.sleep(3)
            break
        time.sleep(3)
    else:
        print("没有发现立即认证")
        return 0

    for i in range(5):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "允许")
        if (point != None):
            d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
            time.sleep(1)
            break
        time.sleep(5)
    else:
        print("没有发现允许")
        return 0


    for i in range(2):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "验证成功")
        if (point != None):
            print("验证chenggong")
            #d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
            time.sleep(1)
            break
        time.sleep(1)
    else:
        print("没有发现验证成功")
        #return 0

    for i in range(1):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "验证失败")
        if (point != None):
            print("验证失败")
            #d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
            time.sleep(1)
            break
        time.sleep(5)
    else:
        print("没有发现验证失败")
        #return 0
    for i in range(8):
        point = ocr.getPoint_by_data_true(all_data, "学生认证")
        if (point != None):
            d.click(67, point[1])
            time.sleep(1)
            break
        time.sleep(3)
    else:
        print("没有发现学生认证")
        return 0

    for i in range(8):
        point = ocr.getPoint_by_data_true(all_data, "学生认证")
        if (point != None):
            d.click(67, point[1])
            time.sleep(1)
            break
        time.sleep(3)
    else:
        print("没有发现学生认证")
        return 0



    for i in range(5):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "行程")
        point_wode  = ocr.getPoint_by_data_back(all_data, "我的")
        if (point != None and point_wode != None):
            print("点击我的")
            d.click(point_wode[0] + random.randint(-3, 3), point_wode[1] + random.randint(-3, 3))
            time.sleep(3)
            break
        time.sleep(3)
    else:
        print("没有进入小程序首页")
        return 0

    for i in range(5):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "支付码")
        point_wode  = ocr.getPoint_by_data_back(all_data, "客服")
        if (point != None and point_wode != None):
            print("进入我的页面成功")
            #d.click(point_wode[0] + random.randint(-3, 3), point_wode[1] + random.randint(-3, 3))
            #time.sleep(5)
            break
        time.sleep(3)
    else:
        print("没有进入我的页面成功")
        return 0

    point_ruhuiyuan = ocr.getPoint_by_data_back(all_data, "如会员")
    point_jinhuiyuan = ocr.getPoint_by_data_back(all_data, "金会员")


    if (point_ruhuiyuan != None ):
        print("d当前是入会员")
        excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "如会员")#phone_path_temp

        take_screenshot_path_tt(d,phone_path_temp,"如会员")

        #return 1
    elif(point_ruhuiyuan == None and point_jinhuiyuan != None):
        print("d当前是jin会员")
        excel_read_w.modify_data(shoujihao_file, int(hanghao), "状态", "金会员")
        take_screenshot_path_tt(d, phone_path_temp,"金会员")
        #return 1
    else:
        print("当前未知错误啊，在我的页面")
    excel_read_w.modify_data(shoujihao_file, int(hanghao), "微信号", str(weixinhao_tt))  # phone_path_temp
    excel_read_w.modify_data(shoujihao_file, int(hanghao), "二维码", str(result_excel_erweima["行号"]-1))  # phone_path_temp

    excel_read_w.increment_success_count(erweima_file,result_excel_erweima)

    point_zhifuma = ocr.getPoint_by_data_back(all_data, "支付码")
    if (point_zhifuma != None ):
        print("点击用户昵称 进入设置页面")
        print(150, point_zhifuma[1] -38)
        d.click(150, point_zhifuma[1] -38)

    for i in range(5):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "退出登录")
        point_wode  = ocr.getPoint_by_data_back(all_data, "会员信息")
        if (point != None and point_wode != None):
            print("进入会员信息页面")
            d.click(point[0] + random.randint(-3, 3), point[1] + random.randint(-3, 3))
            time.sleep(3)
            break
        time.sleep(3)
    else:
        print("没有进入会员信息页面")
        return 0


    for i in range(5):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        point = ocr.getPoint_by_data_back(all_data, "取消")
        point_wode  = ocr.getPoint_by_data_back(all_data, "确定")
        if (point != None and point_wode != None):
            print("点击确定")
            d.click(point_wode[0] + random.randint(-3, 3), point_wode[1] + random.randint(-3, 3))
            time.sleep(1)
            break
        time.sleep(8)
    else:
        print("没有确定按钮")
        return 0

    return 1

def split_num_to_four_parts(num,target_mm, round_type="round"):
    # 计算每份的长度
    part_length = num / 4

    # 定义4个区间的起止值
    parts_range = [
        (part_length, part_length * 2),
        (part_length * 2, part_length * 3),
        (part_length * 3, num)
    ]

    # 计算每一份的中间值（取整）
    middle_values = []
    print(f"将 {num} 分成4等份，每份长度：{part_length}")
    print(f"取整方式：{round_type}")
    print("-" * 60)

    for i, (start, end) in enumerate(parts_range, 1):
        middle = (start + end) / 2  # 先算中间值

        # 根据选择的方式取整
        if round_type == "round":
            middle_int = round(middle)  # 四舍五入
        elif round_type == "int":
            middle_int = int(middle)  # 直接截断小数（比如 135.9 → 135）
        elif round_type == "ceil":
            import math
            middle_int = math.ceil(middle)  # 向上取整（比如 135.1 → 136）
        else:
            middle_int = round(middle)  # 默认四舍五入

        middle_values.append(middle_int)
        print(f"第{i}份区间：{start} ~ {end} → 中间值（原始）：{middle} → 取整后：{middle_int}")

    print("-" * 60)
    return middle_values[target_mm-1]
def back_to_wx_home(d):

    for i in range(10):
        path_photo = take_screenshot(d)
        ocr = get_available_ocr()
        all_data = ocr.yewu(path_photo)
        print(all_data)
        if(str(all_data).count("通讯录")>0 and str(all_data).count("发现")>0  ):
            return 1
        d.press("back")
        time.sleep(2)
    else:
        return 0




def sousuo_douyin_2(serial,d):
    huifuxiaoxiyonghunicheng = get_value_by_key_pkl("shuju_config.pkl", "huifuxiaoxiyonghunicheng")
    tongchengguanjianzi = get_value_by_key_pkl("shuju_config.pkl", "tongchengguanjianzi")
    bbb = None
    if (d(text="搜索").exists(timeout=5)):
        bbb = d(text="搜索")
        random_click_view(d, d(text="搜索"))
        time.sleep(3)
    else:
        print("当前没有搜索按钮")
        return 0

    if (d(className="android.widget.EditText").exists(timeout=15)):
        print("输入关键字搜索")
        sss = d(className="android.widget.EditText")
        print("------->",len(sss))

        for iii in sss:
            print("---------------")
            print(iii)
        if(len(sss)>0):
            sss[len(sss)-1].set_text(huifuxiaoxiyonghunicheng)
            time.sleep(3)
            #sss[len(sss) - 1].info

            # top = sss[len(sss) - 1].info["bounds"]["top"]
            # right =  sss[len(sss) - 1].info["bounds"]["right"]
            # print(top,right)
            # d.click(right+160,top+20)

            if(bbb):
                bbb.click()

        time.sleep(8)
    else:
        print("mei有输入框")
        return 0

    # if (d(text="搜索").exists(timeout=5)):
    #     #random_click_view(d, d(text="搜索"))
    #
    #     sss = d(text="搜索")
    #     print("------->", len(sss))
    #
    #     for iii in sss:
    #         print("---------------")
    #         print(iii)
    #     if (len(sss) > 0):
    #         random_click_view(d,sss[len(sss) - 1])
    #     time.sleep(3)
    #
    #     time.sleep(3)
    # else:
    #     print("当前没有搜索按钮")
    #     return 0

    if (d(text="直播中").exists(timeout=5)):
        random_click_view(d, d(text="直播中"))
        time.sleep(3)
    else:
        print("当前没有直播中")
        return 0


    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "抖音挂机中")
    time.sleep(int(tongchengguanjianzi))
    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "挂机完毕")

    return 1

def sousuo_douyin(serial,d):
    w,h = d.window_size()
    huifuxiaoxiyonghunicheng = get_value_by_key_pkl("shuju_config.pkl", "huifuxiaoxiyonghunicheng")
    tongchengguanjianzi = get_value_by_key_pkl("shuju_config.pkl", "tongchengguanjianzi")
    time.sleep(8)
    bbb = None
    if (d(text="搜索").exists(timeout=5)):
        bbb = d(text="搜索")
        random_click_view(d, d(text="搜索"))
        time.sleep(3)
    else:
        print("当前没有搜索按钮")
        return 0

    if (d(className="android.widget.EditText").exists(timeout=15)):
        print("输入关键字搜索")
        sss = d(className="android.widget.EditText")
        print("------->",len(sss))

        for iii in sss:
            print("---------------")
            print(iii)
        if(len(sss)>0):
            sss[len(sss)-1].set_text(huifuxiaoxiyonghunicheng)
        time.sleep(3)
    else:
        print("mei有输入框")
        return 0

    if(bbb):
        bbb.click()

    # d.click(w-230,170)
    time.sleep(5)

    # if (d(text="搜索").exists(timeout=5)):
    #     random_click_view(d, d(text="搜索"))
    #
    #     # sss = d(text="搜索")
    #     # print("------->", len(sss))
    #     #
    #     # for iii in sss:
    #     #     print("---------------")
    #     #     print(iii)
    #     # if (len(sss) > 0):
    #     #     random_click_view(d,sss[len(sss) - 1])
    #     # time.sleep(3)
    #
    #     time.sleep(3)
    # else:
    #     print("当前没有搜索按钮")
    #     return 0

    if (d(text="直播中").exists(timeout=15)):
        random_click_view(d, d(text="直播中"))
        time.sleep(3)
    else:
        print("当前没有直播中")
        return 0


    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "抖音挂机中")
    time.sleep(int(tongchengguanjianzi))
    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "挂机完毕")

    return 1



def douyin_app(serial,d,package_name,w,h):
    d.app_start(package_name=package_name,stop=True)
    time.sleep(10)
    backToHome_douyin(d)

    goto_zhibo_tab(d)
    time.sleep(3)
    huifuxiaoxiyonghunicheng = get_value_by_key_pkl("shuju_config.pkl", "huifuxiaoxiyonghunicheng")
    tongchengguanjianzi = get_value_by_key_pkl("shuju_config.pkl", "tongchengguanjianzi")
    for i in range(5000):
        print("")

        if (d(textContains=huifuxiaoxiyonghunicheng).exists(timeout=5)):
            d.click(w/2,h/2)
            time.sleep(3)
            print("找到主播了")
            break

        else:
            print("当前没有，往下滑")
            beisaier_small(d)
            time.sleep(random.randint(1,6))
    else:
        return 0 #tongchengguanjianzi
    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "抖音挂机中")
    time.sleep(int(tongchengguanjianzi))
    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "挂机完毕")

    return 1











def wannengyaoshi(serial,d,package_name,w,h):
    d.app_start(package_name=package_name,stop=True)
    time.sleep(10)

    #先判断有没有 弹窗 有的话，点击关闭按钮
    # if (d(resourceId='com.kmxs.reader:id/imageView_close').exists(timeout=3)):
    #     d(resourceId='com.kmxs.reader:id/imageView_close').click()
    #     time.sleep(2)
    # else:
    #     print("当前没有弹窗")
    #     #return "55"

    backToHome_wannengyaoshi(d)


    # 点击douyin 商城
    if (d(text="抖音优惠券").exists(timeout=5)):
        random_click_view(d, d(text="抖音优惠券"))
        time.sleep(3)

    else:
        print("当前没有抖音领券")
        return 0

    if (d(text="搜索").exists(timeout=5)):
        print("当前任务成功")
        return 1

    else:
        print("当前没有我的")
        return 0

def hanxiaoquan(serial,d,package_name,w,h):
    d.app_start(package_name=package_name,stop=True)
    time.sleep(10)

    #先判断有没有 弹窗 有的话，点击关闭按钮
    # if (d(resourceId='com.kmxs.reader:id/imageView_close').exists(timeout=3)):
    #     d(resourceId='com.kmxs.reader:id/imageView_close').click()
    #     time.sleep(2)
    # else:
    #     print("当前没有弹窗")
    #     #return "55"

    backToHome_hanxiaoquan(d)

    #点击我的
    # if (d(text="我的").exists(timeout=5)):
    #     random_click_view(d,d(text="我的"))
    #
    # else:
    #     print("当前没有我的")
    #     return 0
    d.click(w-50,h-130)
    time.sleep(3)

    beisaier_small(d)

    # 点击douyin 商城
    if (d(text="抖音领券").exists(timeout=5)):
        random_click_view(d, d(text="抖音领券"))
        time.sleep(3)

    else:
        print("当前没有抖音领券")
        return 0

    if (d(text="搜索").exists(timeout=5)):
        print("当前任务成功")
        return 1

    else:
        print("当前没有我的")
        return 0


def kuan(serial,d,package_name,w,h):
    d.app_start(package_name=package_name,stop=True)
    time.sleep(10)

    #先判断有没有 弹窗 有的话，点击关闭按钮
    # if (d(resourceId='com.kmxs.reader:id/imageView_close').exists(timeout=3)):
    #     d(resourceId='com.kmxs.reader:id/imageView_close').click()
    #     time.sleep(2)
    # else:
    #     print("当前没有弹窗")
    #     #return "55"

    backToHome_kuan(d)

    #点击我的
    if (d(text="话题").exists(timeout=5)):
        random_click_view(d,d(text="话题"))

    else:
        print("当前没有话题")
        return 0

    # 点击薅羊毛小分队
    if (d(text="薅羊毛小分队").exists(timeout=5)):
        random_click_view(d, d(text="薅羊毛小分队"))

    else:
        print("当前没有话题")
        return 0
    if (d(text="酷安专属,支持搜券").exists(timeout=5)):
        random_click_view(d, d(text="酷安专属,支持搜券"))

    else:
        print("当前没有话题")
        return 0

    #beisaier_small(d)

    if (d(text="搜索").exists(timeout=5)):
        print("当前任务成功")
        return 1

    else:
        print("当前没有我的")
        return 0


def dejian(serial,d,package_name,w,h):
    d.app_start(package_name=package_name,stop=True)
    time.sleep(10)

    #先判断有没有 弹窗 有的话，点击关闭按钮
    if (d(resourceId='com.chaozh.iReader.dj:id/Id_withdraw_close').exists(timeout=3)):
        d(resourceId='com.chaozh.iReader.dj:id/Id_withdraw_close').click()
        time.sleep(2)
    else:
        print("当前没有弹窗")
        #return "55"

    backToHome_dejian(d)

    #点击福利
    d.click(w/2,h-150)
    time.sleep(3)

    # beisaier_small(d)

    # 往右滑动切到 抖音商城页面
    d.swipe(1000,800,100,800,0.3)
    # if (d(text="抖音商城").exists(timeout=5)):
    #     random_click_view(d, d(text="抖音商城"))
    #
    # else:
    #     print("当前没有我的")
    #     return 0
    time.sleep(3)
    if (d(text="搜索").exists(timeout=5)):
        print("当前任务成功")
        return 1
    else:
        print("当前没有我的")
        return 0

def hema(serial,d,package_name):
    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "开始hema")
    d.app_start(package_name=package_name,stop=True)
    time.sleep(10)

    #先判断有没有 弹窗 有的话，点击关闭按钮
    # if (d(resourceId='com.kmxs.reader:id/imageView_close').exists(timeout=3)):
    #     d(resourceId='com.kmxs.reader:id/imageView_close').click()
    #     time.sleep(2)
    # else:
    #     print("当前没有弹窗")
    #     #return "55"

    backToHome_qimao(d)

    #点击我的
    if (d(text="我的").exists(timeout=5)):
        random_click_view(d,d(text="我的"))
        time.sleep(4)

    else:
        print("当前没有我的")
        return 0

    #beisaier_small(d)

    # 点击douyin 商城
    if (d(text="我的商城").exists(timeout=5)):
        random_click_view(d, d(text="我的商城"))
        time.sleep(3)

    else:
        print("当前没有我的商城")
        return 0

    if (d(text="搜索").exists(timeout=5)):
        print("hema任务成功")
        return 1

    else:
        print("当前没有搜索")
        return 0


def qimao(serial,d,package_name,w,h):
    updata_pkl("./shuju/" + serial + ".pkl", "进行的任务", "开始七猫")
    d.app_start(package_name=package_name,stop=True)
    time.sleep(10)

    #先判断有没有 弹窗 有的话，点击关闭按钮
    if (d(resourceId='com.kmxs.reader:id/imageView_close').exists(timeout=3)):
        d(resourceId='com.kmxs.reader:id/imageView_close').click()
        time.sleep(2)
    else:
        print("当前没有弹窗")
        #return "55"

    backToHome_qimao_benmao(d)

    #点击我的
    if (d(text="我的").exists(timeout=5)):
        random_click_view(d,d(text="我的"))

    else:
        d.click(w - 50, h - 130)

    beisaier_small(d)

    # 点击douyin 商城
    if (d(text="抖音商城").exists(timeout=5)):
        random_click_view(d, d(text="抖音商城"))

    else:
        print("当前没有我的")
        return 0

    if (d(text="搜索").exists(timeout=5)):
        return 1

    else:
        print("当前没有我的")
        return 0

def backToHome_kuan(d):
    dd =  0
    time.sleep(1)
    for i in range(10):
        elements = d(text='话题')  # 获取所有文本为'some_text'的元素
        #print(len(elements))
        if(len(elements)>0):
            return "1"
        #time.sleep(1.5)
        print("还得返回")
        d.press("back")
        time.sleep(1.5)
def backToHome_qimao(d):
    dd =  0
    time.sleep(1)
    for i in range(10):
        elements = d(text='我的')  # 获取所有文本为'some_text'的元素
        #print(len(elements))
        if(len(elements)>0):
            return "1"
        #time.sleep(1.5)

        print("还得返回")
        d.press("back")
        time.sleep(1.5)

def backToHome_wannengyaoshi(d):
    dd =  0
    time.sleep(1)
    for i in range(10):
        elements = d(text='抖音优惠券')  # 获取所有文本为'some_text'的元素
        #print(len(elements))
        if(len(elements)>0):
            return "1"
        #time.sleep(1.5)

        print("还得返回")
        d.press("back")
        time.sleep(3.5)
def goto_zhibo_tab(d):
    tuijian_y = 0
    if (d(text="推荐").exists(timeout=5)):
        print("当前在首页了。。。。。。。。")
        tuijian_y = d(text="推荐").info["bounds"]["top"] + 3
    else:
        print("当前bu在首页了。。。。。。。。")
        return
    for i in range(5):

        if (d(text="直播").exists(timeout=1.5)):
            print("当前找到同城乐。。。。。。。。")
            random_click_view(d, d(text="直播"))
            return 1
        d.swipe(200, tuijian_y, 900, tuijian_y, 0.5)
        time.sleep(1)
    return 0
def backToHome_douyin(d):
    dd =  0
    time.sleep(1)
    for i in range(10):
        elements = d(text='推荐')  # 获取所有文本为'some_text'的元素
        #print(len(elements))
        if(len(elements)>0):
            return "1"
        #time.sleep(1.5)

        print("还得返回")
        d.press("back")
        time.sleep(3.5)



def backToHome_qimao_benmao(d):
    w,h = d.window_size()
    dd =  0
    for i in range(10):

        time.sleep(1)
        elements = d(text='发现')  # 获取所有文本为'some_text'的元素
        #print(len(elements))
        if(len(elements)>0):
            return "1"
        #time.sleep(1.5)
        if (d(resourceId='com.kmxs.reader:id/iv_close').exists(timeout=1)):
            d(resourceId='com.kmxs.reader:id/iv_close').click()
            time.sleep(2)
        else:
            print("当前没有弹窗")

        print("还得返回")
        d.press("back")
        time.sleep(1.5)


def backToHome_hanxiaoquan(d):
    dd =  0
    time.sleep(1)
    for i in range(10):
        elements = d(text='精选')  # 获取所有文本为'some_text'的元素
        #print(len(elements))
        if(len(elements)>0):
            return "1"
        #time.sleep(1.5)
        print("还得返回")
        d.press("back")
        time.sleep(1.5)


def backToHome_dejian(d):
    dd =  0
    time.sleep(1)
    for i in range(10):
        elements = d(text='分类')  # 获取所有文本为'some_text'的元素
        #print(len(elements))
        if(len(elements)>0):
            return "1"
        #time.sleep(1.5)
        print("还得返回")
        d.press("back")
        time.sleep(0.5)

def append_to_txt(file_path, content):
    """
    以追加模式向TXT文件写入内容
    :param file_path: TXT文件路径（当前目录直接写文件名）
    :param content: 要写入的内容（字符串）
    """
    # 'a' 模式：追加模式，文件不存在则自动创建；encoding=utf-8 避免中文乱码
    with open(file_path, 'a', encoding='utf-8') as f:
        #f.write(content)
        # 可选：每行内容后加换行符（避免内容挤在一起）
        f.write(content + '\n')
def create_file_if_not_exists(file_path):
    """
    判断文件是否存在，不存在则创建空文件
    :param file_path: 文件的路径（相对/绝对路径均可）
    """
    # 1. 判断文件是否存在
    if not os.path.exists(file_path):
        # 2. 不存在则创建空文件（w 模式会自动新建，a 模式也可）
        with open(file_path, 'w', encoding='utf-8') as f:
            # 可选：写入初始内容（如空文件则无需这行）
            # f.write("初始内容")
            pass
        print(f"文件 {file_path} 不存在，已新建")
    else:
        print(f"文件 {file_path} 已存在，无需创建")
def beisaier_small(d,Diract="up"):
    # 获取屏幕尺寸
    width, height = d.window_size()

    # 设置起点和终点
    if Diract == "up":
        random_start_point_x = random.uniform(0.3, 0.6)
        random_start_point_y = random.uniform(0.7, 0.8)
        random_end_point_x = random.uniform(0.2, 0.99)
        random_end_point_y = 0.3

        start_point = (width * random_start_point_x, height * random_start_point_y)  # 屏幕中下位置
        end_point = (width * random_start_point_x, height * random_end_point_y)  # 屏幕中上位置

    # 设置控制点(可选)
    # 控制点会影响曲线的形状
    control_points = [
        (width * 0.3, height * 0.6),  # 第一个控制点
        (width * 0.7, height * 0.4)  # 第二个控制点
    ]

    # 执行贝塞尔曲线滑动
    swipe_along_bezier(d, start_point, end_point, control_points, steps=15, duration=1.2)

    # 等待一下
    time.sleep(1)
def shanghua(d,Diract="up"):
    # 获取屏幕尺寸
    width, height = d.window_size()

    # 设置起点和终点
    if Diract == "up":
        random_start_point_x = random.uniform(0.4, 0.4)
        random_start_point_y = random.uniform(0.7, 0.7)
        random_end_point_x = random.uniform(0.2, 0.99)
        random_end_point_y = 0.3

        start_point = (width * random_start_point_x, height * random_start_point_y)  # 屏幕中下位置
        end_point = (width * random_start_point_x, height * random_end_point_y)  # 屏幕中上位置
    if Diract == "down":
        random_start_point_x = random.uniform(0.4, 0.4)
        random_end_point_y = random.uniform(0.7, 0.7)
        random_end_point_x = random.uniform(0.2, 0.99)
        random_start_point_y = 0.3

        start_point = (width * random_start_point_x, height * random_start_point_y)  # 屏幕中下位置
        end_point = (width * random_start_point_x, height * random_end_point_y)  # 屏幕中上位置

    # 设置控制点(可选)
    # 控制点会影响曲线的形状
    control_points = [
        (width * 0.3, height * 0.6),  # 第一个控制点
        (width * 0.7, height * 0.4)  # 第二个控制点
    ]

    # 执行贝塞尔曲线滑动
    swipe_along_bezier(d, start_point, end_point, control_points, steps=15, duration=1.2)

    # 等待一下
    time.sleep(1)
def beisaier_random(d,Diract="up"):
    # 获取屏幕尺寸
    width, height = d.window_size()

    # 设置起点和终点
    if Diract == "up":
        random_start_point_x = random.uniform(0.3, 0.6)
        random_start_point_y = random.uniform(0.78, 0.82)
        random_end_point_x = random.uniform(0.2, 0.99)
        random_end_point_y = 0.01

        start_point = (width * random_start_point_x, height * random_start_point_y)  # 屏幕中下位置
        end_point = (width * random_start_point_x, height * random_end_point_y)  # 屏幕中上位置

    # 设置控制点(可选)
    # 控制点会影响曲线的形状
    control_points = [
        (width * 0.3, height * 0.6),  # 第一个控制点
        (width * 0.7, height * 0.4)  # 第二个控制点
    ]

    # 执行贝塞尔曲线滑动
    random_num = random.uniform(0.3, 1.2)
    swipe_along_bezier(d, start_point, end_point, control_points, steps=15, duration=random_num)

    # 等待一下
    time.sleep(1)
def bezier_curve(start_point, end_point, control_points=None, steps=50):
    """
    生成贝塞尔曲线的轨迹点

    参数:
    - start_point: 起点坐标 (x, y)
    - end_point: 终点坐标 (x, y)
    - control_points: 控制点坐标列表，可以是1个或多个控制点 [(x1, y1), (x2, y2), ...]
    - steps: 生成的轨迹点数量

    返回:
    - 贝塞尔曲线上的点列表
    """
    points = [start_point]

    if control_points is None:
        # 如果没有提供控制点，使用一个随机控制点生成二阶贝塞尔曲线
        cx = start_point[0] + random.randint(-50, 50)
        cy = start_point[1] + random.randint(-50, 50)
        control_points = [(cx, cy)]

    # 生成贝塞尔曲线上的点
    for t in range(1, steps + 1):
        t = t / steps
        # 初始化点坐标
        x = 0
        y = 0
        n = len(control_points) + 1  # 阶数

        # 计算组合数系数
        def comb(n, k):
            if k == 0 or k == n:
                return 1
            result = 1
            for i in range(1, k + 1):
                result = result * (n - i + 1) // i
            return result

        # 贝塞尔曲线公式: B(t) = Σ C(n, i) * (1-t)^(n-i) * t^i * P_i
        # 计算曲线上每个点的坐标
        all_points = [start_point] + control_points + [end_point]
        for i in range(n + 1):
            coeff = comb(n, i) * ((1 - t) ** (n - i)) * (t ** i)
            x += coeff * all_points[i][0]
            y += coeff * all_points[i][1]

        points.append((int(x), int(y)))

    return points

def swipe_along_bezier(d, start_point, end_point, control_points=None, steps=15, duration=0.3):
    """
    沿着贝塞尔曲线滑动

    参数:
    - d: uiautomator2设备实例
    - start_point: 起点坐标 (x, y)
    - end_point: 终点坐标 (x, y)
    - control_points: 控制点坐标列表
    - steps: 生成的轨迹点数量
    - duration: 滑动持续时间(秒)
    """
    # 生成贝塞尔曲线上的点
    points = bezier_curve(start_point, end_point, control_points, steps)

    # 计算每个步骤的间隔时间
    interval = duration / len(points)

    # 按下起点
    d.swipe_points(points, duration=interval)
def beisaier(d,Diract="up"):
    # 获取屏幕尺寸
    width, height = d.window_size()

    # 设置起点和终点
    if Diract == "up":
        random_start_point_x = random.uniform(0.3, 0.6)
        random_start_point_y = random.uniform(0.78, 0.82)
        random_end_point_x = random.uniform(0.2, 0.99)
        random_end_point_y = 0.01

        start_point = (width * random_start_point_x, height * random_start_point_y)  # 屏幕中下位置
        end_point = (width * random_start_point_x, height * random_end_point_y)  # 屏幕中上位置

    # 设置控制点(可选)
    # 控制点会影响曲线的形状
    control_points = [
        (width * 0.3, height * 0.6),  # 第一个控制点
        (width * 0.7, height * 0.4)  # 第二个控制点
    ]

    # 执行贝塞尔曲线滑动
    swipe_along_bezier(d, start_point, end_point, control_points, steps=15, duration=0.3)

    # 等待一下
    time.sleep(1)

class PklViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.selected_ids = []
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"小程序注册,欢迎:{get_real_device_id()}")
        self.setGeometry(100, 100, 550, 500)  # 适当增大窗口尺寸

        # ====== 顶部标题和手机列表区域 ======
        self.titleLabel = QLabel(" " * 70 + "手 机 列 表" + " " * 70)
        self.titleLabel.setStyleSheet("""  
            QLabel {  
                font-size: 14px;  
                font-family: "Arial", sans-serif;  
                padding: 10px;  
                font-weight: bold;  
                background-color: #f0f0f0;  
                color: #333;  
            }  
        """)

        # 手机列表表格
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(7)
        self.table_widget.setHorizontalHeaderLabels(
            ['选中', '编号', '昵称', '连接状态', '运行状态', '当前任务', "滑动统计"])
        self.table_widget.setColumnWidth(0, 30)
        self.table_widget.setShowGrid(True)
        self.table_widget.itemChanged.connect(self.on_item_changed)

        # 表格滚动区域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidget(self.table_widget)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setFixedHeight(200)  # 适当减小高度，给Tab区域留出空间

        # 操作区域标题和选择框
        self.caozuo_tiel = QLabel(" " * 70 + "签   到" + " " * 70)
        self.caozuo_tiel.setStyleSheet("""  
            QLabel {  
                font-size: 14px;  
                font-weight: bold;  
                font-family: "Arial", sans-serif;  
                padding: 10px;  
                background-color: #f0f0f0;  
                color: #333;  
            }  
        """)

        self.select_phone_layout = QHBoxLayout()
        self.select_phone_layout.setSpacing(10)  # 组件之间的间距（避免拥挤）

        # 2. 创建“手机选择”标签
        self.select_phone_label = QLabel("选中手机：")
        # 可选：固定标签宽度（避免文字换行）
        self.select_phone_label.setFixedWidth(80)

        # 3. 创建文本输入框
        self.select_phone_input = QLineEdit()
        self.select_phone_input.setPlaceholderText("输入序号（如1、1-5、1,3）")
        # 可选：固定输入框宽度（控制输入区域大小）
        self.select_phone_input.setFixedWidth(200)

        # 4. 创建确认选择按钮
        self.confirm_select_button = QPushButton("确认选择")
        self.confirm_select_button.clicked.connect(self.confirm_phone_selection)
        # 可选：固定按钮宽度
        self.confirm_select_button.setFixedWidth(100)

        # 5. 将三个组件加入水平布局
        self.select_phone_layout.addWidget(self.select_phone_label)
        self.select_phone_layout.addWidget(self.select_phone_input)
        self.select_phone_layout.addWidget(self.confirm_select_button)
        # 可选：添加拉伸（让组件靠左，右侧留空）
        self.select_phone_layout.addStretch(1)

        self.horizontal_layout = QHBoxLayout()
        self.radio_button_select = QPushButton("全选")
        self.radio_button_select.clicked.connect(self.on_item_clicked)
        self.radio_button0 = QLabel("           ")
        self.radio_button1 = QCheckBox("小程序注册")
        self.radio_button1.setChecked(True)
        self.radio_button2 = QCheckBox("小红书养号")
        self.radio_button2.setChecked(False)
        self.radio_button3 = QCheckBox("养号之后是否关闭抖音")
        self.radio_button3.setChecked(True)
        self.radio_button5 = QLabel("           ")

        self.horizontal_layout.addWidget(self.radio_button_select)
        self.horizontal_layout.addWidget(self.radio_button0)
        self.horizontal_layout.addWidget(self.radio_button1)
        self.horizontal_layout.addWidget(self.radio_button5)

        # ====== Tab布局区域 ======
        self.tab_widget = QTabWidget()
        self.tab_widget.setTabPosition(QTabWidget.TabPosition.North)  # 标签在顶部

        # 添加各个Tab页
        self.init_douyin_tab()  # 抖音配置Tab
        self.init_advanced_tab()  # 高级设置Tab

        # ====== 底部按钮区域 ======
        self.button_gang = QHBoxLayout()
        self.execute_button = QPushButton("执行")
        self.execute_button.resize(80, 30)
        self.equit_button = QPushButton("点击记录")
        self.equit_button.resize(80, 30)
        self.clear_task_config_button = QPushButton('保存配置', self)
        #self.clear_task_config_button.clicked.connect(self.clear_task)
        self.execute_button_delete = QPushButton("一键清空")
        self.execute_button_delete.resize(80, 30)
        self.execute_button_reset = QPushButton("重置")
        self.execute_button_reset.resize(80, 30)

        self.button_gang.addWidget(self.execute_button)
        #self.button_gang.addWidget(self.equit_button)
        self.button_gang.addWidget(self.clear_task_config_button)
        self.button_gang.addWidget(self.execute_button_delete)
        self.button_gang.addWidget(self.execute_button_reset)


        # 绑定按钮事件
        self.execute_button.clicked.connect(self.execute_button_clicked)
        self.execute_button_reset.clicked.connect(self.execute_reset_button_clicked)
        self.clear_task_config_button.clicked.connect(self.execute_save_button_clicked)
        self.execute_button_delete.clicked.connect(self.execute_delete_button_clicked)
        self.equit_button.clicked.connect(self.shoudong_zhixing)

        # ====== 主布局组装 ======
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        main_layout.addWidget(self.titleLabel)
        main_layout.addWidget(self.scroll_area)
        #main_layout.addWidget(self.caozuo_tiel)
        #main_layout.addLayout(self.select_phone_layout)
        main_layout.addLayout(self.horizontal_layout) #self.select_phone_layout
        main_layout.addWidget(self.tab_widget)  # 添加Tab控件
        main_layout.addLayout(self.button_gang)

        # 初始化定时器
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.refresh_pkl_files)
        self.timer.start(30000)

        self.timer1 = QTimer(self)
        self.timer1.start(1300)

        self.refresh_pkl_files()
    def confirm_phone_selection(self):
        input_text = self.select_phone_input.text().strip()
        if not input_text:
            return

        # 清空之前的选择
        self.selected_ids = []

        # 解析输入的序号
        selected_numbers = set()
        try:
            # 分割输入为各个部分
            parts = [p.strip() for p in input_text.replace(',', ' ').split()]

            for part in parts:
                if '-' in part:
                    # 处理范围格式 1-5
                    start, end = map(int, part.split('-'))
                    # 确保范围正确（小到大）
                    if start > end:
                        start, end = end, start
                    # 添加范围内所有数字
                    selected_numbers.update(range(start, end + 1))
                else:
                    # 处理单个数字
                    selected_numbers.add(int(part))
        except:
            # 输入格式错误时不做任何选择
            selected_numbers = set()
            print("输入格式错误")

        # 遍历表格中的所有行（行号从0开始，序号从1开始）
        for row in range(self.table_widget.rowCount()):
            # 当前行的序号（行号+1）
            current_number = row + 1  # 关键修改：用行号作为序号

            # 获取当前行的复选框
            checkbox = self.table_widget.cellWidget(row, 0)
            if not isinstance(checkbox, QCheckBox):
                continue

            # 获取当前行的手机ID（用于添加到selected_ids）
            item_id_text = self.table_widget.item(row, 1).text()

            # 检查是否需要选中
            is_selected = current_number in selected_numbers
            print("is_selected=",is_selected)

            # 更新复选框状态
            if(is_selected == True):
                checkbox.setChecked(is_selected)

            #更新选中ID列表
            if is_selected:
                self.selected_ids.append(item_id_text)

        print(f"最终选中的手机序号: {selected_numbers}")
        print(f"最终选中的手机ID: {self.selected_ids}")
        #self.selected_ids = []
        self.refresh_pkl_files()
    def shoudong_zhixing(self):
        print("")
        # Get the current text from the QTextEdit
        # list =  pkl_list("log.pkl")
        # current_text = self.text_edit.toPlainText()
        # for bbb in list:
        #     if(str(bbb) in str(current_text)):
        #         pass
        #     else:
        #         # Append new text (for example, a new line with some content)
        #         new_text = str(bbb)+"--->"+str(list[bbb])  # You can change this to whatever text you want to add
        #         # Set the new text (old text + new text)
        #         self.text_edit.setPlainText(current_text +"\n" +new_text)
        #         # Optionally, move the cursor to the end of the text
        #         cursor = self.text_edit.textCursor()
        #         # cursor.movePosition(cursor.End)
        #         self.text_edit.setTextCursor(cursor)
        #         # Scroll to the bottom (optional, depending on your needs)
        #         scrollbar = self.text_edit.verticalScrollBar()
        #         scrollbar.setValue(scrollbar.maximum())
        # if(os.path.isdir("./task_config")):
        #     shutil.rmtree("./task_config")
        #     self.refresh_pkl_files_test()
        if (self.equit_button.text() == "点击记录"):

            if not os.path.exists("pause.txt"):
                # 如果文件不存在，则创建文件
                with open("pause.txt", 'w') as file:
                    pass  # 这里不需要写入任何内容，只需要创建文件即可

            self.equit_button.setText("结束记录")
        else:
            if os.path.exists("pause.txt"):
                os.remove("pause.txt")
            self.equit_button.setText("点击记录")

    # ====== 初始化各个Tab页 ======
    def init_douyin_tab(self):
        """抖音配置Tab页"""
        douyin_tab = QWidget()
        douyin_tab.setStyleSheet("background-color: #f0f0f0;")
        douyin_tab.setStyleSheet("""
            QWidget {
                background-color: #f5f5f5;  /* 整体背景为浅灰色 */
            }
            QCheckBox {
                background-color: #f5f5f5;  /* 整体背景为浅灰色 */
                border-radius: 1px;
                padding: 1px;
            }
            QLineEdit {
                background-color: white;    /* 输入框保持白色 */
                border: 1px solid #cccccc;  /* 添加边框提升清晰度 */
                border-radius: 3px;
                padding: 2px;
            }
            QPushButton {
                background-color: #e0e0e0;  /* 按钮背景色 */
                border-radius: 3px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #d0d0d0;  /* 按钮悬停效果 */
            }
        """)

        layout = QVBoxLayout(douyin_tab)
        self.horizontal_layout_dy_task1 = QHBoxLayout()
        # 标题标签（跨行吗对齐，设置固定宽度）
        self.radio_button_dy_task0 = QLabel("商城APP选择:")
        self.radio_button_dy_task0.setFixedWidth(105)  # 固定标题宽度，方便换行对齐

        # 原有6个复选框（保留原有宽度设置）
        self.radio_button_dy_task1 = QCheckBox("七猫免费小说")
        self.radio_button_dy_task1.setChecked(False)
        self.radio_button_dy_task1.setFixedWidth(100)

        self.radio_button_dy_task2 = QCheckBox("河马剧场")
        self.radio_button_dy_task2.setChecked(False)
        self.radio_button_dy_task2.setFixedWidth(100)

        self.radio_button_dy_task3 = QCheckBox("得间免费小说")
        self.radio_button_dy_task3.setChecked(False)
        self.radio_button_dy_task3.setFixedWidth(100)

        self.radio_button_dy_task4 = QCheckBox("酷安")
        self.radio_button_dy_task4.setChecked(False)
        self.radio_button_dy_task4.setFixedWidth(100)

        self.radio_button_dy_task5 = QCheckBox("韩小圈")
        self.radio_button_dy_task5.setChecked(False)
        self.radio_button_dy_task5.setFixedWidth(100)

        # 把第一行控件添加到第一行水平布局
        self.horizontal_layout_dy_task1.addWidget(self.radio_button_dy_task0)
        self.horizontal_layout_dy_task1.addWidget(self.radio_button_dy_task1)
        self.horizontal_layout_dy_task1.addWidget(self.radio_button_dy_task2)
        self.horizontal_layout_dy_task1.addWidget(self.radio_button_dy_task3)
        self.horizontal_layout_dy_task1.addWidget(self.radio_button_dy_task4)
        self.horizontal_layout_dy_task1.addWidget(self.radio_button_dy_task5)
        # 第一行末尾添加伸缩项（防止控件拉伸）
        self.horizontal_layout_dy_task1.addStretch()

        # 3. 第二行水平布局（番茄免费小说 + 新增5个选项）
        self.horizontal_layout_dy_task2 = QHBoxLayout()
        # 空标签（占位，和第一行标题对齐）
        self.empty_label = QLabel("")
        self.empty_label.setFixedWidth(105)  # 和标题宽度一致，保证对齐

        # 原有番茄免费小说
        self.radio_button_dy_task6 = QCheckBox("WiFi万能钥匙")
        self.radio_button_dy_task6.setChecked(False)
        self.radio_button_dy_task6.setFixedWidth(100)

        # 新增5个复选框（可自定义名称和宽度）
        self.radio_button_dy_task7 = QCheckBox("抖音APP")
        self.radio_button_dy_task7.setChecked(False)
        self.radio_button_dy_task7.setFixedWidth(100)

        self.radio_button_dy_task8 = QCheckBox("红果免费漫剧")
        self.radio_button_dy_task8.setChecked(False)
        self.radio_button_dy_task8.setFixedWidth(100)

        self.radio_button_dy_task9 = QCheckBox("番茄音乐")
        self.radio_button_dy_task9.setChecked(False)
        self.radio_button_dy_task9.setFixedWidth(100)

        self.radio_button_dy_task10 = QCheckBox("抖音火山版")
        self.radio_button_dy_task10.setChecked(False)
        self.radio_button_dy_task10.setFixedWidth(100)

        # 把第二行控件添加到第二行水平布局
        self.horizontal_layout_dy_task2.addWidget(self.empty_label)  # 占位对齐
        self.horizontal_layout_dy_task2.addWidget(self.radio_button_dy_task6)
        self.horizontal_layout_dy_task2.addWidget(self.radio_button_dy_task7)
        # self.horizontal_layout_dy_task2.addWidget(self.radio_button_dy_task8)
        # self.horizontal_layout_dy_task2.addWidget(self.radio_button_dy_task9)
        # self.horizontal_layout_dy_task2.addWidget(self.radio_button_dy_task10)
        # 第二行末尾添加伸缩项
        self.horizontal_layout_dy_task2.addStretch()

        # ===================== 新增第三行布局 =====================
        self.horizontal_layout_dy_task3 = QHBoxLayout()
        # 空标签占位（和标题对齐）
        self.empty_label2 = QLabel("")
        self.empty_label2.setFixedWidth(80)

        # 第三行的5个复选框（自定义名称/宽度）
        self.radio_button_dy_task11 = QCheckBox("悟空浏览器")
        self.radio_button_dy_task11.setChecked(True)
        self.radio_button_dy_task11.setFixedWidth(100)
        self.radio_button_dy_task12 = QCheckBox("今日头条")
        self.radio_button_dy_task12.setChecked(True)
        self.radio_button_dy_task12.setFixedWidth(100)

        self.radio_button_dy_task13 = QCheckBox("西瓜视频")
        self.radio_button_dy_task13.setChecked(True)
        self.radio_button_dy_task13.setFixedWidth(100)




        # 添加第三行控件到布局
        self.horizontal_layout_dy_task3.addWidget(self.empty_label2)
        self.horizontal_layout_dy_task3.addWidget(self.radio_button_dy_task11)
        self.horizontal_layout_dy_task3.addWidget(self.radio_button_dy_task12)
        self.horizontal_layout_dy_task3.addWidget(self.radio_button_dy_task13)

        self.horizontal_layout_dy_task3.addStretch()


        #任务市场
        self.renwushichang_layout = QHBoxLayout()
        self.renwushichang_layout = QHBoxLayout()
        self.renwushichang_layout.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.renwushichang_layout.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben_renwushichang = QLabel('dy任务执行时长:')
        self.douyin_guanjianzi_wenben_renwushichang.setFixedWidth(90)

        douyinrenwuzhixingshichang = get_value_by_key_pkl("shuju_config.pkl", "douyinrenwuzhixingshichang")
        if (douyinrenwuzhixingshichang != None):
            self.douyinrenwuzhixingshichang = QLineEdit(douyinrenwuzhixingshichang)
        else:
            self.douyinrenwuzhixingshichang = QLineEdit("50")
        self.douyinrenwuzhixingshichang.setFixedWidth(40)

        self.baifenbi_2 = QLabel("分钟              ")

        self.renwushichang_layout.addWidget(self.douyin_guanjianzi_wenben_renwushichang)
        self.renwushichang_layout.addWidget(self.douyinrenwuzhixingshichang)
        self.renwushichang_layout.addWidget(self.baifenbi_2)

        # 搜索文件选择
        self.h_layout_dir = QHBoxLayout()
        self.label_file = QLabel("请选择配置文件大路径:")

        file_temp_path = get_value_by_key_pkl("shuju_config.pkl", "file_path")
        if (file_temp_path != None):
            self.file_textbox = QLineEdit(file_temp_path)
        else:
            self.file_textbox = QLineEdit("请选择配置文件大路径")
        self.file_textbox.setFixedWidth(400)
        self.file_button = QPushButton("选择文件", self)
        self.temp = QLabel("                          ")
        self.h_layout_dir.addWidget(self.label_file)
        self.h_layout_dir.addWidget(self.file_textbox)
        self.h_layout_dir.addWidget(self.file_button)
        #self.h_layout_dir.addWidget(self.temp)

        # 评论文件选择
        self.h_layout_dir_comment = QHBoxLayout()
        self.label_file_comment = QLabel("请选择回复文件路径:")

        file_temp_path_comment = get_value_by_key_pkl("shuju_config.pkl", "file_path_comment111")
        if (file_temp_path_comment != None):
            self.file_textbox_comment = QLineEdit(file_temp_path_comment)
        else:
            self.file_textbox_comment = QLineEdit("请输入回复文件路径")
        self.file_textbox_comment.setFixedWidth(400)
        self.file_button_comment = QPushButton("选择文件", self)
        self.temp_comment = QLabel("                          ")
        self.h_layout_dir_comment.addWidget(self.label_file_comment)
        self.h_layout_dir_comment.addWidget(self.file_textbox_comment)
        self.h_layout_dir_comment.addWidget(self.file_button_comment)
        #self.h_layout_dir_comment.addWidget(self.temp_comment)


        self.h_layout_diwuhang = QHBoxLayout()
        self.h_layout_diwuhang.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_diwuhang.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben = QLabel('抖音视频观看时长:')
        self.douyin_guanjianzi_wenben.setFixedWidth(110)

        douyinshipinguankanshichang_xiao = get_value_by_key_pkl("shuju_config.pkl", "douyinshipinguankanshichang_xiao")
        if (douyinshipinguankanshichang_xiao != None):
            self.douyinshipinguankanshichang_xiao = QLineEdit(douyinshipinguankanshichang_xiao)
        else:
            self.douyinshipinguankanshichang_xiao = QLineEdit("50")
        self.douyinshipinguankanshichang_xiao.setFixedWidth(40)
        self.baifenbi_1 = QLabel("至")

        douyinshipinguankanshichang_da = get_value_by_key_pkl("shuju_config.pkl", "douyinshipinguankanshichang_da")
        if (douyinshipinguankanshichang_da != None):
            self.douyinshipinguankanshichang_da = QLineEdit(douyinshipinguankanshichang_da)
        else:
            self.douyinshipinguankanshichang_da = QLineEdit("50")
        self.douyinshipinguankanshichang_da.setFixedWidth(40)

        self.baifenbi_2 = QLabel("秒              ")

        self.h_layout_diwuhang.addWidget(self.douyin_guanjianzi_wenben)
        self.h_layout_diwuhang.addWidget(self.douyinshipinguankanshichang_xiao)
        self.h_layout_diwuhang.addWidget(self.baifenbi_1)

        self.h_layout_diwuhang.addWidget(self.douyinshipinguankanshichang_da)
        self.h_layout_diwuhang.addWidget(self.baifenbi_2)

        #抖音视频推荐单次滑动次数
        self.douyin_guanjianzi_wenben_huadongcishu = QLabel('每次推荐滑动次数:')
        self.douyin_guanjianzi_wenben_huadongcishu.setFixedWidth(105)

        meicituijianhuadongcishu_xiao = get_value_by_key_pkl("shuju_config.pkl", "meicituijianhuadongcishu_xiao")
        if (meicituijianhuadongcishu_xiao != None):
            self.meicituijianhuadongcishu_xiao = QLineEdit(meicituijianhuadongcishu_xiao)
        else:
            self.meicituijianhuadongcishu_xiao = QLineEdit("50")
        self.meicituijianhuadongcishu_xiao.setFixedWidth(40)
        self.baifenbi_6 = QLabel("至")

        meicituijianhuadongcishu_da = get_value_by_key_pkl("shuju_config.pkl", "meicituijianhuadongcishu_da")
        if (meicituijianhuadongcishu_da != None):
            self.meicituijianhuadongcishu_da = QLineEdit(meicituijianhuadongcishu_da)
        else:
            self.meicituijianhuadongcishu_da = QLineEdit("50")
        self.meicituijianhuadongcishu_da.setFixedWidth(40)

        self.baifenbi_7 = QLabel("次")

        self.h_layout_diwuhang.addWidget(self.douyin_guanjianzi_wenben_huadongcishu)
        self.h_layout_diwuhang.addWidget(self.meicituijianhuadongcishu_xiao)
        self.h_layout_diwuhang.addWidget(self.baifenbi_6)

        self.h_layout_diwuhang.addWidget(self.meicituijianhuadongcishu_da)
        self.h_layout_diwuhang.addWidget(self.baifenbi_7)

        self.h_layout_diwuhang.addStretch(1)

        #同城相关配置
        self.h_layout_tongcheng = QHBoxLayout()
        self.h_layout_tongcheng.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_tongcheng.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben_tongcheng = QLabel('二维码1成功注册个数:')
        self.douyin_guanjianzi_wenben_tongcheng.setFixedWidth(130)

        tongchengguanjianzi = get_value_by_key_pkl("shuju_config.pkl", "tongchengguanjianzi")
        if (tongchengguanjianzi != None):
            self.tongchengguanjianzi = QLineEdit(tongchengguanjianzi)
        else:
            self.tongchengguanjianzi = QLineEdit("单位为分钟")
        self.tongchengguanjianzi.setFixedWidth(350)

        self.h_layout_tongcheng.addWidget(self.douyin_guanjianzi_wenben_tongcheng)
        self.h_layout_tongcheng.addWidget(self.tongchengguanjianzi)
        self.h_layout_tongcheng.addStretch(1)

        # 消息回复相关配置
        self.h_layout_xiaoxihuifu = QHBoxLayout()
        self.h_layout_xiaoxihuifu.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_xiaoxihuifu.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben_huifu = QLabel('二维码2成功注册个数:')
        self.douyin_guanjianzi_wenben_huifu.setFixedWidth(130)

        huifuxiaoxiyonghunicheng = get_value_by_key_pkl("shuju_config.pkl", "huifuxiaoxiyonghunicheng")
        if (huifuxiaoxiyonghunicheng != None):
            self.huifuxiaoxiyonghunicheng = QLineEdit(huifuxiaoxiyonghunicheng)
        else:
            self.huifuxiaoxiyonghunicheng = QLineEdit("")
        self.huifuxiaoxiyonghunicheng.setFixedWidth(350)

        self.h_layout_xiaoxihuifu.addWidget(self.douyin_guanjianzi_wenben_huifu)
        self.h_layout_xiaoxihuifu.addWidget(self.huifuxiaoxiyonghunicheng)
        self.h_layout_xiaoxihuifu.addStretch(1)

        self.h_layout_xiaoxihuifu_weixin_1 = QHBoxLayout()
        self.h_layout_xiaoxihuifu_weixin_1.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_xiaoxihuifu_weixin_1.setSpacing(0)  # 设置控件间距

        self.h_layout_xiaoxihuifu_1 = QHBoxLayout()
        self.h_layout_xiaoxihuifu_1.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_xiaoxihuifu_1.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben_huifu_1 = QLabel('二维码3成功注册个数:')
        self.douyin_guanjianzi_wenben_huifu_1.setFixedWidth(130)

        huifuxiaoxiyonghunicheng_1 = get_value_by_key_pkl("shuju_config.pkl", "huifuxiaoxiyonghunicheng_1")
        if (huifuxiaoxiyonghunicheng_1 != None):
            self.huifuxiaoxiyonghunicheng_1 = QLineEdit(huifuxiaoxiyonghunicheng_1)
        else:
            self.huifuxiaoxiyonghunicheng_1 = QLineEdit("")
        self.huifuxiaoxiyonghunicheng_1.setFixedWidth(350)

        self.h_layout_xiaoxihuifu_1.addWidget(self.douyin_guanjianzi_wenben_huifu_1)
        self.h_layout_xiaoxihuifu_1.addWidget(self.huifuxiaoxiyonghunicheng_1)
        self.h_layout_xiaoxihuifu_1.addStretch(1)

        self.h_layout_xiaoxihuifu_weixin = QHBoxLayout()
        self.h_layout_xiaoxihuifu_weixin.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_xiaoxihuifu_weixin.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben_huifu_weixin = QLabel('请选择微信:')
        self.douyin_guanjianzi_wenben_huifu_weixin.setFixedWidth(130)
        self.combo = QComboBox()
        self.combo.addItems(["微信1", "微信2"])
        self.combo.setFixedWidth(80)
        self.h_layout_xiaoxihuifu_weixin.addWidget(self.douyin_guanjianzi_wenben_huifu_weixin)
        self.h_layout_xiaoxihuifu_weixin.addWidget(self.combo)
        self.h_layout_xiaoxihuifu_weixin.addStretch(1)

        #分享用户配置
        self.h_layout_fenxianggei = QHBoxLayout()
        self.h_layout_fenxianggei.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_fenxianggei.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben_fenxiang = QLabel('分享用户昵称配置:')
        self.douyin_guanjianzi_wenben_fenxiang.setFixedWidth(110)

        fenxiangyonghunicheng = get_value_by_key_pkl("shuju_config.pkl", "fenxiangyonghunicheng")
        if (fenxiangyonghunicheng != None):
            self.fenxiangyonghunicheng = QLineEdit(fenxiangyonghunicheng)
        else:
            self.fenxiangyonghunicheng = QLineEdit("多个名称用'_'隔开")
        self.fenxiangyonghunicheng.setFixedWidth(330)

        self.h_layout_fenxianggei.addWidget(self.douyin_guanjianzi_wenben_fenxiang)
        self.h_layout_fenxianggei.addWidget(self.fenxiangyonghunicheng)
        self.h_layout_fenxianggei.addStretch(1)

        # 低于多少个不点赞，低于多少个不收藏
        self.h_layout_shoucang_dianzan_yuzhi = QHBoxLayout()
        self.h_layout_shoucang_dianzan_yuzhi.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_shoucang_dianzan_yuzhi.setSpacing(0)  # 设置控件间距

        self.douyin_guanjianzi_wenben_shoucang_dianzan_yuzhi = QLabel('点赞低于')
        self.douyin_guanjianzi_wenben_shoucang_dianzan_yuzhi.setFixedWidth(50)

        dianzanyuzhi = get_value_by_key_pkl("shuju_config.pkl", "dianzanyuzhi")
        if (dianzanyuzhi != None):
            self.dianzanyuzhi = QLineEdit(dianzanyuzhi)
        else:
            self.dianzanyuzhi = QLineEdit("100")
        self.dianzanyuzhi.setFixedWidth(50)

        self.douyin_guanjianzi_wenben_shoucang_dianzan_yuzhi_budianzan = QLabel('不点赞')
        self.douyin_guanjianzi_wenben_shoucang_dianzan_yuzhi_budianzan.setFixedWidth(50)

        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.douyin_guanjianzi_wenben_shoucang_dianzan_yuzhi)
        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.dianzanyuzhi)
        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.douyin_guanjianzi_wenben_shoucang_dianzan_yuzhi_budianzan)


        self.douyin_guanjianzi_wenben_shoucang_yuzhi_shoucang = QLabel('收藏低于')
        self.douyin_guanjianzi_wenben_shoucang_yuzhi_shoucang.setFixedWidth(50)

        shoucangyuzhi = get_value_by_key_pkl("shuju_config.pkl", "shoucangyuzhi")
        if (shoucangyuzhi != None):
            self.shoucangyuzhi = QLineEdit(shoucangyuzhi)
        else:
            self.shoucangyuzhi = QLineEdit("100")
        self.shoucangyuzhi.setFixedWidth(50)

        self.douyin_guanjianzi_wenben_shoucang_yuzhi_budianzan = QLabel('不收藏')
        self.douyin_guanjianzi_wenben_shoucang_yuzhi_budianzan.setFixedWidth(50)

        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.douyin_guanjianzi_wenben_shoucang_yuzhi_shoucang)
        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.shoucangyuzhi)
        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.douyin_guanjianzi_wenben_shoucang_yuzhi_budianzan)


        self.douyin_guanjianzi_wenben_shoucang_yuzhi_pinglun = QLabel('评论低于')
        self.douyin_guanjianzi_wenben_shoucang_yuzhi_pinglun.setFixedWidth(50)

        pinglunyuzhi = get_value_by_key_pkl("shuju_config.pkl", "pinglunyuzhi")
        if (pinglunyuzhi != None):
            self.pinglunyuzhi = QLineEdit(pinglunyuzhi)
        else:
            self.pinglunyuzhi = QLineEdit("100")
        self.pinglunyuzhi.setFixedWidth(50)

        self.douyin_guanjianzi_wenben_punglun_yuzhi_budianzan = QLabel('不评论')
        self.douyin_guanjianzi_wenben_punglun_yuzhi_budianzan.setFixedWidth(50)

        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.douyin_guanjianzi_wenben_shoucang_yuzhi_pinglun)
        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.pinglunyuzhi)
        self.h_layout_shoucang_dianzan_yuzhi.addWidget(self.douyin_guanjianzi_wenben_punglun_yuzhi_budianzan)


        self.h_layout_shoucang_dianzan_yuzhi.addStretch(1)


        #点赞 收藏 评论 概率
        self.h_layout_gailv = QHBoxLayout()
        self.h_layout_gailv.setContentsMargins(1, 0, 0, 0)  # 设置布局边距
        self.h_layout_gailv.setSpacing(0)  # 设置控件间距
        self.h_layout_gailv.setAlignment(Qt.AlignmentFlag.AlignLeft)


        self.douyin_dianzan_gailv = QLabel('点赞概率')
        self.douyin_dianzan_gailv.setFixedWidth(50)

        dianzan_gailv = get_value_by_key_pkl("shuju_config.pkl", "dianzan_gailv")
        if (dianzan_gailv != None):
            self.dianzan_gailv = QLineEdit(dianzan_gailv)
        else:
            self.dianzan_gailv = QLineEdit("50")
        self.dianzan_gailv.setFixedWidth(50)

        self.douyin_dianzan_gailv_baifenhao = QLabel('%')
        self.douyin_dianzan_gailv_baifenhao.setFixedWidth(50)
        self.h_layout_gailv.addWidget(self.douyin_dianzan_gailv)
        self.h_layout_gailv.addWidget(self.dianzan_gailv)
        self.h_layout_gailv.addWidget(self.douyin_dianzan_gailv_baifenhao)

        self.douyin_shoucang_gailv = QLabel('收藏概率')
        self.douyin_shoucang_gailv.setFixedWidth(50)

        shoucang_gailv = get_value_by_key_pkl("shuju_config.pkl", "shoucang_gailv")
        if (shoucang_gailv != None):
            self.shoucang_gailv = QLineEdit(shoucang_gailv)
        else:
            self.shoucang_gailv = QLineEdit("50")
        self.shoucang_gailv.setFixedWidth(50)

        self.douyin_dianzan_gailv_baifenhao1 = QLabel('%')
        self.douyin_dianzan_gailv_baifenhao1.setFixedWidth(50)
        self.h_layout_gailv.addWidget(self.douyin_shoucang_gailv)
        self.h_layout_gailv.addWidget(self.shoucang_gailv)
        self.h_layout_gailv.addWidget(self.douyin_dianzan_gailv_baifenhao1)

        self.douyin_pinglun_gailv = QLabel('评论概率')
        self.douyin_pinglun_gailv.setFixedWidth(50)

        pinglun_gailv = get_value_by_key_pkl("shuju_config.pkl", "pinglun_gailv")
        if (pinglun_gailv != None):
            self.pinglun_gailv = QLineEdit(pinglun_gailv)
        else:
            self.pinglun_gailv = QLineEdit("50")
        self.pinglun_gailv.setFixedWidth(50)

        self.douyin_dianzan_gailv_baifenhao2 = QLabel('%')
        self.douyin_dianzan_gailv_baifenhao2.setFixedWidth(50)
        self.h_layout_gailv.addWidget(self.douyin_pinglun_gailv)
        self.h_layout_gailv.addWidget(self.pinglun_gailv)
        self.h_layout_gailv.addWidget(self.douyin_dianzan_gailv_baifenhao2)

        self.douyin_fenxiang_gailv = QLabel('分享概率')
        self.douyin_fenxiang_gailv.setFixedWidth(50)

        fenxiang_gailv = get_value_by_key_pkl("shuju_config.pkl", "fenxiang_gailv")
        if (fenxiang_gailv != None):
            self.fenxiang_gailv = QLineEdit(fenxiang_gailv)
        else:
            self.fenxiang_gailv = QLineEdit("50")
        self.fenxiang_gailv.setFixedWidth(50)

        self.douyin_dianzan_gailv_baifenhao3 = QLabel('%')
        self.douyin_dianzan_gailv_baifenhao3.setFixedWidth(50)
        self.h_layout_gailv.addWidget(self.douyin_fenxiang_gailv)
        self.h_layout_gailv.addWidget(self.fenxiang_gailv)
        self.h_layout_gailv.addWidget(self.douyin_dianzan_gailv_baifenhao3)


        # layout.addLayout(self.horizontal_layout_dy_task1)
        # layout.addLayout(self.horizontal_layout_dy_task2)
        #layout.addLayout(self.horizontal_layout_dy_task3)
        #layout.addLayout(self.renwushichang_layout)
        layout.addLayout(self.h_layout_dir)
        #layout.addLayout(self.h_layout_dir_comment)
        #layout.addLayout(self.h_layout_diwuhang)
        layout.addLayout(self.h_layout_tongcheng)
        layout.addLayout(self.h_layout_xiaoxihuifu)
        layout.addLayout(self.h_layout_xiaoxihuifu_1)
        layout.addLayout(self.h_layout_xiaoxihuifu_weixin)
        #layout.addLayout(self.h_layout_fenxianggei)
        #layout.addLayout(self.h_layout_shoucang_dianzan_yuzhi)
        #layout.addLayout(self.h_layout_gailv)
        # layout.addLayout(self.h_layout)
        layout.addStretch()  # 底部留白

        # 绑定文件选择事件
        self.file_button.clicked.connect(self.on_file_button_clicked)
        self.file_button_comment.clicked.connect(self.showDialog_comment)

        self.tab_widget.addTab(douyin_tab, "设置")


    def init_advanced_tab(self):
        """高级设置Tab页"""
        advanced_tab = QWidget()
        layout = QVBoxLayout

    def on_item_changed(self,item: QTableWidgetItem):
        if self.table_widget.currentColumn() == 2:
            # 获取新的数据并打印（或保存到其他地方）
            new_data = item.text()
            #print(item.column())

            item.row()
            #print(self.table_widget.item(item.row(),1).text())
            phone_name = self.table_widget.item(item.row(),1).text()
            #print(f"New data in row {item.column()}, column 2: {new_data}")
            # 你可以在这里添加保存数据的逻辑，比如保存到数据库或文件中
            updata_pkl_config("config.pkl",phone_name,new_data)
            #print(pkl_list("config.pkl"))
    def on_item_clicked(self):

        current_devices = get_connected_devices()
        current_device_ids = {device[0] for device in current_devices}
        self.selected_ids = current_device_ids
        self.refresh_pkl_files()
        print(self.selected_ids)
        print("全选")

    def shell_neibu(self,cmd):
        os.system(cmd)

    def on_file_button_clicked(self):
        folder_path = QFileDialog.getExistingDirectory(self, 'Select Folder')
        if folder_path:
            # If a folder is selected, update the QLabel
            self.file_textbox.setText(folder_path)
            updata_pkl_config_mianban("file_path",folder_path)

            # self.file_textbox.setText(selected_file)
            # updata_pkl_config_mianban("file_path", selected_file)
        else:
            self.file_textbox.setText('No folder selected')

    def showDialog(self):
        # 设置文件过滤器
        print("1")
        filters = "All Files (*)"

        # 创建文件对话框
        dialog = QFileDialog(self, "Open Excel File", "", filters)
        dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        dialog.setDefaultSuffix("txt")
        dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptOpen)
        # 显示对话框并获取用户选择
        if dialog.exec() == QFileDialog.DialogCode.Accepted:
            selected_files = dialog.selectedFiles()
            if selected_files:
                selected_file = selected_files[0]  # 获取第一个选中的文件（假设只选择一个文件）
                self.file_textbox.setText(selected_file)
                updata_pkl_config_mianban("file_path", selected_file)
                #self.excel_file = selected_file
                #self.import_config()
                #self.refresh_pkl_files_test()
    def showDialog_comment(self):
        # 设置文件过滤器
        print("1")
        filters = "Excel Files (*.txt);;All Files (*)"

        # 创建文件对话框
        dialog = QFileDialog(self, "Open Excel File", "", filters)
        dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        dialog.setDefaultSuffix("txt")
        dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptOpen)
        # 显示对话框并获取用户选择
        if dialog.exec() == QFileDialog.DialogCode.Accepted:
            selected_files = dialog.selectedFiles()
            if selected_files:
                selected_file = selected_files[0]  # 获取第一个选中的文件（假设只选择一个文件）
                self.file_textbox_comment.setText(selected_file)
                updata_pkl_config_mianban("file_path_comment", selected_file)
                #self.excel_file = selected_file
                #self.import_config()
                #self.refresh_pkl_files_test()
    def import_config(self):
        print("")
        path_dir = "task_config"
        create_directory_if_not_exists(path_dir)

        # path = path_dir+"/"+"config.pkl"
        # self.judge_pkl_creat(path)
        # 指定文件的路径
        with video_lock:
            file_path = self.file_textbox.text()
            # 使用 with 语句打开文件，这样可以确保文件在读取完毕后自动关闭
            with open(file_path, 'r', encoding='utf-8') as file:
                # 逐行读取文件内容并打印
                for line in file:
                    print(line, end='')  # 使用 end='' 是为了避免打印每行末尾的额外换行符
                    if((str(line).count("_")>0) and (str(line).count("/")>0)):
                        file_name = str(line).split("/")[-2]
                        new_data = {"url":str(line).split("_")[0],"BIG_COUNT":int(str(line).split("_")[1]),"TONGJI":0}
                        print("newdata=",new_data)
                        file_name = path_dir + "/" +file_name +".pkl"
                        print("file_name=", file_name)
                        self.judge_pkl_creat(file_name,new_data)
        # 注意：使用 with 语句后，不需要手动关闭文件，它会在块结束时自动关闭

    def judge_pkl_creat(self,pkl_file,new_data):
        # 指定文件的路径
        file_path = pkl_file
        # 检查文件是否存在
        if not os.path.exists(file_path):
            # 如果文件不存在，则创建一个新的字典
            new_data = new_data
            # 使用 with 语句打开（或创建）文件，并写入数据
            with open(file_path, 'wb') as file:
                pickle.dump(new_data, file)

            print(f"文件 {file_path} 已创建并写入新数据。")
        else:
            print(f"文件 {file_path} 已存在。")

        # 注意：'wb' 模式用于以二进制写入方式打开文件，这是 pickle 所需的。
    def updata_pkl_config_video(self,pklfile, key, value):
        # dic = {}
        if not os.path.exists(pklfile):
            # 如果文件不存在，创建一个新的字典（或其他对象）
            data = {key: value}  # 这里可以替换为你想要保存的任何Python对象
            # 使用pickle将对象序列化并保存到文件中
            with open(pklfile, 'wb') as file:
                pickle.dump(data, file)
        else:
            with open(pklfile, 'rb') as pkl_file:
                dic = pickle.load(pkl_file)
                # print("li----------------",dic)
            dic[key] = value
            # print("----------------------------------",dic)
            with open(pklfile, 'wb') as pkl_file:
                pickle.dump(dic, pkl_file)


    def save_config(self):
        try:

            updata_pkl_config_mianban("douyinrenwuzhixingshichang", self.douyinrenwuzhixingshichang.text())
            updata_pkl_config_mianban("douyinshipinguankanshichang_xiao", self.douyinshipinguankanshichang_xiao.text())
            updata_pkl_config_mianban("douyinshipinguankanshichang_da", self.douyinshipinguankanshichang_da.text())
            updata_pkl_config_mianban("meicituijianhuadongcishu_xiao", self.meicituijianhuadongcishu_xiao.text())
            updata_pkl_config_mianban("meicituijianhuadongcishu_da", self.meicituijianhuadongcishu_da.text())
            updata_pkl_config_mianban("tongchengguanjianzi", self.tongchengguanjianzi.text())
            updata_pkl_config_mianban("dianzanyuzhi", self.dianzanyuzhi.text())
            updata_pkl_config_mianban("shoucangyuzhi", self.shoucangyuzhi.text())
            updata_pkl_config_mianban("dianzan_gailv", self.dianzan_gailv.text())
            updata_pkl_config_mianban("shoucang_gailv", self.shoucang_gailv.text())
            updata_pkl_config_mianban("pinglun_gailv", self.pinglun_gailv.text())
            updata_pkl_config_mianban("fenxiang_gailv", self.fenxiang_gailv.text())
            updata_pkl_config_mianban("pinglunyuzhi", self.pinglunyuzhi.text())
            updata_pkl_config_mianban("fenxiangyonghunicheng", self.fenxiangyonghunicheng.text())
            updata_pkl_config_mianban("huifuxiaoxiyonghunicheng", self.huifuxiaoxiyonghunicheng.text())
            updata_pkl_config_mianban("weixin_count", self.combo.currentText())
            task = ""
            # if (self.radio_button_dy_task6.isChecked()):
            #     task = "task-suiji"
            # else:
            if (self.radio_button_dy_task1.isChecked()):
                task = task + "_" + "七猫免费小说"
            if (self.radio_button_dy_task2.isChecked()):
                task = task + "_" + "河马剧场"
            if (self.radio_button_dy_task3.isChecked()):
                task = task + "_" + "得间免费小说"
            if (self.radio_button_dy_task4.isChecked()):
                task = task + "_" + "酷安"
            if (self.radio_button_dy_task5.isChecked()):
                task = task + "_" + "韩小圈"

            if (self.radio_button_dy_task6.isChecked()):
                task = task + "_" + "WiFi万能钥匙"
            if (self.radio_button_dy_task7.isChecked()):
                task = task + "_" + "抖音APP"
            # if (self.radio_button_dy_task8.isChecked()):
            #     task = task + "_" + "红果免费漫剧"
            # if (self.radio_button_dy_task9.isChecked()):
            #     task = task + "_" + "番茄音乐"
            # if (self.radio_button_dy_task10.isChecked()):
            #     task = task + "_" + "抖音火山版"
            #
            # if (self.radio_button_dy_task11.isChecked()):
            #     task = task + "_" + "悟空浏览器"
            # if (self.radio_button_dy_task12.isChecked()):
            #     task = task + "_" + "今日头条"
            # if (self.radio_button_dy_task13.isChecked()):
            #     task = task + "_" + "西瓜视频"
            if(self.radio_button3.isChecked()):
                updata_pkl_config_mianban("shifouguanbidouyin", "1")
            else:
                updata_pkl_config_mianban("shifouguanbidouyin", "0")
            updata_pkl_config_mianban("task_app", task)
            print("task=",task)
            if(task == ""):
                print("请选择APP签到")
                return


            task = ""
            if (self.radio_button1.isChecked()):
                task = "douyinyanghao"
            # if (self.radio_button2.isChecked()):
            #     task = "xiaohongshuyanghao"
            updata_pkl_config_mianban("task", task)

        except BaseException as e:
            print(f"发生崩溃了: {e}")
            error_info = traceback.format_exc()
            print("完整错误信息:")
            print(error_info)



    def execute_button_clicked(self):

        self.save_config()
        if(self.selected_ids == []):
            print("当前没有选择手机")
            return
        for temp in self.selected_ids:
            #print(temp)
            updata_pkl("./shuju/"+temp+".pkl","执行状态","运行中")

        self.refresh_pkl_files()
        tasks = []
        print("tasks------------",tasks)
        # if(os.path.isfile(self.file_textbox.text())):
        #     print("搜索文件加载")
        # else:
        #     print("搜索文件buzai")
        #     return
        # if (os.path.isfile(self.file_textbox_comment.text())):
        #     print("评论文件加载")
        # else:
        #     print("评论文件不在")
            #return
        for serial in self.selected_ids:
            thread = threading.Thread(target=operate_device, args=(serial,))
            #搜索路径、评论文件路径、任务列表、运行时长、更换频率小、更换频率大、视频滑动间隔小、视频滑动间隔大、视频滑动次数、收藏、评论、点赞、关注
            #threads.append(thread)
            thread.start()

        self.selected_ids = []

    def execute_delete_button_clicked(self):
        print("---------------")
        # if(self.selected_ids == []):
        #     return
        # for temp in self.selected_ids:
        #     #print(temp)
        #     if(os.path.isfile("./shuju/" + temp + ".pkl")):
        #         os.remove("./shuju/" + temp + ".pkl")
        # self.refresh_pkl_files()
        # self.selected_ids = []

        threading.Thread(target=self.thread_temp,args=(self.file_textbox.text(),)).start()

    def clear_xlsx_data(self,root_dir: str):
        from openpyxl import load_workbook
        max_depth = 2  # 只往下找2级目录
        for current_root, dirs, files in os.walk(root_dir):
            depth = current_root.count(os.sep) - root_dir.count(os.sep)
            if depth > max_depth:
                continue

            for file in files:
                if file.lower().endswith(".xlsx") and not file.startswith("~$"):
                    file_path = os.path.join(current_root, file)
                    try:
                        wb = load_workbook(file_path)
                        for ws in wb.worksheets:
                            max_row = ws.max_row
                            if max_row <= 1:
                                continue
                            # ==============================================
                            # 真正删除：从第2行删到最后一行！
                            # ==============================================
                            ws.delete_rows(2, max_row - 1)
                        wb.save(file_path)
                        wb.close()
                        print(f"✅ 已删除数据：{file_path}")
                    except Exception as e:
                        print(f"❌ 失败：{file_path} | {e}")

    def thread_temp(self,filePath):
        thread_list = []
        self.execute_button.setEnabled(False)
        if(not os.path.isdir(filePath)):
            print("当前目录配置错误")
            return
        self.clear_xlsx_data(filePath)
        self.execute_button.setEnabled(True)
    def execute_reset_button_clicked(self):
        #print("execute_reset_button_clicked")
        directory = './shuju'
        for filename in sorted(os.listdir(directory)):
            if filename.endswith('.pkl'):
                filepath = os.path.join(directory, filename)
                updata_pkl(filepath, "执行状态", "运行结束")
                updata_pkl(filepath, "进行的任务", "空闲")
        self.refresh_pkl_files()

    def execute_save_button_clicked(self):
        print("baoc")
        self.save_config()

    def refresh_pkl_files(self):
        # 保存当前滚动位置
        current_pos = self.table_widget.verticalScrollBar().value()

        #print("current_scroll_position",current_scroll_position)
        # 清除旧数据
        self.table_widget.setRowCount(0)
        # 遍历目录中的所有文件
        directory = './shuju'
        row_index = 0

        sorted_data = dict(sorted(pkl_list("config.pkl").items(), key=lambda item: item[1]))
        #print("sorted_data=",sorted_data)

        for device_id,v in sorted_data.items():
            #print("device_id---->",device_id)
            file_name = directory+"/"+str(device_id)+".pkl"
            #print("file_name---",file_name)
            if(os.path.isfile(file_name)):
                try:
                    with open(file_name, 'rb') as file:
                        data = pickle.load(file)

                        #print("data-----------,",data)

                        # 假设数据是一个字典
                        if isinstance(data, dict):
                            # 插入新行
                            self.table_widget.insertRow(row_index)
                            # 添加复选框
                            checkbox = QCheckBox(self)
                            #print("self.selected_ids=",self.selected_ids)
                            #print("os.path.splitext(file_name)[0]",os.path.splitext(file_name)[0].split("/")[2])
                            if os.path.splitext(file_name)[0].split("/")[2] in self.selected_ids:
                                checkbox.setChecked(True)
                            if(data.get('执行状态', 'N/A') == "运行中"):
                                checkbox.setEnabled(False)
                            else:
                                checkbox.setEnabled(True)
                            # if (data.get('连接状态', 'N/A') == "中断连接"):
                            #     checkbox.setEnabled(True)
                            # else:
                            #     checkbox.setEnabled(True)
                            checkbox.stateChanged.connect(lambda state, row=row_index: self.update_selected_ids(state, row))
                            self.table_widget.setCellWidget(row_index, 0, checkbox)
                            # 设置文件名（去除后缀）
                            self.table_widget.setItem(row_index, 1, QTableWidgetItem(device_id))
                            # 设置其他数据
                            phone_name = get_value_by_key_pkl("config.pkl",data.get('name', 'N/A'))
                            if(phone_name != None):
                                item_i = QTableWidgetItem(phone_name)
                            else:
                                item_i = QTableWidgetItem(data.get('nick_name', 'N/A'))
                            #item_i.setForeground(QBrush(QColor(255,0,0)))

                            self.table_widget.setItem(row_index, 2, item_i)
                            if(data.get('连接状态', 'N/A') == "中断连接"):
                                item_lianjie = QTableWidgetItem(data.get('连接状态', 'N/A'))
                                item_lianjie.setForeground(QBrush(QColor(255,0,0)))
                            else:
                                item_lianjie = QTableWidgetItem(data.get('连接状态', 'N/A'))
                                item_lianjie.setForeground(QBrush(QColor(0, 0, 0)))
                            item_lianjie.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            self.table_widget.setItem(row_index, 3, item_lianjie)

                            #item_i = QTableWidgetItem(data.get('执行状态', 'N/A'))
                            # item_i.setForeground(QBrush(QColor(255,0,0)))
                            item_zhuangtai = QTableWidgetItem(data.get('执行状态', 'N/A'))
                            item_zhuangtai.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            self.table_widget.setItem(row_index, 4, item_zhuangtai)

                            # item_zhuangage = QTableWidgetItem(data.get('age', 'N/A'))
                            # item_zhuangage.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            # self.table_widget.setItem(row_index, 5, item_zhuangage)
                            #
                            # item_add = QTableWidgetItem(data.get('add', 'N/A'))
                            # item_add.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            # self.table_widget.setItem(row_index, 6, item_add)

                            item_renwu = QTableWidgetItem(data.get('进行的任务', 'N/A'))
                            item_renwu.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            self.table_widget.setItem(row_index, 5, item_renwu)

                            button111 = QTableWidgetItem(data.get('tongji', 'N/A'))
                            # button111.clicked.connect(lambda: print("Button clicked!"))
                            self.table_widget.setItem(row_index, 6,button111)

                            row_index += 1
                except Exception as e:
                    print(f"读取文件2 {file_name} 时出错: {e}")
         # 恢复滚动位置
        self.table_widget.verticalScrollBar().setSliderPosition(current_pos)

    def update_selected_ids(self, state, row):
        # 更新选中的编号
        item_id = self.table_widget.item(row, 1).text()  # 获取编号
        #print("item_id=",item_id)
        if item_id not in self.selected_ids:
            self.selected_ids.append(item_id)  # 添加到选中的编号
        else:
            if item_id in self.selected_ids:
                self.selected_ids.remove(item_id)  # 从选中的编号中移除
        # 打印当前选中的编号
        #print("当前选中的编号:", self.selected_ids)

def judge():
    shebeima = get_real_device_id()
    final_str = encrypt_and_modify(shebeima)
    if (os.path.isfile(final_str)):
        return True
    else:
        return False
def encrypt_and_modify(shebeima):
    """对输入的字符串进行Base64编码，并在特定位置插入字符"""
    input_text = shebeima

    if not input_text:
        return

    # 进行Base64编码
    encoded_bytes = base64.b64encode(input_text.encode('utf-8'))
    encoded_str = encoded_bytes.decode('utf-8')

    # 在特定位置插入字符
    modified_str = list(encoded_str)

    # 确保字符串足够长，以避免索引错误
    if len(modified_str) >= 1:
        modified_str[0] += 'a'
    if len(modified_str) >= 3:
        modified_str[2] += 'b'
    if len(modified_str) >= 5:
        modified_str[4] += 'f'
    if len(modified_str) >= 2:
        modified_str[-2] += 'g'

    # 将列表转换回字符串
    final_str = ''.join(modified_str)
    return final_str

def get_real_device_id():
    """获取更真实的设备唯一标识，并返回缩短后的版本"""
    try:
        # 收集各种硬件和系统信息
        info = [
            platform.node(),  # 计算机名
            platform.machine(),  # 机器类型
            platform.processor(),  # 处理器信息
            platform.system(),  # 操作系统名称
            platform.release(),  # 操作系统版本
            str(os.environ.get('COMPUTERNAME', '')),  # Windows计算机名
            str(os.environ.get('USERNAME', '')),  # 用户名
            str(uuid.getnode()),  # MAC地址
        ]

        # 创建哈希作为设备ID
        hash_obj = hashlib.sha256()
        hash_obj.update(''.join(info).encode('utf-8'))
        full_hash = hash_obj.hexdigest()

        # 返回缩短后的唯一码（例如前8个字符）
        return full_hash[:18]  # 取前8个字符作为缩短的唯一码
    except Exception as e:
        return f"ERR-{str(e)[:18]}"  # 错误情况下也返回缩短的字符串


def get_format_time():
    import time
    from datetime import datetime
    # 获取当前时间的时间戳
    timestamp = time.time()
    # 将时间戳转换为datetime对象
    current_time = datetime.fromtimestamp(timestamp)
    # 格式化datetime对象为字符串
    formatted_date = current_time.strftime("%Y-%m-%d-%H:%M:%S")
    return formatted_date


def pkl_add(pkl,dic):
    with open(pkl, 'wb') as pkl_file:
        pickle.dump(dic, pkl_file)
lock111 = threading.Lock()
def pkl_list(pklfile):
    # 使用with语句自动管理锁的获取和释放
    try:
        with lock111:
            if(not os.path.exists(pklfile)):
                data = {}
                with open(pklfile, 'wb') as file:
                    pickle.dump(data, file)
            if pklfile == "log.pkl":
                if os.path.isfile(pklfile):
                    print()  # 这里只是打印了一个空行，可能需要根据实际需求修改
                else:
                    data = {}
                    with open(pklfile, 'wb') as file:
                        pickle.dump(data, file)

                        # 读取文件
            with open(pklfile, 'rb') as pkl_file:
                my_object111 = pickle.load(pkl_file)
                return my_object111
    except BaseException:
        data = {}
        with open(pklfile, 'wb') as file:
            pickle.dump(data, file)

# 重新序列化对象
lock222 = threading.Lock()
def get_value_by_key_pkl(pklfile, key):
    # 使用with语句自动管理锁的获取和释放
    with lock222:
        if not os.path.isfile(pklfile):
            return None

        dic = {}
        try:
            with open(pklfile, 'rb') as pkl_file:
                dic = pickle.load(pkl_file)
        except (EOFError, pickle.PickleError):
            # 处理文件为空或损坏的情况
            print(f"Warning: Unable to load pickle file {pklfile}")
            return None

            # 检查键是否在字典中
        if key in dic:
            return dic[key]
        else:
            return None

def updata_pkl_config(pklfile,key,value):
    #dic = {}
    if not os.path.exists(pklfile):
        # 如果文件不存在，创建一个新的字典（或其他对象）
        data = {key:value}  # 这里可以替换为你想要保存的任何Python对象
        # 使用pickle将对象序列化并保存到文件中
        with open(pklfile, 'wb') as file:
            pickle.dump(data, file)
    else:
        with open(pklfile, 'rb') as pkl_file:
            dic = pickle.load(pkl_file)
            #print("li----------------",dic)
        dic[key] = value
        #print("----------------------------------",dic)
        with open(pklfile, 'wb') as pkl_file:
            pickle.dump(dic, pkl_file)

def updata_pkl(pklfile,key,value):
    #dic = {}
    if(os.path.isfile(pklfile)):
        with open(pklfile, 'rb') as pkl_file:
            dic = pickle.load(pkl_file)
        dic[key] = value
        #print("----------------------------------",dic)
        with open(pklfile, 'wb') as pkl_file:
            pickle.dump(dic, pkl_file)

import subprocess
import time
def get_connected_devices():
    # Run the adb devices command
    result = subprocess.run(['adb', 'devices'], capture_output=True, text=True)
    devices = result.stdout.strip().split('\n')[1:]  # Skip the first line (header)
    #print("连接的设备有。。。",devices)
    connected_devices = []
    for device in devices:
        if device.strip():
            device_info = device.split('\t')
            connected_devices.append((device_info[0], device_info[1]))  # (device_id, status)

    return connected_devices

def updata_pkl_config_mianban(key,value):
    pklfile = "shuju_config.pkl"
    if not os.path.isfile(pklfile):
        my_dict = {"file_path": "请输入文件夹路径"}
        # If it doesn't exist, create the file
        with open(pklfile, 'wb') as pkl_file:
            pickle.dump(my_dict, pkl_file)
    dic = {}
    with open(pklfile, 'rb') as pkl_file:
        dic = pickle.load(pkl_file)
    dic[key]=value
    with open(pklfile, 'wb') as pkl_file:
        pickle.dump(dic, pkl_file)
def monitor_devices():
    known_devices = set()
    create_directory_if_not_exists("shuju")
    delete_directory_contents("shuju")
    if(os.path.exists("config.pkl")):
        print()

    while True:
        current_devices = get_connected_devices()
        current_device_ids = {device[0] for device in current_devices}

        # Check for new connections
        new_devices = current_device_ids - known_devices
        for device_id in new_devices:
            #print(f"Device connected: {device_id}")
            dic = {"name":device_id,"连接状态":"已连接","执行状态":"空闲中","age":"1811","add":"bj1","xingbie":"nan","进行的任务":"空闲","nick_name":"昵称点击可编辑","tongji":"0"}
            pkl_add("./shuju/"+device_id+".pkl",dic)
            if(device_id not in dict(sorted(pkl_list("config.pkl").items(), key=lambda item: item[1]))):
                updata_pkl_config("config.pkl", device_id,"昵称点击可编辑")
        # Check for disconnections
        disconnected_devices = known_devices - current_device_ids
        for device_id in disconnected_devices:
            #print(f"Device disconnected: {device_id}")
            updata_pkl("./shuju/"+device_id+".pkl","连接状态","中断连接")
        # Update the known devices set
        known_devices = current_device_ids

        time.sleep(15)  # Check every 5 seconds
def delete_directory_contents(directory):
    shutil.rmtree(directory)
    os.makedirs(directory)  # 重新创建空文件夹


import os
import pickle


    # 示例用法
if __name__ == "__main__":
    thread = threading.Thread(target=monitor_devices)
    thread.start()
    app = QApplication(sys.argv)
    viewer = PklViewer()
    viewer.show()
    sys.exit(app.exec())
