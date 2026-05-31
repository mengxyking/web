# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('e:/pythonProject/douyin_hotlist/微博城市热搜_20260531_131027.xlsx')
ws = wb.active

print(f'Sheet: {ws.title}')
print(f'总行数 (含表头): {ws.max_row}')
print(f'总列数: {ws.max_column}')
print()

# 检查空行
empty_rows = []
partial_empty = []  # 部分字段为空
for row_idx in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
    non_empty = [v for v in row_vals if v is not None and str(v).strip() != '']
    if not non_empty:
        empty_rows.append(row_idx)
    else:
        # 检查关键字段是否为空（标题=col5, 对应内容=col10, 发布时间=col11）
        title = row_vals[4]  # col5
        content = row_vals[9]  # col10
        pubtime = row_vals[10]  # col11
        if not title or not content or not pubtime:
            partial_empty.append({
                'row': row_idx,
                'city': row_vals[3],
                'rank': row_vals[5],
                'title': title,
                'content': str(content)[:30] if content else '',
                'pubtime': pubtime,
            })

print(f'完全空行: {len(empty_rows)} 行')
if empty_rows:
    print(f'  行号: {empty_rows[:10]}')

print(f'关键字段部分为空: {len(partial_empty)} 行')
for r in partial_empty[:10]:
    print(f"  行{r['row']}: 城市={r['city']} 排名={r['rank']} 标题={r['title']} 内容={r['content']} 发布时间={r['pubtime']}")

# 打印表头
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print(f'\n表头: {headers}')
