# 先安装依赖：pip install openpyxl
from openpyxl import load_workbook


# ====================== 1. 读取 Excel 表格 ======================
def read_excel(file_path):
    """读取 Excel 所有数据"""
    wb = load_workbook(file_path)
    ws = wb.active  # 打开第一个工作表

    print("📄 当前 Excel 所有数据：")
    for row in ws.iter_rows(values_only=True):
        print(row)

    wb.close()


# ====================== 2. 添加一行数据 ======================
def add_row(file_path, new_data):
    """在最后一行追加一条数据"""
    wb = load_workbook(file_path)
    ws = wb.active

    ws.append(new_data)  # 添加一行
    wb.save(file_path)
    wb.close()
    print(f"✅ 已添加数据：{new_data}")


# ====================== 3. 根据 行、列 修改数据 ======================
def modify_cell(file_path, row_num, col_num, new_value):
    """
    修改指定单元格
    row_num: 行号（从 1 开始）
    col_num: 列号（从 1 开始）
    """
    wb = load_workbook(file_path)
    ws = wb.active

    ws.cell(row=row_num, column=col_num, value=new_value)
    wb.save(file_path)
    wb.close()
    print(f"✅ 已修改 第{row_num}行 第{col_num}列 为：{new_value}")


# ====================== 使用示例 ======================
if __name__ == '__main__':
    excel_file = "data.xlsx"  # 你的 Excel 文件名

    # 1. 读取
    read_excel(excel_file)

    # 2. 添加一行（你可以改成自己的数据）
    # new_row = ["张三", 25, "工程师", 8000]
    # add_row(excel_file, new_row)

    # 3. 修改数据：第3行 第2列 → 改成 30
    modify_cell(excel_file, row_num=3, col_num=2, new_value=30)

    # 再读一遍看结果
    print("\n📊 修改后数据：")
    read_excel(excel_file)