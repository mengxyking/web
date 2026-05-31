"""
抖音城市热搜 全量采集 → Excel
==========================================================
核心接口（城市热搜）：
  GET https://so-landing.douyin.com/aweme/v1/hot/search/list/
      ?board_type=1&board_sub_type={城市行政区划代码}&detail_list=1

  返回当前城市热搜列表
  每条数据字段：
    word=话题标题, position=排名, hot_value=热度,
    view_count=浏览量, video_count=视频数,
    event_time=事件时间(Unix), word_cover=封面图

视频搜索接口（可选，需要 Cookie）：
  GET https://www.douyin.com/aweme/v1/web/aweme/search/
      ?keyword={keyword}&search_source=hot_search_board_web

Cookie 自动从 Chrome 读取（rookiepy），Chrome 保持登录 douyin.com 即可。
无 Cookie 时：热搜列表 + 封面图正常采集，视频内容留空（基础模式）。
"""

import urllib.parse
import json
import os
import random
import re
import time
import uuid
from datetime import datetime

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import rookiepy
    _ROOKIEPY_OK = True
except ImportError:
    _ROOKIEPY_OK = False

# ─── 城市（行政区划代码 → 城市名）────────────────────────────────────────────
CITIES = {
    "110000": "北京",
    "310000": "上海",
    "440100": "广州",
    "440300": "深圳",
    "510100": "成都",
    "330100": "杭州",
    "500000": "重庆",
    "420100": "武汉",
    "320500": "苏州",
    "610100": "西安",
    "320100": "南京",
    "430100": "长沙",
    "410100": "郑州",
    "120000": "天津",
    "340100": "合肥",
    "370200": "青岛",
    "441900": "东莞",
    "330200": "宁波",
    "440600": "佛山",
    "230100": "哈尔滨",
    "210100": "沈阳",
    "370100": "济南",
    "350100": "福州",
    "360100": "南昌",
    "530100": "昆明",
    "140100": "太原",
}

_NAME_TO_CODE = {v: k for k, v in CITIES.items()}

TARGET_CITIES = [
    "北京",
    # "上海", "广州", "深圳",
    # "成都",
    # "杭州", "重庆", "武汉", "苏州", "西安",
    # "南京", "长沙", "郑州",  "合肥",
    # "青岛", "东莞", "宁波", "佛山",
    "天津",
]

# 每个城市最多采集条数（0 = 全部）
TOP_N = 0

API_URL = "https://so-landing.douyin.com/aweme/v1/hot/search/list/"

UA_DESKTOP = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/132.0.0.0 Safari/537.36"
)

COLUMNS = [
    "平台", "网站域名", "内容分类", "城市",
    "标题", "排名", "热度", "标签",
    "对应内容url", "对应内容",
    "发布时间", "采集时间", "批次id", "原始来源",
    "外网图片链接", "内网图片链接", "图片分辨率",
]

# label_type 映射
_LABEL = {1: "爆", 2: "新", 3: "热", 4: "推荐"}

# ─── Cookie 池（可选，用于视频内容抓取）────────────────────────────────────────
# dy_cookies/ 目录下每个 .txt 文件 = 一个账号的 cookie（一行完整 Cookie 字符串）
# 无 Cookie 时自动以基础模式运行（热搜列表+封面图正常，视频内容留空）

_COOKIE_POOL: list[tuple[str, str]] = []
_POOL_IDX:    int                   = 0
_SESSION: requests.Session | None   = None
_BASIC_MODE: bool                   = False


def _status_path() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "dy_cookies", "_status.json")


def _load_status() -> dict:
    try:
        with open(_status_path(), encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_status(status: dict):
    os.makedirs(os.path.dirname(_status_path()), exist_ok=True)
    with open(_status_path(), "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=2)


def mark_cookie_banned(name: str):
    status = _load_status()
    status[name] = {"ok": False, "banned_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    _save_status(status)
    print(f"  [Cookie池] 已标记「{name}」为不可用")


def _load_cookie_pool() -> list[tuple[str, str]]:
    global _COOKIE_POOL, _BASIC_MODE
    if _COOKIE_POOL:
        return _COOKIE_POOL
    if _BASIC_MODE:
        return []

    pool   = []
    status = _load_status()

    def _is_banned(name: str) -> bool:
        return not status.get(name, {}).get("ok", True) and "banned_at" in status.get(name, {})

    # ① 从 Chrome 读取 douyin.com 的 Cookie
    if _ROOKIEPY_OK:
        try:
            all_c = rookiepy.chrome(domains=["douyin.com", "www.douyin.com"])
            # 抖音视频搜索接口所需的关键 cookie 字段（按优先级排列）
            _pick = lambda name: next((c["value"] for c in all_c if c["name"] == name), "")
            sid_guard  = _pick("sid_guard")
            uid_tt     = _pick("uid_tt")
            sessionid  = _pick("sessionid")
            ttwid      = _pick("ttwid")
            ac_nonce   = _pick("__ac_nonce")
            odin_tt    = _pick("odin_tt")
            csrf_token = _pick("passport_csrf_token")
            ms_token   = _pick("msToken")
            if sid_guard or uid_tt:
                parts = []
                if sid_guard:  parts.append(f"sid_guard={sid_guard}")
                if uid_tt:     parts.append(f"uid_tt={uid_tt}")
                if sessionid:  parts.append(f"sessionid={sessionid}")
                if ttwid:      parts.append(f"ttwid={ttwid}")
                if ac_nonce:   parts.append(f"__ac_nonce={ac_nonce}")
                if odin_tt:    parts.append(f"odin_tt={odin_tt}")
                if csrf_token: parts.append(f"passport_csrf_token={csrf_token}")
                if ms_token:   parts.append(f"msToken={ms_token}")
                name = "Chrome账号"
                if not _is_banned(name):
                    pool.append((name, "; ".join(parts)))
                    print(f"  [Cookie池] {name} 已加入（douyin.com）")
            else:
                print("  [Cookie池] Chrome 中未检测到 douyin.com 登录态，请先在 Chrome 登录抖音")
        except Exception as e:
            print(f"  [Cookie池] Chrome 读取失败: {e}")

    # ② dy_cookies/ 目录下 .txt 文件
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    cookies_dir = os.path.join(script_dir, "dy_cookies")
    if os.path.isdir(cookies_dir):
        for fname in sorted(os.listdir(cookies_dir)):
            if not fname.endswith(".txt"):
                continue
            try:
                cookie = open(os.path.join(cookies_dir, fname),
                              encoding="utf-8").read().strip()
                if cookie:
                    name = fname[:-4]
                    if _is_banned(name):
                        print(f"  [Cookie池] {name} 已封禁，跳过")
                    else:
                        pool.append((name, cookie))
                        print(f"  [Cookie池] {name} 已加入")
            except Exception:
                pass

    if not pool:
        print("  [Cookie池] 无可用 Cookie → 基础模式（热搜列表+封面图正常，对应内容为空）")
        print("  提示：在 Chrome 登录 douyin.com 后重新运行，可自动获取视频内容\n")
        _BASIC_MODE = True
    else:
        print(f"  [Cookie池] 共 {len(pool)} 个可用账号\n")
        _COOKIE_POOL = pool
    return pool


def _current_cookie() -> str:
    pool = _load_cookie_pool()
    return pool[_POOL_IDX][1] if pool else ""


def _current_account() -> str:
    pool = _load_cookie_pool()
    return pool[_POOL_IDX][0] if pool else ""


def rotate_cookie() -> bool:
    global _POOL_IDX, _SESSION
    pool = _load_cookie_pool()
    _POOL_IDX += 1
    if _POOL_IDX >= len(pool):
        return False
    _SESSION = None
    print(f"\n  [Cookie切换] → 账号 {pool[_POOL_IDX][0]}  (共{len(pool)}个)")
    return True


def _make_session() -> requests.Session:
    s = requests.Session()
    retry = Retry(total=2, backoff_factor=0.3,
                  status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry,
                          pool_connections=4, pool_maxsize=8)
    s.mount("https://", adapter)
    s.mount("http://",  adapter)
    return s


def _get_session() -> requests.Session:
    global _SESSION
    if _SESSION is None:
        _SESSION = _make_session()
    return _SESSION


def _hotlist_headers() -> dict:
    return {
        "User-Agent":      UA_DESKTOP,
        "Accept":          "*/*",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Cache-Control":   "no-cache",
        "Pragma":          "no-cache",
        "Referer":         "https://so-landing.douyin.com/landings/hotlist",
    }


def _video_headers() -> dict:
    return {
        "User-Agent":      UA_DESKTOP,
        "Accept":          "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Referer":         "https://www.douyin.com/",
        "Cookie":          _current_cookie(),
    }


# ─── 工具 ──────────────────────────────────────────────────────────────────────

def _strip_html(t: str) -> str:
    return re.sub(r"<[^>]+>", "", t or "").strip()


def _unix_to_str(ts) -> str:
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


# ─── 接口一：城市热搜列表 ──────────────────────────────────────────────────────

def get_city_hot_list(city: str, top_n: int = 0) -> list[dict]:
    """
    获取抖音城市热搜列表。
    top_n: 最多返回条数，0 = 不限（返回全部）。
    """
    city_code = _NAME_TO_CODE.get(city)
    if not city_code:
        raise ValueError(f"城市 '{city}' 不在列表中")

    params = urllib.parse.urlencode({
        "aid":            "581610",
        "detail_list":    "1",
        "board_type":     "1",
        "board_sub_type": city_code,
        "need_board_tab": "true",
        "need_covid_tab": "false",
        "version_code":   "32.3.0",
    })
    try:
        resp = _get_session().get(
            API_URL + "?" + params,
            headers=_hotlist_headers(),
            timeout=10,
        )
        raw = resp.json()
    except Exception as e:
        raise RuntimeError(f"接口请求失败: {e}")

    items = (raw.get("data") or {}).get("word_list") or []
    result = []
    for item in items:
        word = item.get("word", "").strip()
        if not word:
            continue

        # 封面图
        cover     = item.get("word_cover") or {}
        url_list  = cover.get("url_list") or []
        cover_url = url_list[0] if url_list else ""
        cover_w   = int(cover.get("width")  or 0)
        cover_h   = int(cover.get("height") or 0)

        result.append({
            "rank":        item.get("position") or (len(result) + 1),
            "word":        word,
            "hot_value":   item.get("hot_value", 0),
            "view_count":  item.get("view_count", 0),
            "video_count": item.get("video_count", 0),
            "sentence_id": item.get("sentence_id", ""),
            "event_time":  item.get("event_time", 0),
            "cover_url":   cover_url,
            "cover_w":     cover_w,
            "cover_h":     cover_h,
            "label":       item.get("label", 0),   # 正确字段名（非 label_type）
            "group_id":    str(item.get("group_id") or ""),  # 关联视频ID，无需cookie直取详情
            "city":        city,
            "city_code":   city_code,
        })
        if top_n > 0 and len(result) >= top_n:
            break

    return result


# ─── Playwright 浏览器（关键词搜索视频内容）────────────────────────────────

try:
    from playwright.sync_api import sync_playwright as _sync_playwright
    _PW_AVAILABLE = True
except ImportError:
    _PW_AVAILABLE = False

_pw_obj     = None   # sync_playwright().__enter__() 实例
_pw_context = None   # BrowserContext（全程共享，避免反复启动浏览器）


def _chrome_user_data() -> str:
    """获取 Chrome 用户数据目录路径"""
    import platform
    if platform.system() == "Windows":
        return os.path.join(os.environ.get("LOCALAPPDATA", ""),
                            "Google", "Chrome", "User Data")
    if platform.system() == "Darwin":
        return os.path.expanduser("~/Library/Application Support/Google/Chrome")
    return os.path.expanduser("~/.config/google-chrome")


def _init_playwright() -> bool:
    """
    用系统已安装的 Chrome（channel=chrome）启动 persistent context，
    直接继承用户的 Chrome 登录态和 Cookie，绕过机器人检测。
    返回 True 表示初始化成功。
    """
    global _pw_obj, _pw_context
    if _pw_context is not None:
        return True
    if not _PW_AVAILABLE:
        return False

    user_data = _chrome_user_data()
    if not os.path.isdir(user_data):
        print(f"  [Playwright] 未找到 Chrome 用户目录: {user_data}")
        return False

    try:
        _pw_obj     = _sync_playwright().__enter__()
        # launch_persistent_context 使用真实 Chrome profile，自带 Cookie 和登录态
        _pw_context = _pw_obj.chromium.launch_persistent_context(
            user_data_dir = user_data,
            channel       = "chrome",
            headless      = False,      # 有头模式：绕过抖音验证码检测
            args          = [
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--window-position=-32000,-32000",  # 窗口移出屏幕外，用户不可见
            ],
            viewport      = {"width": 1920, "height": 1080},
            locale        = "zh-CN",
            timezone_id   = "Asia/Shanghai",
        )
        print(f"  [Playwright] Chrome 已启动（继承本机 Cookie，有头隐藏模式）")
        # 热身：先访问 douyin.com 让浏览器完全就绪，避免第一条数据丢失
        try:
            _warmup = _pw_context.new_page()
            _warmup.goto("https://www.douyin.com/", wait_until="domcontentloaded", timeout=10000)
            _warmup.wait_for_timeout(2000)
            _warmup.close()
        except Exception:
            pass
        return True
    except Exception as e:
        print(f"  [Playwright] Chrome 启动失败: {e}")
        # 回退：Chromium + 手动注入 Cookie
        try:
            browser     = _pw_obj.chromium.launch(
                headless=False,
                args=["--no-sandbox", "--disable-blink-features=AutomationControlled",
                      "--window-position=-32000,-32000"],
            )
            _pw_context = browser.new_context(
                user_agent  = UA_DESKTOP,
                viewport    = {"width": 1920, "height": 1080},
                locale      = "zh-CN",
                timezone_id = "Asia/Shanghai",
            )
            if _ROOKIEPY_OK:
                try:
                    all_c = rookiepy.chrome(domains=["douyin.com", "www.douyin.com"])
                    pw_cookies = []
                    for c in all_c:
                        entry = {"name": c["name"], "value": c["value"],
                                 "domain": c.get("domain", ".douyin.com"),
                                 "path":   c.get("path", "/")}
                        ss = c.get("sameSite", "")
                        if ss in ("Strict", "Lax", "None"):
                            entry["sameSite"] = ss
                        pw_cookies.append(entry)
                    if pw_cookies:
                        _pw_context.add_cookies(pw_cookies)
                except Exception:
                    pass
            print(f"  [Playwright] Chromium 回退模式已启动")
            try:
                _warmup = _pw_context.new_page()
                _warmup.goto("https://www.douyin.com/", wait_until="domcontentloaded", timeout=10000)
                _warmup.wait_for_timeout(2000)
                _warmup.close()
            except Exception:
                pass
            return True
        except Exception as e2:
            print(f"  [Playwright] 回退也失败: {e2}")
            return False


def _close_playwright():
    """关闭 Playwright 浏览器，释放资源"""
    global _pw_obj, _pw_context
    if _pw_context:
        try:
            _pw_context.browser.close()
        except Exception:
            pass
        _pw_context = None
    if _pw_obj:
        try:
            _pw_obj.__exit__(None, None, None)
        except Exception:
            pass
        _pw_obj = None


def _parse_search_response(body_bytes: bytes) -> dict:
    """
    从搜索响应 body 里提取第一条有 desc 的视频，支持 JSON 和 NDJSON 两种格式。
    兜底：正则从原始文本提取 desc / aweme_id / nickname。
    """
    raw_str = body_bytes.decode("utf-8", errors="ignore")

    def _find_in_obj(raw: dict) -> dict:
        for src in [raw.get("data") or [], raw.get("aweme_list") or []]:
            if not isinstance(src, list):
                continue
            for item in src:
                v = item.get("aweme_info") or item
                if isinstance(v, dict) and v.get("desc"):
                    return v
        return {}

    # 整体 JSON
    try:
        found = _find_in_obj(json.loads(raw_str))
        if found:
            return found
    except Exception:
        pass

    # NDJSON：每行一个 JSON 对象
    for line in raw_str.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        try:
            found = _find_in_obj(json.loads(line))
            if found:
                return found
        except Exception:
            pass

    # 正则兜底：直接从文本里提取关键字段
    m_desc = re.search(r'"desc"\s*:\s*"([^"]{5,})"', raw_str)
    m_nick = re.search(r'"nickname"\s*:\s*"([^"]+)"', raw_str)
    m_id   = re.search(r'"aweme_id"\s*:\s*"(\d+)"', raw_str)
    m_ts   = re.search(r'"create_time"\s*:\s*(\d+)', raw_str)
    if m_desc:
        return {
            "desc":        m_desc.group(1),
            "aweme_id":    m_id.group(1)   if m_id   else "",
            "create_time": int(m_ts.group(1)) if m_ts else 0,
            "author":      {"nickname": m_nick.group(1) if m_nick else ""},
        }
    return {}


def fetch_top_video(keyword: str) -> dict:
    """
    用 Playwright 打开抖音搜索页，拦截搜索流式 API 响应，返回第一条视频信息。
    实际端点：/aweme/v1/web/general/search/stream/
    """
    if not _PW_AVAILABLE or _pw_context is None:
        return {}

    result: dict = {}

    def _on_response(resp):
        if result:
            return
        if "search" not in resp.url:
            return
        try:
            body  = resp.body()
            found = _parse_search_response(body)
            if found:
                result.update(found)
        except Exception:
            pass

    page = _pw_context.new_page()
    page.on("response", _on_response)
    try:
        kw_q = urllib.parse.quote(keyword)
        page.goto(
            f"https://www.douyin.com/search/{kw_q}?source=hot_search_board_web",
            wait_until="domcontentloaded",
            timeout=15000,
        )
        # 等待搜索 API 响应（最多 15 秒）
        for _ in range(150):
            if result:
                break
            page.wait_for_timeout(100)
    except Exception:
        pass
    finally:
        page.close()

    return result


def _video_cover_url(video: dict) -> str:
    """从视频详情里提取封面 URL"""
    cover    = (video.get("video") or {}).get("cover") or {}
    url_list = cover.get("url_list") or []
    return url_list[0] if url_list else ""


# ─── 图片下载 ──────────────────────────────────────────────────────────────────

_IMG_HEADERS = {
    "User-Agent": UA_DESKTOP,
    "Referer":    "https://www.douyin.com/",
    "Accept":     "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
}


def download_image(url: str, save_dir: str, filename: str) -> str:
    """下载图片到本地，返回绝对路径；失败返回空字符串。"""
    if not url:
        return ""
    os.makedirs(save_dir, exist_ok=True)
    ext = os.path.splitext(url.split("?")[0])[-1]
    if not ext.startswith(".") or len(ext) > 5:
        ext = ".jpg"
    local_path = os.path.join(save_dir, filename + ext)
    if os.path.exists(local_path):
        return local_path
    try:
        resp = _get_session().get(url, headers=_IMG_HEADERS, timeout=8)
        if resp.status_code == 200:
            with open(local_path, "wb") as f:
                f.write(resp.content)
            return local_path
    except Exception:
        pass
    return ""


# ─── 构建记录 ──────────────────────────────────────────────────────────────────

def build_record(city: str, hot_item: dict, video: dict,
                 collect_time: str, batch_id: str,
                 img_dir: str = "") -> dict:
    word = hot_item["word"]
    rank = hot_item["rank"]

    # 视频文案：有 cookie 时取视频描述，无 cookie 则留空
    content = _strip_html(video["desc"]) if (video and video.get("desc")) else ""

    # 发布时间：优先视频 create_time，次用热搜 event_time
    create_ts = (video.get("create_time") or 0) if video else 0
    pub_time  = _unix_to_str(create_ts or hot_item.get("event_time", 0))

    # 视频 ID：优先 fetch 返回的 aweme_id，次用热搜列表里的 group_id
    aweme_id = (video.get("aweme_id") or "") if video else ""
    if not aweme_id:
        aweme_id = hot_item.get("group_id", "")
    if aweme_id:
        content_url = f"https://www.douyin.com/video/{aweme_id}"
    else:
        content_url = (
            f"https://www.douyin.com/search/{urllib.parse.quote(word)}"
            "?source=hot_search_board_web"
        )

    # 原始来源：视频作者昵称
    author_nick = (video.get("author", {}).get("nickname") or "") if video else ""

    # 封面图：优先视频封面，次用热搜 word_cover
    ext_img = (_video_cover_url(video) if video else "") or hot_item.get("cover_url", "")

    # 分辨率：优先 API 返回的尺寸
    cw, ch = hot_item.get("cover_w", 0), hot_item.get("cover_h", 0)
    resolution = f"{cw}x{ch}" if cw and ch else ""

    # 标签：label 文字映射（0=无标签）
    tag = _LABEL.get(hot_item.get("label", 0), "")

    # 下载图片到本地
    local_img = ""
    if ext_img and img_dir:
        fname     = f"{city}_{rank:02d}_{aweme_id or uuid.uuid4().hex[:8]}"
        local_img = download_image(ext_img, img_dir, fname)

    # 图片分辨率兜底：API 未返回时从本地文件读取
    if not resolution and local_img:
        try:
            from PIL import Image as _PILImage
            with _PILImage.open(local_img) as _img:
                resolution = f"{_img.width}x{_img.height}"
        except Exception:
            pass

    return {
        "平台":        "抖音",
        "网站域名":    "douyin.com",
        "内容分类":    "城市热搜",
        "城市":        city,
        "标题":        word,
        "排名":        rank,
        "热度":        hot_item["hot_value"],
        "标签":        tag,
        "对应内容url":  content_url,
        "对应内容":    content,
        "发布时间":    pub_time,
        "采集时间":    collect_time,
        "批次id":      batch_id,
        "原始来源":    author_nick or f"抖音{city}城市热搜",
        "外网图片链接": ext_img,
        "内网图片链接": local_img,
        "图片分辨率":   resolution,
    }


# ─── Excel 写出 ────────────────────────────────────────────────────────────────

_CITY_COLORS = [
    "DDEEFF", "EEF5FF", "DDF0E8", "EEF8F3",
    "FFF3DD", "FFF9EE", "F0DDEE", "F8EEF5",
    "DDEEEE", "EEF5F5",
]
_HEADER_FILL  = PatternFill("solid", fgColor="1B4F9A")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
_DATA_FONT    = Font(name="微软雅黑", size=9)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN         = Side(style="thin", color="CCCCCC")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COL_WIDTHS = {
    "平台": 8, "网站域名": 14, "内容分类": 12, "城市": 8,
    "标题": 28, "排名": 6, "热度": 10, "标签": 8,
    "对应内容url": 52, "对应内容": 48,
    "发布时间": 20, "采集时间": 20, "批次id": 24, "原始来源": 18,
    "外网图片链接": 55, "内网图片链接": 55, "图片分辨率": 12,
}
_CENTER_COLS = {
    "平台", "网站域名", "内容分类", "城市", "排名", "热度", "标签",
    "发布时间", "采集时间", "批次id", "原始来源", "图片分辨率",
}


def save_excel(records: list[dict], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "抖音城市热搜"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 16)
    ws.row_dimensions[1].height = 22

    city_order = []
    for r in records:
        c = r.get("城市", "")
        if c not in city_order:
            city_order.append(c)
    city_color = {c: _CITY_COLORS[i % len(_CITY_COLORS)] for i, c in enumerate(city_order)}

    for row_idx, rec in enumerate(records, 2):
        fill = PatternFill("solid", fgColor=city_color.get(rec.get("城市", ""), "FFFFFF"))
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col_name, ""))
            cell.fill      = fill
            cell.font      = _DATA_FONT
            cell.border    = _BORDER
            cell.alignment = _ALIGN_CENTER if col_name in _CENTER_COLS else _ALIGN_LEFT
        ws.row_dimensions[row_idx].height = 45

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(filename)
    print(f"  Excel 已保存：{filename}  ({len(records)} 行)")


# ─── 全量采集主流程 ────────────────────────────────────────────────────────────

def collect_all(cities: list[str] = None, delay: float = 1.5,
                img_dir: str = "", top_n: int = 0) -> tuple:
    if cities is None:
        cities = TARGET_CITIES

    # 初始化 Cookie 池（可选）
    _load_cookie_pool()

    # 启动 Playwright 浏览器（用于获取视频内容）
    pw_ok = _init_playwright()

    batch_id     = datetime.now().strftime("%Y%m%d%H%M%S") + "_" + str(uuid.uuid4())[:6]
    collect_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_records  = []
    start_dt     = datetime.now()

    if not img_dir:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        img_dir    = os.path.join(script_dir, "dy_images", batch_id)
    os.makedirs(img_dir, exist_ok=True)

    mode_label = "Playwright模式" if pw_ok else ("API模式" if not _BASIC_MODE else "基础模式（无Cookie）")
    n_label    = f"每城 TOP {top_n}" if top_n > 0 else "每城全量"
    print(f"\n批次 ID : {batch_id}")
    print(f"城市数量 : {len(cities)}   {n_label}   {mode_label}")
    print(f"图片目录 : {img_dir}")
    print(f"开始时间 : {collect_time}")
    print("=" * 70)

    for city_idx, city in enumerate(cities, 1):
        print(f"\n[{city_idx:02d}/{len(cities)}] {city}", end="", flush=True)
        try:
            hot_list = get_city_hot_list(city, top_n=top_n)
        except Exception as e:
            print(f"  热榜失败: {e}")
            continue

        print(f"  共 {len(hot_list)} 条", end="", flush=True)
        city_records = []

        for idx, hot_item in enumerate(hot_list):
            # Playwright 按关键词搜索，拦截 API 响应拿第一条视频
            video = fetch_top_video(hot_item["word"])

            rec     = build_record(city, hot_item, video, collect_time, batch_id, img_dir)
            city_records.append(rec)

            has_img     = bool(rec["内网图片链接"])
            has_content = bool(rec["对应内容"])
            symbol = "o" if has_img else "-"
            if has_content:
                symbol = symbol.upper()
            print(f" {symbol}", end="", flush=True)

            # 每条之间固定随机等待 1-3 秒（不管有没有 Cookie）
            time.sleep(random.uniform(1, 3))

        all_records.extend(city_records)
        print(f"  OK {len(city_records)} 条入库")

        # 城市间间隔
        if city_idx < len(cities):
            city_delay = 1.0 if _BASIC_MODE else 2.0
            time.sleep(city_delay)

    elapsed = int((datetime.now() - start_dt).total_seconds())
    print(f"\n{'='*70}")
    print(f"采集完成  总计 {len(all_records)} 条  耗时 {elapsed} 秒")

    _close_playwright()
    return all_records, batch_id


# ─── 主函数 ───────────────────────────────────────────────────────────────────

def main():
    records, batch_id = collect_all(TARGET_CITIES, top_n=TOP_N)
    if not records:
        print("无数据")
        return

    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx   = f"抖音城市热搜_{ts}.xlsx"
    json_f = f"抖音城市热搜_{ts}.json"

    save_excel(records, xlsx)

    with open(json_f, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"  JSON 备份：{json_f}")


if __name__ == "__main__":
    main()
