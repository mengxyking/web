"""
微博城市同城热搜 全量采集 → Excel
==========================================================
核心接口（正确的同城热搜接口）：
  GET https://m.weibo.cn/api/container/getIndex
      ?containerid=106003type%3D25%26t%3D3%26disable_hot%3D1%26filter_type%3Dregion
      &lon={经度}&lat={纬度}

  返回当前城市（由经纬度确定）的同城热搜列表
  每条数据字段：title_sub=话题标题, desc=同城热度+来源, scheme=内容URL

话题详情接口：
  GET https://m.weibo.cn/api/container/getIndex
      ?containerid=100103type%3D1%26q%3D{keyword}

Cookie 自动从 Chrome 读取（rookiepy），Chrome 保持登录 weibo.com 即可。
"""

import urllib.parse
import json
import os
import random
import re
import time
import uuid
from datetime import datetime
from email.utils import parsedate_to_datetime

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import rookiepy
    _ROOKIEPY_OK = True
except ImportError:
    _ROOKIEPY_OK = False

# ─── 城市坐标（经度，纬度）────────────────────────────────────────────────────
CITIES = {
    "北京":    (116.3972, 39.9093),
    "上海":    (121.4737, 31.2304),
    "广州":    (113.2644, 23.1292),
    "深圳":    (114.0579, 22.5431),
    "成都":    (104.0657, 30.6595),
    "杭州":    (120.1551, 30.2741),
    "重庆":    (106.5516, 29.5630),
    "武汉":    (114.3054, 30.5932),
    "苏州":    (120.5853, 31.2990),
    "西安":    (108.9402, 34.3416),
    "南京":    (118.7963, 32.0603),
    "长沙":    (112.9388, 28.2282),
    "郑州":    (113.6253, 34.7466),
    "天津":    (117.1995, 39.1439),
    "合肥":    (117.2272, 31.8206),
    "青岛":    (120.3826, 36.0671),
    "东莞":    (113.7465, 23.0457),
    "宁波":    (121.5440, 29.8683),
    "佛山":    (113.1220, 23.0220),
    "沈阳":    (123.4295, 41.7963),
    "哈尔滨":  (126.5350, 45.8038),
    "长春":    (125.3235, 43.8171),
    "昆明":    (102.8329, 24.8801),
    "福州":    (119.3062, 26.0753),
    "厦门":    (118.0894, 24.4798),
    "南昌":    (115.8581, 28.6829),
    "济南":    (117.0009, 36.6758),
    "石家庄":  (114.5149, 38.0428),
    "太原":    (112.5489, 37.8706),
    "贵阳":    (106.7135, 26.5783),
    "南宁":    (108.3667, 22.8170),
    "海口":    (110.1999, 20.0440),
    "呼和浩特":(111.7519, 40.8415),
    "兰州":    (103.8343, 36.0611),
    "银川":    (106.2782, 38.4664),
    "西宁":    (101.7782, 36.6171),
    "乌鲁木齐":(87.6168,  43.8256),
    "拉萨":    (91.1409,  29.6456),
    "温州":    (120.6722, 28.0000),
    "无锡":    (120.3119, 31.4912),
    "大连":    (121.6147, 38.9140),
}

TARGET_CITIES = [
    "北京",
    "上海", "广州", "深圳",
    "成都",
    "杭州", "重庆", "武汉", "苏州", "西安",
    "南京", "长沙", "郑州", "天津", "合肥",
    "青岛", "东莞", "宁波", "佛山",
]

# 每个城市最多采集条数（0 = 全部）
TOP_N = 0

UA_MOBILE = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Mobile/15E148 MicroMessenger/8.0.30"
)

COLUMNS = [
    "平台", "网站域名", "内容分类", "城市",
    "标题", "排名", "热度", "标签",
    "对应内容url", "对应内容",
    "发布时间", "采集时间", "批次id", "原始来源",
    "外网图片链接", "内网图片链接", "图片分辨率",
]

# ─── Cookie 池 ────────────────────────────────────────────────────────────────
# cookies/ 目录下每个 .txt 文件 = 一个账号的 cookie（一行：SUB=xxx; SUBP=xxx）
# 触发验证码时自动切换到下一个账号，所有账号都被封才停止。
# 若 cookies/ 目录为空或不存在，则自动从 Chrome 读取。

_COOKIE_POOL: list[tuple[str, str]] = []   # [(账号名, cookie字符串), ...]
_POOL_IDX:    int                   = 0    # 当前使用的 cookie 下标
_SESSION: requests.Session | None   = None

# 状态文件路径（启动时加载，被封时写入）
def _status_path() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "cookies", "_status.json")

def _load_status() -> dict:
    """读取 cookie 状态表，格式：{账号名: {"ok": bool, "banned_at": "时间"}}"""
    try:
        with open(_status_path(), encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _save_status(status: dict):
    os.makedirs(os.path.dirname(_status_path()), exist_ok=True)
    with open(_status_path(), "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=2)

def mark_cookie_banned(name: str):
    """把指定账号标记为已封禁，写入状态文件"""
    status = _load_status()
    status[name] = {"ok": False, "banned_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    _save_status(status)
    print(f"  [Cookie池] 已标记「{name}」为不可用，下次运行自动跳过")


def _load_cookie_pool() -> list[tuple[str, str]]:
    """
    构建 Cookie 池，优先级：
      1. Chrome 当前登录账号（rookiepy 读取）—— 排在最前
      2. cookies/ 目录下所有 .txt 文件（每文件一个账号）
    被封后按顺序切换，全部失效才停止。
    """
    global _COOKIE_POOL
    if _COOKIE_POOL:
        return _COOKIE_POOL

    pool   = []
    status = _load_status()   # 读取封禁状态

    def _is_banned(name: str) -> bool:
        return not status.get(name, {}).get("ok", True) and "banned_at" in status.get(name, {})

    # ① 先尝试从 Chrome 读取
    if _ROOKIEPY_OK:
        try:
            all_c   = rookiepy.chrome(domains=["weibo.com", "weibo.cn", "passport.weibo.com"])
            sub_cn  = next((c["value"] for c in all_c if c["name"] == "SUB"  and c.get("domain") == ".weibo.cn"), "")
            subp_cn = next((c["value"] for c in all_c if c["name"] == "SUBP" and c.get("domain") == ".weibo.cn"), "")
            if sub_cn:
                name = "Chrome账号"
                if _is_banned(name):
                    print(f"  [Cookie池] {name} 已被标记封禁，跳过（如需重用请删除 cookies/_status.json 中对应条目）")
                else:
                    pool.append((name, f"SUB={sub_cn}; SUBP={subp_cn}"))
                    print(f"  [Cookie池] {name} 已加入")
        except Exception as e:
            print(f"  [Cookie池] Chrome 读取失败: {e}")

    # ② 再从 cookies/ 目录读取 txt 文件
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    cookies_dir = os.path.join(script_dir, "cookies")
    if os.path.isdir(cookies_dir):
        for fname in sorted(os.listdir(cookies_dir)):
            if not fname.endswith(".txt"):
                continue
            try:
                cookie = open(os.path.join(cookies_dir, fname), encoding="utf-8").read().strip()
                if cookie:
                    name = fname[:-4]
                    if _is_banned(name):
                        banned_at = status[name].get("banned_at", "")
                        print(f"  [Cookie池] {name} 已被封禁（{banned_at}），跳过")
                    else:
                        pool.append((name, cookie))
                        print(f"  [Cookie池] {name} 已加入")
            except Exception:
                pass

    if not pool:
        raise RuntimeError(
            "所有 Cookie 均不可用（已被封禁或未找到）！\n"
            "方式1：保持 Chrome 登录 weibo.com（自动读取）\n"
            "方式2：在 cookies/ 目录放置 账号名.txt（内容：SUB=xxx; SUBP=xxx）\n"
            "解封方式：删除 cookies/_status.json 中对应账号条目，或删除整个文件"
        )

    print(f"  [Cookie池] 共 {len(pool)} 个可用账号\n")
    _COOKIE_POOL = pool
    return pool


def _current_cookie() -> str:
    """返回当前账号的 cookie 字符串"""
    pool = _load_cookie_pool()
    return pool[_POOL_IDX][1]


def _current_account() -> str:
    """返回当前账号名"""
    pool = _load_cookie_pool()
    return pool[_POOL_IDX][0]


def rotate_cookie() -> bool:
    """
    切换到下一个 cookie。
    返回 True 表示切换成功，False 表示所有 cookie 都已用尽。
    """
    global _POOL_IDX, _SESSION
    pool = _load_cookie_pool()
    _POOL_IDX += 1
    if _POOL_IDX >= len(pool):
        return False
    _SESSION = None   # 新账号需要重建 Session
    print(f"\n  [Cookie切换] → 账号 {pool[_POOL_IDX][0]}  (共{len(pool)}个)")
    return True


def _make_session() -> requests.Session:
    """创建带连接池和自动重试的 Session"""
    s = requests.Session()
    retry = Retry(total=2, backoff_factor=0.3,
                  status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry,
                          pool_connections=4, pool_maxsize=8)
    s.mount("https://", adapter)
    s.mount("http://",  adapter)
    return s


def _get_session() -> requests.Session:
    global _SESSION
    if _SESSION is None:
        _SESSION = _make_session()
    return _SESSION


def _mh() -> dict:
    """移动端请求头（自动使用当前 cookie 池中的账号）"""
    return {
        "User-Agent":       UA_MOBILE,
        "Accept":           "application/json",
        "Referer":          "https://m.weibo.cn/",
        "Cookie":           _current_cookie(),
        "X-Requested-With": "XMLHttpRequest",
    }


class CaptchaError(RuntimeError):
    """微博触发验证码，需要在 Chrome 里手动完成后重试"""


def _get(url: str, timeout: int = 5) -> dict:
    """GET 请求，返回 JSON dict；触发验证码时抛 CaptchaError"""
    try:
        resp = _get_session().get(url, headers=_mh(), timeout=timeout)
        if resp.headers.get("Content-Type", "").startswith("text/html"):
            raise CaptchaError("m.weibo.cn 返回 HTML（验证码页面），请在 Chrome 打开 m.weibo.cn 完成验证后重试")
        data = resp.json()
        if data.get("ok") == -100 and "captcha" in data.get("url", ""):
            raise CaptchaError("触发微博验证码，请在 Chrome 打开 m.weibo.cn 完成验证后重试")
        return data
    except CaptchaError:
        raise
    except Exception:
        return {}


# ─── 工具 ─────────────────────────────────────────────────────────────────────

def _strip_html(t: str) -> str:
    return re.sub(r"<[^>]+>", "", t or "").strip()


def _parse_time(s: str) -> str:
    try:
        return parsedate_to_datetime(s).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return s


# ─── 接口一：城市同城热搜（正确接口）─────────────────────────────────────────

def get_city_hot_list(city: str, top_n: int = 0) -> list[dict]:
    """
    用经纬度获取城市同城热搜。
    top_n: 最多返回条数，0 表示不限（返回全部）。
    接口：m.weibo.cn/api/container/getIndex?containerid=106003type%3D25%26...filter_type%3Dregion&lon=...&lat=...
    """
    if city not in CITIES:
        raise ValueError(f"城市 '{city}' 不在列表中")
    lon, lat = CITIES[city]
    url = (
        "https://m.weibo.cn/api/container/getIndex"
        "?containerid=106003type%3D25%26t%3D3%26disable_hot%3D1%26filter_type%3Dregion"
        f"&lon={lon}&lat={lat}"
    )
    raw = _get(url, timeout=8)
    if raw.get("ok") != 1:
        raise RuntimeError(f"接口异常: {raw.get('msg', raw)}")

    items = []
    rank = 0
    for card in raw.get("data", {}).get("cards", []):
        for cg in card.get("card_group") or []:
            if cg.get("card_type") != 25 or not cg.get("title_sub"):
                continue
            rank += 1
            desc    = cg.get("desc", "")
            scheme  = cg.get("scheme", "")

            # 解析热度值
            m_hot = re.search(r"同城热度(\d+)", desc)
            hot_val = int(m_hot.group(1)) if m_hot else 0

            # 解析来源账号 (@xxx)
            source_user = ""
            if "|" in desc:
                after = desc.split("|")[-1].strip()
                if after.startswith("@"):
                    source_user = after

            # 关键词直接用 title_sub（原始标题，干净无噪音）
            # 去掉首尾的 # 号（部分条目带话题标记）
            keyword = cg.get("title_sub", "").strip("#").strip()

            # 从 scheme 里提取 city_code_set
            city_code_match = re.search(r"city_code_set=(\d+)", scheme)
            city_code_set = city_code_match.group(1) if city_code_match else ""

            # 构建简洁的内容URL
            topic_q = urllib.parse.quote(keyword)
            content_url = f"https://s.weibo.com/weibo?q={topic_q}&Refer=top"

            items.append({
                "rank":         rank,
                "title":        cg.get("title_sub", ""),
                "hot_value":    hot_val,
                "source_user":  source_user,
                "desc":         desc,
                "keyword":      keyword,
                "content_url":  content_url,
                "city_code_set": city_code_set,
                "scheme":       scheme,
            })
            if top_n > 0 and len(items) >= top_n:
                return items
    return items


# ─── 接口二：话题搜索 → 热门第一帖 ────────────────────────────────────────────

def fetch_top_post(keyword: str) -> dict:
    """
    搜索关键词，返回第一条 card_type=9 的 mblog。
    优先用 t=10 + #keyword# 话题格式（与同城热搜 scheme 一致），
    找不到再降级用 t=1 普通关键词兜底。
    """
    def _search(containerid: str) -> dict:
        url = f"https://m.weibo.cn/api/container/getIndex?containerid={containerid}"
        raw = _get(url, timeout=5)
        for card in raw.get("data", {}).get("cards", []):
            # 方式一：顶层 card_type=9（最常见）
            if card.get("card_type") == 9:
                m = card.get("mblog", {})
                if m:
                    return m
            # 方式二：card_type=11 容器内嵌套 mblog（部分话题搜索结果）
            for item in card.get("card_group") or []:
                m = item.get("mblog", {})
                if m and m.get("text"):
                    return m
        return {}

    # 方案一：t=10 + #keyword# 话题格式（命中率高）
    kw_hash  = urllib.parse.quote(f"#{keyword}#")
    result   = _search(f"100103type%3D1%26t%3D10%26q%3D{kw_hash}")
    if result:
        return result

    # 方案二：t=1 普通关键词兜底
    kw_plain = urllib.parse.quote(keyword)
    return _search(f"100103type%3D1%26q%3D{kw_plain}")


def _image_info(mblog: dict) -> tuple[str, str, str]:
    """
    返回 (外网图片链接, 内网图片链接, 图片分辨率)
    外网 = large 高清图；内网 = orj360/orj480 缩略图
    """
    def _to_large(url: str) -> str:
        """把缩略尺寸替换成 large"""
        return re.sub(r"/(orj\d+|thumbnail|square|mw690)/", "/large/", url or "")

    pics = mblog.get("pics") or []
    if pics:
        pic   = pics[0]
        large = pic.get("large", {})
        inn   = pic.get("url", "")                       # 缩略图作内网
        ext   = large.get("url") or _to_large(inn)       # 高清图作外网
        geo   = large.get("geo", {})
        res   = f"{geo.get('width','')}x{geo.get('height','')}" if geo.get("width") else ""
        return ext, inn, res

    # 无 pics，兜底取视频/文章封面
    pp = mblog.get("page_info", {}).get("page_pic", {})
    if pp.get("url"):
        inn = pp["url"]
        ext = _to_large(inn)
        res = f"{pp.get('width','')}x{pp.get('height','')}" if pp.get("width") else ""
        return ext, inn, res

    return "", "", ""


# ─── 图片下载 ─────────────────────────────────────────────────────────────────

_IMG_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/132.0.0.0 Safari/537.36",
    "Referer":    "https://weibo.com/",
    "Accept":     "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
}


def _thumb_url(url: str) -> str:
    """把 large/mw2000 换成 thumbnail（小图，下载快）"""
    return re.sub(r"/(large|mw2000|mw690|orj480)/", "/thumbnail/", url or "")


def download_image(url: str, save_dir: str, filename: str) -> str:
    """
    下载缩略图到本地，返回绝对路径；失败返回空字符串。
    使用全局 Session（复用连接，快 5-10x）。
    """
    if not url:
        return ""
    thumb = _thumb_url(url)         # 用缩略图，体积小速度快
    os.makedirs(save_dir, exist_ok=True)
    ext = os.path.splitext(thumb.split("?")[0])[-1] or ".jpg"
    if not ext.startswith("."):
        ext = ".jpg"
    local_path = os.path.join(save_dir, filename + ext)
    if os.path.exists(local_path):
        return local_path
    try:
        resp = _get_session().get(thumb, headers=_IMG_HEADERS, timeout=5)
        if resp.status_code == 200:
            with open(local_path, "wb") as f:
                f.write(resp.content)
            return local_path
    except Exception:
        pass
    return ""


# ─── 构建记录 ─────────────────────────────────────────────────────────────────

def build_record(city: str, hot_item: dict, mblog: dict,
                 collect_time: str, batch_id: str,
                 img_dir: str = "") -> dict:
    text       = _strip_html(mblog.get("text", "")) if mblog else ""
    pub_time   = _parse_time(mblog.get("created_at", "")) if mblog else ""
    uid        = mblog.get("user", {}).get("id", "") if mblog else ""
    mid        = mblog.get("mid", "") if mblog else ""
    post_url   = f"https://weibo.com/{uid}/{mid}" if uid and mid else hot_item["content_url"]
    ext_img, _, resolution = _image_info(mblog) if mblog else ("", "", "")

    # 原始来源：热搜列表中的 @媒体账号 → 兜底用发帖人昵称
    source_user = hot_item.get("source_user", "")
    if not source_user and mblog:
        author_name = mblog.get("user", {}).get("screen_name", "")
        if author_name:
            source_user = f"@{author_name}"

    # 标签：发帖人认证身份（如"央视新闻官方账号"），无则留空
    tag = ""
    if mblog:
        tag = mblog.get("user", {}).get("verified_reason", "")

    # 下载图片到本地，内网图片链接 = 本地文件路径
    local_img = ""
    if ext_img and img_dir:
        fname = f"{city}_{hot_item['rank']:02d}_{mid or uuid.uuid4().hex[:8]}"
        local_img = download_image(ext_img, img_dir, fname)

    # 图片分辨率兜底：API 未返回 geo 时从本地文件读取实际尺寸
    if not resolution and local_img:
        try:
            from PIL import Image as _PILImage
            with _PILImage.open(local_img) as _img:
                resolution = f"{_img.width}x{_img.height}"
        except Exception:
            pass

    return {
        "平台":       "微博",
        "网站域名":   "weibo.com",
        "内容分类":   "城市热搜",
        "城市":       city,
        "标题":       hot_item["title"],
        "排名":       hot_item["rank"],
        "热度":       hot_item["hot_value"],
        "标签":       tag,
        "对应内容url": post_url,
        "对应内容":   text,
        "发布时间":   pub_time,
        "采集时间":   collect_time,
        "批次id":     batch_id,
        "原始来源":   source_user or f"微博{city}同城热搜",
        "外网图片链接": ext_img,
        "内网图片链接": local_img,   # 本地文件路径
        "图片分辨率":  resolution,
    }


# ─── Excel 写出 ────────────────────────────────────────────────────────────────

_CITY_COLORS = [
    "DDEEFF", "EEF5FF", "DDF0E8", "EEF8F3",
    "FFF3DD", "FFF9EE", "F0DDEE", "F8EEF5",
    "DDEEEE", "EEF5F5",
]
_HEADER_FILL  = PatternFill("solid", fgColor="1B4F9A")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
_DATA_FONT    = Font(name="微软雅黑", size=9)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN         = Side(style="thin", color="CCCCCC")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COL_WIDTHS = {
    "平台": 8, "网站域名": 14, "内容分类": 12, "城市": 8,
    "标题": 28, "排名": 6, "热度": 10, "标签": 16,
    "对应内容url": 52, "对应内容": 48,
    "发布时间": 20, "采集时间": 20, "批次id": 24, "原始来源": 18,
    "外网图片链接": 55, "内网图片链接": 55, "图片分辨率": 12,
}
_CENTER_COLS = {"平台","网站域名","内容分类","城市","排名","热度",
                "发布时间","采集时间","批次id","原始来源","图片分辨率"}


def save_excel(records: list[dict], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "微博城市热搜"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 16)
    ws.row_dimensions[1].height = 22

    city_order = []
    for r in records:
        c = r.get("城市", "")
        if c not in city_order:
            city_order.append(c)
    city_color = {c: _CITY_COLORS[i % len(_CITY_COLORS)] for i, c in enumerate(city_order)}

    for row_idx, rec in enumerate(records, 2):
        fill = PatternFill("solid", fgColor=city_color.get(rec.get("城市",""), "FFFFFF"))
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col_name, ""))
            cell.fill      = fill
            cell.font      = _DATA_FONT
            cell.border    = _BORDER
            cell.alignment = _ALIGN_CENTER if col_name in _CENTER_COLS else _ALIGN_LEFT
        ws.row_dimensions[row_idx].height = 45

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(filename)
    print(f"  Excel 已保存：{filename}  ({len(records)} 行)")


# ─── 全量采集主流程 ────────────────────────────────────────────────────────────

def collect_all(cities: list[str] = None, delay: float = 5.5,
                img_dir: str = "", top_n: int = 0) -> tuple:
    if cities is None:
        cities = TARGET_CITIES

    batch_id     = datetime.now().strftime("%Y%m%d%H%M%S") + "_" + str(uuid.uuid4())[:6]
    collect_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_records  = []
    start_dt     = datetime.now()

    # 图片保存目录：images/{batch_id}/
    if not img_dir:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        img_dir    = os.path.join(script_dir, "images", batch_id)
    os.makedirs(img_dir, exist_ok=True)

    n_label = f"每城 TOP {top_n}" if top_n > 0 else "每城全量"
    print(f"\n批次 ID : {batch_id}")
    print(f"城市数量 : {len(cities)}   {n_label}")
    print(f"图片目录 : {img_dir}")
    print(f"开始时间 : {collect_time}")
    print("=" * 70)

    for city_idx, city in enumerate(cities, 1):
        print(f"\n[{city_idx:02d}/{len(cities)}] {city}", end="", flush=True)
        try:
            hot_list = get_city_hot_list(city, top_n=top_n)
        except CaptchaError:
            banned_name = _current_account()
            mark_cookie_banned(banned_name)          # 写入状态文件
            print(f"\n  [验证码] 账号「{banned_name}」已封禁并记录", end="")
            if rotate_cookie():
                print(f"，切换到「{_current_account()}」，重试...")
                try:
                    hot_list = get_city_hot_list(city, top_n=top_n)
                except CaptchaError:
                    mark_cookie_banned(_current_account())
                    print(f"\n  [验证码] 账号「{_current_account()}」也被封，所有账号已耗尽，中止采集")
                    break
                except Exception as e2:
                    print(f"  切换后仍失败: {e2}")
                    continue
            else:
                print(f"\n  Cookie 池已耗尽，中止采集！")
                print(f"  请在 Chrome 完成验证码或在 cookies/ 中补充新账号")
                break
        except Exception as e:
            print(f"  热榜失败: {e}")
            continue

        print(f"  共 {len(hot_list)} 条", end="", flush=True)
        city_records = []

        for idx, hot_item in enumerate(hot_list):
            mblog = fetch_top_post(hot_item["keyword"])
            rec   = build_record(city, hot_item, mblog, collect_time, batch_id, img_dir)
            city_records.append(rec)
            has_img = bool(rec["内网图片链接"])
            print(" o" if has_img else " -", end="", flush=True)

            # 每条之间随机等待 delay ± 1s
            if delay > 0:
                time.sleep(delay + random.uniform(-1, 1))
            # 每采 10 条额外休息 8-12s，模拟人工翻页停顿
            if (idx + 1) % 10 == 0:
                pause = random.uniform(8, 12)
                print(f" [休息{pause:.0f}s]", end="", flush=True)
                time.sleep(pause)

        all_records.extend(city_records)
        print(f"  OK {len(city_records)} 条入库")

        if city_idx < len(cities):
            time.sleep(0.5)

    elapsed = int((datetime.now() - start_dt).total_seconds())
    print(f"\n{'='*70}")
    print(f"采集完成  总计 {len(all_records)} 条  耗时 {elapsed} 秒")
    return all_records, batch_id


# ─── 主函数 ───────────────────────────────────────────────────────────────────

def main():
    records, batch_id = collect_all(TARGET_CITIES, delay=5.5, top_n=TOP_N)
    if not records:
        print("无数据")
        return

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx     = f"微博城市热搜_{ts}.xlsx"
    json_f   = f"微博城市热搜_{ts}.json"

    save_excel(records, xlsx)

    with open(json_f, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"  JSON 备份：{json_f}")


# ─── Cookie 验证工具 ──────────────────────────────────────────────────────────

def _test_cookie(cookie_str: str) -> str:
    """
    用一条 cookie 字符串请求轻量级接口，返回状态描述：
      "OK"       → 登录有效
      "CAPTCHA"  → 触发验证码
      "EXPIRED"  → 未登录 / cookie 过期
      "ERROR"    → 网络或解析异常
    """
    url = "https://m.weibo.cn/api/config"
    headers = {
        "User-Agent":       UA_MOBILE,
        "Accept":           "application/json",
        "Referer":          "https://m.weibo.cn/",
        "Cookie":           cookie_str,
        "X-Requested-With": "XMLHttpRequest",
    }
    try:
        s = _make_session()
        resp = s.get(url, headers=headers, timeout=8)
        ct = resp.headers.get("Content-Type", "")
        if ct.startswith("text/html"):
            return "CAPTCHA"
        data = resp.json()
        # data.data.login == True 表示已登录
        if data.get("data", {}).get("login"):
            return "OK"
        return "EXPIRED"
    except Exception as e:
        return f"ERROR({e})"


def verify_cookies():
    """
    验证所有 cookie（包括已封禁的），自动更新 _status.json。

    来源：
      1. Chrome 当前登录（rookiepy）
      2. cookies/ 目录下所有 .txt 文件

    结果：
      OK      → 解除封禁，下次采集可用
      CAPTCHA → 仍需在 Chrome 完成验证码再重试
      EXPIRED → Cookie 已失效，需重新登录并更新文件
    """
    print("\n" + "=" * 60)
    print("  微博 Cookie 验证工具")
    print("=" * 60)

    script_dir  = os.path.dirname(os.path.abspath(__file__))
    cookies_dir = os.path.join(script_dir, "cookies")
    status      = _load_status()
    changed     = False

    candidates: list[tuple[str, str]] = []   # [(账号名, cookie字符串)]

    # ① Chrome 账号
    if _ROOKIEPY_OK:
        try:
            all_c   = rookiepy.chrome(domains=["weibo.com", "weibo.cn", "passport.weibo.com"])
            sub_cn  = next((c["value"] for c in all_c if c["name"] == "SUB"  and c.get("domain") == ".weibo.cn"), "")
            subp_cn = next((c["value"] for c in all_c if c["name"] == "SUBP" and c.get("domain") == ".weibo.cn"), "")
            if sub_cn:
                candidates.append(("Chrome账号", f"SUB={sub_cn}; SUBP={subp_cn}"))
        except Exception as e:
            print(f"  [Chrome] 读取失败: {e}")

    # ② cookies/*.txt 文件
    if os.path.isdir(cookies_dir):
        for fname in sorted(os.listdir(cookies_dir)):
            if not fname.endswith(".txt"):
                continue
            try:
                cookie = open(os.path.join(cookies_dir, fname), encoding="utf-8").read().strip()
                if cookie:
                    candidates.append((fname[:-4], cookie))
            except Exception:
                pass

    if not candidates:
        print("  未找到任何 cookie（Chrome 未登录，cookies/ 目录也为空）")
        return

    print(f"\n  共找到 {len(candidates)} 个账号，逐一验证...\n")

    ok_count = 0
    for name, cookie_str in candidates:
        was_banned = (name in status and not status[name].get("ok", True))
        result = _test_cookie(cookie_str)

        is_chrome = (name == "Chrome账号")
        banned_info = f"  [封禁于 {status.get(name,{}).get('banned_at','')}]" if was_banned else ""
        print(f"  {name}{banned_info}")

        if result == "OK":
            status[name] = {"ok": True}
            changed = True
            ok_count += 1
            print(f"    → ✅ OK" + ("（已解封，下次采集可用）" if was_banned else ""))

        elif result == "CAPTCHA":
            if is_chrome:
                print(f"    → ⚠️  CAPTCHA")
                print(f"       处理方式：Chrome 打开 https://m.weibo.cn 完成验证码")
                print(f"                 完成后重新运行 --verify 即可自动恢复")
            else:
                print(f"    → ⚠️  CAPTCHA")
                print(f"       处理方式：Chrome【退出当前账号】→ 用「{name}」的账号密码重新登录微博")
                print(f"                 完成验证码后，重新导出 SUB/SUBP 更新到 cookies/{name}.txt")
                print(f"                 再运行 --verify 验证")

        elif result == "EXPIRED":
            status[name] = {"ok": False, "banned_at": status.get(name, {}).get("banned_at", ""), "reason": "expired"}
            changed = True
            if is_chrome:
                print(f"    → ❌ EXPIRED（Chrome 微博未登录，请先登录）")
            else:
                print(f"    → ❌ EXPIRED（Cookie 已彻底失效）")
                print(f"       处理方式：用「{name}」账号密码重新登录微博，导出新 SUB/SUBP 覆盖 cookies/{name}.txt")
        else:
            print(f"    → ❓ {result}")

        time.sleep(1)

    if changed:
        _save_status(status)
        print(f"\n  状态已更新 → {_status_path()}")

    print(f"\n  验证完成：{ok_count}/{len(candidates)} 个可用")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--verify":
        verify_cookies()
    else:
        main()
