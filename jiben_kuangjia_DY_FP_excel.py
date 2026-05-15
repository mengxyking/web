# ===================== 1. 导入依赖 =====================
import openpyxl
from openpyxl.styles import Font, Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter
import threading
import os
from copy import copy

# ===================== 2. 全局配置（仅需修改Excel路径） =====================
# 你的 Excel 文件路径
# 全局线程锁（多线程共享，保证同一时间仅1个线程写入）
excel_write_lock = threading.Lock()

# ===================== 3. 样式复制工具函数（保留原模板样式） =====================
def copy_row_style(src_row, dst_row, worksheet):
    """复制源行所有样式到目标行，保证新增数据和原模板样式完全一致"""
    for col_idx in range(1, worksheet.max_column + 1):
        src_cell = worksheet.cell(row=src_row, column=col_idx)
        dst_cell = worksheet.cell(row=dst_row, column=col_idx)
        dst_cell.font = copy(src_cell.font)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format

# ===================== 4. 核心写入函数（直接匹配你的字典Key，无额外映射） =====================
def write_order_to_excel(order_data,EXCEL_FILE_PATH):
    """
    线程安全的订单数据写入Excel
    :param order_data: 订单字典，Key和Excel表头完全一一对应
    :return: 写入成功返回True，失败返回False
    """
    # 校验空数据
    if not order_data:
        print(f"❌ 线程 {threading.current_thread().name} 订单数据为空，写入终止")
        return False

    # 获取线程锁（获取不到会自动阻塞，直到锁释放）
    excel_write_lock.acquire()
    print(f"✅ 线程 {threading.current_thread().name} 已获取锁，开始写入订单：{order_data.get('订单编号')}")

    try:
        # 打开/创建Excel文件
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            worksheet = workbook.active
        else:
            # 文件不存在时，自动用订单Key创建表头
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            # 写入表头（直接用订单的Key）
            for col_idx, header in enumerate(order_data.keys(), 1):
                worksheet.cell(row=1, column=col_idx, value=header)
            # 表头基础样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_align = Alignment(horizontal="center", vertical="center")
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        # 确定写入行号（最后一行的下一行）
        write_row = worksheet.max_row + 1

        # 写入订单数据（直接按Key匹配Excel表头列）
        for data_key, data_value in order_data.items():
            # 找到表头对应的列号
            col_idx = None
            for col in range(1, worksheet.max_column + 1):
                if worksheet.cell(row=1, column=col).value == data_key:
                    col_idx = col
                    break
            if col_idx:
                worksheet.cell(row=write_row, column=col_idx, value=data_value)

        # 复制上一行样式，保持模板统一
        if write_row > 2:
            copy_row_style(write_row - 1, write_row, worksheet)

        # 自动调整列宽，保证内容完整显示
        for col in range(1, worksheet.max_column + 1):
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                cell_value = str(worksheet.cell(row=row, column=col).value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            worksheet.column_dimensions[get_column_letter(col)].width = max_length + 2

        # 保存文件
        workbook.save(EXCEL_FILE_PATH)
        workbook.close()
        print(f"✅ 线程 {threading.current_thread().name} 写入完成，订单：{order_data.get('订单编号')}")
        return True

    except Exception as e:
        print(f"❌ 线程 {threading.current_thread().name} 写入异常：{str(e)}")
        return False

    finally:
        # 无论成功失败，必须释放锁，避免死锁
        excel_write_lock.release()
        print(f"🔓 线程 {threading.current_thread().name} 已释放锁")

# ===================== 5. 测试示例（用你给的订单数据） =====================
# if __name__ == "__main__":
#     # 你给的订单数据（直接用，无需修改）
#     EXCEL_FILE_PATH = r"E:\360MoveData\Users\Administrator\Desktop\dy产品信息表(1).xlsx"
#
#     test_order = {
#         '订单状态': '交易完成',
#         '店铺名称': '万千里运动专营店1',
#         '产品': '李宁速干短袖男款2026夏1.',
#         '原价': '￥64.001',
#         '实付款金额': '实付款￥50.001',
#         '数量文本': 'x11',
#         '数量': '1',
#         '快递单号': '中通快递 76923899261051',
#         '订单编号': '6952269417147995526',
#         '实付': '¥50'
#     }
#
#     # 单线程测试写入
#     write_order_to_excel(test_order,EXCEL_FILE_PATH)
#
#     # 多线程并发测试（取消注释即可测试）
#     # threads = []
#     # for i in range(3):
#     #     thread = threading.Thread(
#     #         target=write_order_to_excel,
#     #         args=(test_order,),
#     #         name=f"写入线程-{i+1}"
#     #     )
#     #     threads.append(thread)
#     #     thread.start()
#     # for thread in threads:
#     #     thread.join()
#
#     print("\n🎉 写入操作完成")